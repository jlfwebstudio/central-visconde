import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from tkinter import Tk, filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from baixar_relatorios_roteirizacao import baixar_relatorios_automaticamente


BASE_DIR = Path(__file__).resolve().parent.parent
ARQUIVO_REGRAS = BASE_DIR / "bases" / "regras_roteirizacao.xlsx"
PASTA_SAIDA = BASE_DIR / "outputs" / "roteirizacao"
PASTA_HISTORICO = PASTA_SAIDA / "historico"
ARQUIVO_ATUAL = PASTA_SAIDA / "roteirizacao_atual.xlsx"
ARQUIVO_RESOLUCOES_TEMPORARIAS = PASTA_SAIDA / "resolucoes_temporarias.json"
PASTA_DOWNLOADS_ROTEIRO = BASE_DIR / "downloads" / "roteirizacao"
PASTA_DOWNLOADS_MOBYAN = PASTA_DOWNLOADS_ROTEIRO / "mobyan"
PASTA_DOWNLOADS_OGEA = PASTA_DOWNLOADS_ROTEIRO / "ogea"

COR_AZUL = "1F4E78"
COR_CABECALHO = "FFFFFF"
COR_MOBYAN = "FF0000"
COR_OGEA = "0070C0"
COR_VERDE = "D9EAD3"
COR_AMARELO = "FFF2CC"
COR_VERMELHO = "F4CCCC"
COR_CINZA = "D9E2F3"
COR_BORDA = "D9E2F3"

TECNICOS_PADRAO = [
    "Adriano", "Emerson", "Fabio", "Fernando", "Gerson", "Geziel",
    "Howard", "Johnny", "Leonardo", "Lucindo", "Matheus", "Vladimir",
]


def normalizar_texto(valor):
    if pd.isna(valor):
        return ""

    texto = str(valor).strip()
    if texto == "" or texto.lower() in {"nan", "none", "null", "nat"}:
        return ""

    # Corrige caracteres estranhos encontrados na planilha original.
    texto = texto.replace("Μ", "A").replace("µ", "A")
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.upper()
    texto = re.sub(r"[^A-Z0-9]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def limpar_os(valor):
    texto = "" if pd.isna(valor) else str(valor).strip()
    if texto.endswith(".0"):
        texto = texto[:-2]
    return re.sub(r"\D", "", texto)


def status_aceito(valor):
    return "ENCAMINH" in normalizar_texto(valor)


def ler_csv_robusto(caminho):
    tentativas = [
        ("utf-8-sig", ";"),
        ("latin1", ";"),
        ("cp1252", ";"),
        ("utf-8-sig", ","),
        ("latin1", ","),
        ("cp1252", ","),
    ]

    ultimo_erro = None
    for encoding, sep in tentativas:
        try:
            df = pd.read_csv(
                caminho,
                dtype=str,
                encoding=encoding,
                sep=sep,
                keep_default_na=False,
                na_filter=False,
                low_memory=False,
            )
            if len(df.columns) > 1:
                return df
        except Exception as erro:
            ultimo_erro = erro

    raise RuntimeError(f"Não consegui ler o CSV {caminho}. Último erro: {ultimo_erro}")


def exigir_colunas(df, colunas, origem):
    faltantes = [c for c in colunas if c not in df.columns]
    if faltantes:
        raise ValueError(
            f"O relatório {origem} não possui as colunas esperadas: {', '.join(faltantes)}"
        )


def preparar_mobyan(caminho):
    df = ler_csv_robusto(caminho)
    exigir_colunas(
        df,
        ["Chamado", "Status", "Cidade", "Bairro", "Nome Cliente", "Serviço", "Técnico", "Data Limite", "Endereço"],
        "Mobyan",
    )

    resultado = pd.DataFrame({
        "Origem": "MOBYAN",
        "OS": df["Chamado"].map(limpar_os),
        "Técnico Atual": df["Técnico"].fillna("").astype(str).str.strip(),
        "Cidade": df["Cidade"].fillna("").astype(str).str.strip(),
        "Bairro / Distrito": df["Bairro"].fillna("").astype(str).str.strip(),
        "Cliente": df["Nome Cliente"].fillna("").astype(str).str.strip(),
        "Serviço": df["Serviço"].fillna("").astype(str).str.strip(),
        "Status": df["Status"].fillna("").astype(str).str.strip(),
        "Data Limite": df["Data Limite"].fillna("").astype(str).str.strip(),
        "Endereço": df["Endereço"].fillna("").astype(str).str.strip(),
    })

    resultado = resultado[resultado["Status"].map(status_aceito)].copy()
    resultado = resultado[resultado["OS"] != ""].copy()
    return resultado


def preparar_ogea(caminho):
    df = ler_csv_robusto(caminho)
    exigir_colunas(
        df,
        ["Código", "Status", "Cidade", "Distrito", "Cliente", "Serviço", "Técnico", "Data Limite", "Endereço"],
        "OGEA",
    )

    resultado = pd.DataFrame({
        "Origem": "OGEA",
        "OS": df["Código"].map(limpar_os),
        "Técnico Atual": df["Técnico"].fillna("").astype(str).str.strip(),
        "Cidade": df["Cidade"].fillna("").astype(str).str.strip(),
        "Bairro / Distrito": df["Distrito"].fillna("").astype(str).str.strip(),
        "Cliente": df["Cliente"].fillna("").astype(str).str.strip(),
        "Serviço": df["Serviço"].fillna("").astype(str).str.strip(),
        "Status": df["Status"].fillna("").astype(str).str.strip(),
        "Data Limite": df["Data Limite"].fillna("").astype(str).str.strip(),
        "Endereço": df["Endereço"].fillna("").astype(str).str.strip(),
    })

    resultado = resultado[resultado["Status"].map(status_aceito)].copy()
    resultado = resultado[resultado["OS"] != ""].copy()
    return resultado


def ler_regras():
    if not ARQUIVO_REGRAS.exists():
        raise FileNotFoundError(
            f"Base de regras não encontrada: {ARQUIVO_REGRAS}\n"
            "Copie regras_roteirizacao.xlsx para a pasta bases do projeto."
        )

    regras = pd.read_excel(
        ARQUIVO_REGRAS,
        sheet_name="Regras",
        dtype=str,
        keep_default_na=False,
        na_filter=False,
    )

    aliases = pd.read_excel(
        ARQUIVO_REGRAS,
        sheet_name="Aliases",
        dtype=str,
        keep_default_na=False,
        na_filter=False,
    )

    colunas_regras = [
        "Ativo", "Prioridade", "Técnico", "Tipo de Regra", "Cidade",
        "Bairro / Localidade Normalizada", "Origem", "Regra Original", "Observação",
    ]
    faltantes = [c for c in colunas_regras if c not in regras.columns]
    if faltantes:
        raise ValueError(f"Base de regras sem colunas obrigatórias: {', '.join(faltantes)}")

    regras = regras[
        regras["Ativo"].map(normalizar_texto).isin(["SIM", "S", "YES", "Y"])
    ].copy()

    regras["_CIDADE"] = regras["Cidade"].map(normalizar_texto)
    regras["_BAIRRO"] = regras["Bairro / Localidade Normalizada"].map(normalizar_texto)
    regras["_TIPO"] = regras["Tipo de Regra"].map(normalizar_texto)
    regras["_ORIGEM"] = regras["Origem"].map(normalizar_texto)
    regras["_PRIORIDADE"] = pd.to_numeric(regras["Prioridade"], errors="coerce").fillna(999999)

    for coluna in ["Ativo", "Cidade", "Nome recebido", "Nome considerado", "Origem"]:
        if coluna not in aliases.columns:
            aliases[coluna] = "AMBOS" if coluna == "Origem" else ""

    aliases = aliases[
        aliases["Ativo"].map(normalizar_texto).isin(["SIM", "S", "YES", "Y"])
    ].copy()
    aliases["_CIDADE"] = aliases["Cidade"].map(normalizar_texto)
    aliases["_RECEBIDO"] = aliases["Nome recebido"].map(normalizar_texto)
    aliases["_CONSIDERADO"] = aliases["Nome considerado"].map(normalizar_texto)
    aliases["_ORIGEM"] = aliases["Origem"].map(normalizar_texto).replace("", "AMBOS")

    try:
        tecnicos_df = pd.read_excel(
            ARQUIVO_REGRAS,
            sheet_name="Técnicos",
            dtype=str,
            keep_default_na=False,
            na_filter=False,
        )
        tecnicos = [
            str(v).strip()
            for v in tecnicos_df.get("Técnico", pd.Series(dtype=str)).tolist()
            if str(v).strip()
        ]
    except Exception:
        tecnicos = []

    if not tecnicos:
        tecnicos = sorted(regras["Técnico"].astype(str).str.strip().replace("", pd.NA).dropna().unique())

    return regras, aliases, tecnicos


def aplicar_alias(cidade_norm, bairro_norm, origem, aliases):
    correspondencias = aliases[
        (aliases["_CIDADE"] == cidade_norm)
        & (aliases["_RECEBIDO"] == bairro_norm)
        & (aliases["_ORIGEM"].isin(["AMBOS", origem]))
    ].copy()

    if correspondencias.empty:
        return bairro_norm, ""

    # Um alias específico da origem vence um alias genérico.
    correspondencias["_ESPECIFICO"] = (correspondencias["_ORIGEM"] == origem).astype(int)
    linha = correspondencias.sort_values("_ESPECIFICO").iloc[-1]
    return linha["_CONSIDERADO"], str(linha.get("Nome considerado", "")).strip()


def ler_resolucoes_temporarias():
    if not ARQUIVO_RESOLUCOES_TEMPORARIAS.exists():
        return {}
    try:
        dados = json.loads(ARQUIVO_RESOLUCOES_TEMPORARIAS.read_text(encoding="utf-8"))
    except Exception:
        return {}
    mapa = {}
    for item in dados if isinstance(dados, list) else []:
        chave = (
            normalizar_texto(item.get("origem", "")),
            normalizar_texto(item.get("cidade", "")),
            normalizar_texto(item.get("bairro", "")),
        )
        tecnico = str(item.get("tecnico", "")).strip()
        if all(chave) and tecnico:
            mapa[chave] = tecnico
    return mapa


def roteirizar_linha(linha, regras, aliases, resolucoes_temporarias=None):
    origem = normalizar_texto(linha.get("Origem", ""))
    cidade_norm = normalizar_texto(linha.get("Cidade", ""))
    bairro_recebido_norm = normalizar_texto(linha.get("Bairro / Distrito", ""))

    temporario = (resolucoes_temporarias or {}).get((origem, cidade_norm, bairro_recebido_norm))
    if temporario:
        return pd.Series({
            "Resultado": "ROTEIRIZADA",
            "Técnico Roteirizado": temporario,
            "Regra Aplicada": "RESOLUÇÃO TEMPORÁRIA",
            "Alias Aplicado": "",
            "Candidatos": temporario,
        })

    bairro_norm, alias_aplicado = aplicar_alias(cidade_norm, bairro_recebido_norm, origem, aliases)

    regras_cidade = regras[
        (regras["_CIDADE"] == cidade_norm)
        & (regras["_ORIGEM"].isin(["AMBOS", origem]))
    ]

    regras_inteiras = regras_cidade[
        regras_cidade["_TIPO"].str.contains("CIDADE INTEIRA", na=False)
    ]

    if not regras_inteiras.empty:
        menor_prioridade = regras_inteiras["_PRIORIDADE"].min()
        regras_escolhidas = regras_inteiras[regras_inteiras["_PRIORIDADE"] == menor_prioridade]
        candidatos = sorted(set(regras_escolhidas["Técnico"].astype(str).str.strip()))
        regra_aplicada = "CIDADE INTEIRA"
    else:
        regras_bairro = regras_cidade[
            regras_cidade["_TIPO"].str.contains("BAIRRO", na=False)
            & (regras_cidade["_BAIRRO"] == bairro_norm)
        ]
        if not regras_bairro.empty:
            menor_prioridade = regras_bairro["_PRIORIDADE"].min()
            regras_bairro = regras_bairro[regras_bairro["_PRIORIDADE"] == menor_prioridade]
        candidatos = sorted(set(regras_bairro["Técnico"].astype(str).str.strip()))
        regra_aplicada = "CIDADE + BAIRRO"

    candidatos = [c for c in candidatos if c]

    # Regra validada: Leonardo não trabalha com Mobyan.
    if origem == "MOBYAN" and "Leonardo" in candidatos:
        candidatos = [c for c in candidatos if c != "Leonardo"]
        if "Emerson" not in candidatos:
            candidatos.append("Emerson")

    # Nas áreas compartilhadas Emerson/Leonardo, o OGEA fica com Leonardo.
    if origem == "OGEA" and "Leonardo" in candidatos:
        candidatos = ["Leonardo"]

    candidatos = sorted(set(candidatos))

    if len(candidatos) == 1:
        return pd.Series({
            "Resultado": "ROTEIRIZADA",
            "Técnico Roteirizado": candidatos[0],
            "Regra Aplicada": regra_aplicada,
            "Alias Aplicado": alias_aplicado,
            "Candidatos": candidatos[0],
        })

    if len(candidatos) > 1:
        texto = " / ".join(candidatos)
        return pd.Series({
            "Resultado": "CONFLITO",
            "Técnico Roteirizado": "",
            "Regra Aplicada": regra_aplicada,
            "Alias Aplicado": alias_aplicado,
            "Candidatos": texto,
        })

    motivo = "Cidade ou bairro sem regra cadastrada"
    if cidade_norm == "":
        motivo = "Cidade não informada"
    elif bairro_norm == "" and regras_inteiras.empty:
        motivo = "Bairro/Distrito não informado"

    return pd.Series({
        "Resultado": "SEM ROTA",
        "Técnico Roteirizado": "",
        "Regra Aplicada": motivo,
        "Alias Aplicado": alias_aplicado,
        "Candidatos": "",
    })


def executar_motor(caminho_mobyan, caminho_ogea, usar_resolucoes_temporarias=False):
    print("Lendo relatório Mobyan...")
    mobyan = preparar_mobyan(caminho_mobyan)
    print(f"Mobyan após filtro de status: {len(mobyan)} OS(s).")

    print("Lendo relatório OGEA...")
    ogea = preparar_ogea(caminho_ogea)
    print(f"OGEA após filtro de status: {len(ogea)} OS(s).")

    print("Carregando regras de roteirização...")
    regras, aliases, tecnicos = ler_regras()

    base = pd.concat([mobyan, ogea], ignore_index=True)
    base = base.drop_duplicates(subset=["Origem", "OS"], keep="last").reset_index(drop=True)

    resolucoes_temporarias = ler_resolucoes_temporarias() if usar_resolucoes_temporarias else {}
    if resolucoes_temporarias:
        print(f"Aplicando {len(resolucoes_temporarias)} resolução(ões) temporária(s)...")
    print("Aplicando regras de cidade e bairro...")
    resultado = base.apply(
        lambda linha: roteirizar_linha(linha, regras, aliases, resolucoes_temporarias),
        axis=1,
    )

    df = pd.concat([base, resultado], axis=1)

    colunas = [
        "Origem", "OS", "Resultado", "Técnico Roteirizado", "Técnico Atual",
        "Cidade", "Bairro / Distrito", "Cliente", "Serviço", "Status",
        "Data Limite", "Endereço", "Regra Aplicada", "Alias Aplicado", "Candidatos",
    ]
    df = df[colunas].copy()

    df["_ORIGEM_ORDEM"] = df["Origem"].map({"MOBYAN": 1, "OGEA": 2}).fillna(9)
    df["_OS_NUM"] = pd.to_numeric(df["OS"], errors="coerce")
    df = df.sort_values(
        by=["Resultado", "Técnico Roteirizado", "_ORIGEM_ORDEM", "_OS_NUM"],
        ascending=[True, True, True, False],
        na_position="last",
    ).drop(columns=["_ORIGEM_ORDEM", "_OS_NUM"])

    return df.reset_index(drop=True), regras, aliases, tecnicos


def montar_resumo(df, tecnicos):
    roteirizadas = df[df["Resultado"] == "ROTEIRIZADA"].copy()

    indicadores = pd.DataFrame([
        ["Total de OSs roteirizadas", len(roteirizadas)],
        ["Mobyan", len(roteirizadas[roteirizadas["Origem"] == "MOBYAN"])],
        ["OGEA", len(roteirizadas[roteirizadas["Origem"] == "OGEA"])],
        ["Conflitos", len(df[df["Resultado"] == "CONFLITO"])],
        ["Sem rota", len(df[df["Resultado"] == "SEM ROTA"])],
    ], columns=["Indicador", "Quantidade"])

    linhas_tecnicos = []
    for tecnico in tecnicos:
        recorte = roteirizadas[roteirizadas["Técnico Roteirizado"] == tecnico]
        linhas_tecnicos.append([
            tecnico,
            len(recorte[recorte["Origem"] == "MOBYAN"]),
            len(recorte[recorte["Origem"] == "OGEA"]),
            len(recorte),
        ])

    resumo_tecnicos = pd.DataFrame(
        linhas_tecnicos,
        columns=["Técnico", "Mobyan", "OGEA", "Total"],
    )
    return indicadores, resumo_tecnicos


def criar_listas_pdf(df, tecnicos):
    roteirizadas = df[df["Resultado"] == "ROTEIRIZADA"].copy()
    linhas = []

    for tecnico in tecnicos:
        base_tecnico = roteirizadas[roteirizadas["Técnico Roteirizado"] == tecnico]

        mobyan = sorted(
            {limpar_os(v) for v in base_tecnico[base_tecnico["Origem"] == "MOBYAN"]["OS"] if limpar_os(v)},
            key=lambda x: int(x),
            reverse=True,
        )
        ogea = sorted(
            {limpar_os(v) for v in base_tecnico[base_tecnico["Origem"] == "OGEA"]["OS"] if limpar_os(v)},
            key=lambda x: int(x),
            reverse=True,
        )

        linhas.append({
            "Técnico": tecnico,
            "Qtd. Mobyan": len(mobyan),
            "OSs Mobyan": ",".join(mobyan),
            "Qtd. OGEA": len(ogea),
            "OSs OGEA": ",".join(ogea),
            "Total": len(mobyan) + len(ogea),
        })

    return pd.DataFrame(linhas)


def gerar_planilha(df, regras, aliases, tecnicos):
    PASTA_SAIDA.mkdir(parents=True, exist_ok=True)

    indicadores, resumo_tecnicos = montar_resumo(df, tecnicos)
    conflitos = df[df["Resultado"] == "CONFLITO"].copy()
    sem_rota = df[df["Resultado"] == "SEM ROTA"].copy()
    listas_pdf = criar_listas_pdf(df, tecnicos)

    caminho_temp = PASTA_SAIDA / "roteirizacao_em_processamento.xlsx"

    with pd.ExcelWriter(caminho_temp, engine="openpyxl") as writer:
        indicadores.to_excel(writer, index=False, sheet_name="Resumo", startrow=3, startcol=0)
        resumo_tecnicos.to_excel(writer, index=False, sheet_name="Resumo", startrow=3, startcol=3)
        df.to_excel(writer, index=False, sheet_name="Roteiro Geral")
        conflitos.to_excel(writer, index=False, sheet_name="Conflitos")
        sem_rota.to_excel(writer, index=False, sheet_name="Sem Rota")
        listas_pdf.to_excel(writer, index=False, sheet_name="Listas PDF")

        regras.drop(columns=[c for c in regras.columns if c.startswith("_")], errors="ignore").to_excel(
            writer, index=False, sheet_name="Regras"
        )
        aliases.drop(columns=[c for c in aliases.columns if c.startswith("_")], errors="ignore").to_excel(
            writer, index=False, sheet_name="Aliases"
        )

        # Roteiro por Técnico: somente números, Mobyan primeiro e OGEA abaixo.
        roteiro = df[df["Resultado"] == "ROTEIRIZADA"].copy()
        max_linhas = 0
        dados_colunas = {}
        origens_colunas = {}

        for tecnico in tecnicos:
            recorte = roteiro[roteiro["Técnico Roteirizado"] == tecnico]
            mobyan = sorted(
                {int(limpar_os(v)) for v in recorte[recorte["Origem"] == "MOBYAN"]["OS"] if limpar_os(v)},
                reverse=True,
            )
            ogea = sorted(
                {int(limpar_os(v)) for v in recorte[recorte["Origem"] == "OGEA"]["OS"] if limpar_os(v)},
                reverse=True,
            )
            dados_colunas[tecnico] = mobyan + ogea
            origens_colunas[tecnico] = (["MOBYAN"] * len(mobyan)) + (["OGEA"] * len(ogea))
            max_linhas = max(max_linhas, len(dados_colunas[tecnico]))

        matriz = {}
        for tecnico in tecnicos:
            valores = dados_colunas[tecnico]
            matriz[tecnico] = valores + [None] * (max_linhas - len(valores))

        pd.DataFrame(matriz).to_excel(writer, index=False, sheet_name="Roteiro por Técnico")

    formatar_planilha(caminho_temp, df, tecnicos, origens_colunas)

    if ARQUIVO_ATUAL.exists():
        try:
            ARQUIVO_ATUAL.unlink()
        except PermissionError as erro:
            raise PermissionError(
                "A planilha roteirizacao_atual.xlsx está aberta no Excel. "
                "Feche essa planilha e clique novamente em salvar/gerar. "
                "A regra cadastrada já foi preservada."
            ) from erro
        except OSError as erro:
            raise OSError(
                f"Não consegui substituir a roteirização atual: {erro}"
            ) from erro

    try:
        caminho_temp.replace(ARQUIVO_ATUAL)
    except PermissionError as erro:
        raise PermissionError(
            "A planilha roteirizacao_atual.xlsx está em uso. "
            "Feche o Excel e clique novamente em salvar/gerar. "
            "A regra cadastrada já foi preservada."
        ) from erro

    # O usuário optou por não manter histórico.
    # Remove arquivos antigos caso a pasta de histórico já exista.
    if PASTA_HISTORICO.exists():
        for item in PASTA_HISTORICO.iterdir():
            try:
                if item.is_dir():
                    shutil.rmtree(item)
                else:
                    item.unlink()
            except Exception:
                pass

    return ARQUIVO_ATUAL


def formatar_cabecalho(ws):
    preenchimento = PatternFill("solid", fgColor=COR_AZUL)
    fonte = Font(color=COR_CABECALHO, bold=True)
    borda = Border(
        left=Side(style="thin", color=COR_BORDA),
        right=Side(style="thin", color=COR_BORDA),
        top=Side(style="thin", color=COR_BORDA),
        bottom=Side(style="thin", color=COR_BORDA),
    )

    for cell in ws[1]:
        cell.fill = preenchimento
        cell.font = fonte
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24


def formatar_planilha(caminho, df, tecnicos, origens_colunas):
    wb = load_workbook(caminho)

    # Resumo
    ws = wb["Resumo"]
    ws["A1"] = "ROTEIRIZAÇÃO DO DIA — MOBYAN + OGEA"
    ws.merge_cells("A1:G1")
    ws["A1"].fill = PatternFill("solid", fgColor=COR_AZUL)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"] = "Somente OSs com status que contenha ENCAMINH. OSs EM CAMPO ficam fora."
    ws.merge_cells("A2:G2")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.sheet_view.showGridLines = False

    for row in [4]:
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.fill = PatternFill("solid", fgColor=COR_AZUL)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    # Roteiro Geral, conflitos, sem rota, listas e bases.
    for nome in ["Roteiro Geral", "Conflitos", "Sem Rota", "Listas PDF", "Regras", "Aliases"]:
        ws = wb[nome]
        formatar_cabecalho(ws)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=False)

    larguras_geral = {
        "A": 12, "B": 13, "C": 16, "D": 22, "E": 22, "F": 20,
        "G": 28, "H": 38, "I": 28, "J": 22, "K": 20, "L": 48,
        "M": 24, "N": 22, "O": 28,
    }
    for nome in ["Roteiro Geral", "Conflitos", "Sem Rota"]:
        ws = wb[nome]
        for letra, largura in larguras_geral.items():
            ws.column_dimensions[letra].width = largura

    ws = wb["Listas PDF"]
    for letra, largura in {"A":20,"B":14,"C":90,"D":14,"E":90,"F":12}.items():
        ws.column_dimensions[letra].width = largura

    # Roteiro por Técnico.
    ws = wb["Roteiro por Técnico"]
    formatar_cabecalho(ws)
    ws.auto_filter.ref = None
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A2"

    max_row = ws.max_row
    max_col = ws.max_column

    # Remove qualquer preenchimento e deixa só a cor da fonte por origem.
    for col_idx, tecnico in enumerate(tecnicos, start=1):
        letra = get_column_letter(col_idx)
        ws.column_dimensions[letra].width = 15
        origens = origens_colunas.get(tecnico, [])

        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.number_format = "0"

            origem = origens[row_idx - 2] if row_idx - 2 < len(origens) else ""
            if origem == "MOBYAN":
                cell.font = Font(color=COR_MOBYAN, size=11)
            elif origem == "OGEA":
                cell.font = Font(color=COR_OGEA, size=11)
            else:
                cell.font = Font(color="000000", size=11)

    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 20

    # Coloração de resultado nas abas detalhadas.
    for nome in ["Roteiro Geral", "Conflitos", "Sem Rota"]:
        ws = wb[nome]
        headers = [cell.value for cell in ws[1]]
        if "Resultado" not in headers:
            continue
        col_resultado = headers.index("Resultado") + 1
        for row_idx in range(2, ws.max_row + 1):
            resultado = str(ws.cell(row=row_idx, column=col_resultado).value or "")
            if resultado == "ROTEIRIZADA":
                fill = PatternFill("solid", fgColor=COR_VERDE)
            elif resultado == "CONFLITO":
                fill = PatternFill("solid", fgColor=COR_AMARELO)
            else:
                fill = PatternFill("solid", fgColor=COR_VERMELHO)
            ws.cell(row=row_idx, column=col_resultado).fill = fill
            ws.cell(row=row_idx, column=col_resultado).font = Font(bold=True)

    ordem = [
        "Resumo", "Roteiro Geral", "Roteiro por Técnico", "Conflitos",
        "Sem Rota", "Listas PDF", "Regras", "Aliases",
    ]
    wb._sheets = [wb[n] for n in ordem if n in wb.sheetnames]
    wb.active = 2 if "Roteiro por Técnico" in wb.sheetnames else 0
    wb.save(caminho)


def arquivo_mais_recente(pasta):
    pasta = Path(pasta)
    candidatos = [
        p for p in pasta.rglob("*")
        if p.is_file() and p.suffix.lower() in {".csv", ".txt", ".xlsx", ".xls"}
    ] if pasta.exists() else []
    if not candidatos:
        return None
    return max(candidatos, key=lambda p: p.stat().st_mtime)


def localizar_relatorios_atuais():
    mobyan = arquivo_mais_recente(PASTA_DOWNLOADS_MOBYAN)
    ogea = arquivo_mais_recente(PASTA_DOWNLOADS_OGEA)
    if not mobyan or not ogea:
        raise FileNotFoundError(
            "Não encontrei os dois relatórios já baixados. Gere uma roteirização normal primeiro."
        )
    return mobyan, ogea


def selecionar_arquivos():
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    try:
        caminho_mobyan = filedialog.askopenfilename(
            title="Selecione o relatório da MOBYAN",
            filetypes=[("Relatório CSV", "*.csv"), ("Todos os arquivos", "*.*")],
        )
        if not caminho_mobyan:
            return None, None

        caminho_ogea = filedialog.askopenfilename(
            title="Selecione o relatório do OGEA",
            filetypes=[("Relatório CSV", "*.csv"), ("Todos os arquivos", "*.*")],
        )
        if not caminho_ogea:
            return None, None

        return Path(caminho_mobyan), Path(caminho_ogea)
    finally:
        root.destroy()


def abrir_arquivo(caminho):
    try:
        if sys.platform == "darwin":
            subprocess.run(["open", str(caminho)], check=False)
        elif os.name == "nt":
            os.startfile(str(caminho))
        else:
            subprocess.run(["xdg-open", str(caminho)], check=False)
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser(description="Gera roteirização Mobyan + OGEA")
    parser.add_argument("--mobyan", help="Caminho do relatório CSV da Mobyan")
    parser.add_argument("--ogea", help="Caminho do relatório CSV do OGEA")
    parser.add_argument(
        "--manual",
        action="store_true",
        help="Seleciona manualmente os relatórios em vez de baixá-los",
    )
    parser.add_argument("--sem-abrir", action="store_true", help="Não abre o Excel ao terminar")
    parser.add_argument("--reprocessar-atual", action="store_true", help="Reusa os últimos relatórios baixados")
    parser.add_argument("--usar-resolucoes-temporarias", action="store_true", help="Aplica decisões válidas somente para este roteiro")
    args = parser.parse_args()

    if args.mobyan and args.ogea:
        caminho_mobyan = Path(args.mobyan).expanduser().resolve()
        caminho_ogea = Path(args.ogea).expanduser().resolve()
    elif args.reprocessar_atual:
        caminho_mobyan, caminho_ogea = localizar_relatorios_atuais()
    elif args.manual:
        caminho_mobyan, caminho_ogea = selecionar_arquivos()
        if not caminho_mobyan or not caminho_ogea:
            print("Seleção cancelada. Nenhum arquivo foi alterado.")
            return 0
    else:
        caminho_mobyan, caminho_ogea = baixar_relatorios_automaticamente()

    print(f"Relatório Mobyan: {caminho_mobyan}")
    print(f"Relatório OGEA: {caminho_ogea}")

    df, regras, aliases, tecnicos = executar_motor(
        caminho_mobyan, caminho_ogea,
        usar_resolucoes_temporarias=args.usar_resolucoes_temporarias,
    )
    atual = gerar_planilha(df, regras, aliases, tecnicos)

    if args.usar_resolucoes_temporarias and ARQUIVO_RESOLUCOES_TEMPORARIAS.exists():
        try:
            ARQUIVO_RESOLUCOES_TEMPORARIAS.unlink()
            print("Resoluções temporárias consumidas e removidas.")
        except Exception:
            pass

    roteirizadas = len(df[df["Resultado"] == "ROTEIRIZADA"])
    conflitos = len(df[df["Resultado"] == "CONFLITO"])
    sem_rota = len(df[df["Resultado"] == "SEM ROTA"])

    print("")
    print("=" * 70)
    print(f"Roteirizadas: {roteirizadas}")
    print(f"Conflitos: {conflitos}")
    print(f"Sem rota: {sem_rota}")
    print(f"Planilha atual: {atual}")
    print("=" * 70)

    if not args.sem_abrir:
        abrir_arquivo(atual)

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as erro:
        print("")
        print("ERRO AO GERAR ROTEIRIZAÇÃO")
        print(str(erro))
        raise
