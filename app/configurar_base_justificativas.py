from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
ARQUIVO_BASE = BASE_DIR / "bases" / "base_justificativas.xlsx"

PRESTADORES = [
    "RS-SMART",
    "RS-SMART - CAXIAS DO SUL",
    "RS-SMART - CAPAO DA CANOA",
    "RS-SMART - PASSO FUNDO",
    "RS-SMART - PELOTAS",
    "RS-SMART - SANTA MARIA",
    "RS-SMART - SANTA CRUZ DO SUL",
    "RS-SMART - SANTA VITORIA DO",
    "RS-SMART - SANTANA DO",
    "RS-SMART - SANTO ANGELO",
    "RS-SMART - URUGUAIANA",
    "RS-SMART - VALE DOS SINOS",
    "RS-SMART TAPES",
]

MOTIVOS = [
    "Falta de Equipamento",
    "Falta de Bobina",
    "Falta de Rota",
    "Aguardando Estoque",
    "Instabilidade Sistema Mobyan",
    "Instabilidade Sistema Contratante",
    "Erro de Roteirização",
    "Técnico Indisponível",
    "Condição Climática",
    "Cliente Ausente",
    "OS Recebida em Atraso",
    "Outro",
]

ORIGENS = [
    "Base",
    "Técnico",
    "Interno",
    "Sistema",
]

VALIDADO = [
    "Sim",
    "Não",
]

CABECALHOS = [
    "Chamado",
    "Prestador",
    "Motivo",
    "Observação",
    "Data Registro",
    "Origem",
    "Validado",
]


def criar_ou_abrir_workbook():
    ARQUIVO_BASE.parent.mkdir(parents=True, exist_ok=True)

    if ARQUIVO_BASE.exists():
        wb = load_workbook(ARQUIVO_BASE)
    else:
        wb = Workbook()

    return wb


def garantir_aba(wb, nome):
    if nome in wb.sheetnames:
        return wb[nome]

    return wb.create_sheet(nome)


def configurar_aba_justificativas(wb):
    if "Justificativas" in wb.sheetnames:
        ws = wb["Justificativas"]
    else:
        # Se for workbook novo, usa a planilha ativa
        ws = wb.active
        ws.title = "Justificativas"

    # Garante cabeçalhos sem apagar os dados existentes
    for col, cabecalho in enumerate(CABECALHOS, start=1):
        ws.cell(row=1, column=col).value = cabecalho

    # Formatação
    preenchimento_cabecalho = PatternFill("solid", fgColor="1F4E78")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)
    borda_fina = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for cell in ws[1]:
        cell.fill = preenchimento_cabecalho
        cell.font = fonte_cabecalho
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda_fina

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G1000"

    larguras = {
        "A": 14,  # Chamado
        "B": 30,  # Prestador
        "C": 34,  # Motivo
        "D": 46,  # Observação
        "E": 16,  # Data Registro
        "F": 16,  # Origem
        "G": 12,  # Validado
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    ws.row_dimensions[1].height = 24

    for row in range(2, 1001):
        ws.row_dimensions[row].height = 18

        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = borda_fina
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Formato de data
        ws.cell(row=row, column=5).number_format = "dd/mm/yyyy"

    # Validações
    dv_prestador = DataValidation(
        type="list",
        formula1="=Listas!$A$2:$A$100",
        allow_blank=True
    )

    dv_motivo = DataValidation(
        type="list",
        formula1="=Listas!$B$2:$B$100",
        allow_blank=True
    )

    dv_origem = DataValidation(
        type="list",
        formula1="=Listas!$C$2:$C$100",
        allow_blank=True
    )

    dv_validado = DataValidation(
        type="list",
        formula1="=Listas!$D$2:$D$20",
        allow_blank=True
    )

    dv_data = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2020,1,1)",
        formula2="DATE(2035,12,31)",
        allow_blank=True
    )

    dv_data.error = "Digite uma data válida."
    dv_data.errorTitle = "Data inválida"

    ws.add_data_validation(dv_prestador)
    ws.add_data_validation(dv_motivo)
    ws.add_data_validation(dv_origem)
    ws.add_data_validation(dv_validado)
    ws.add_data_validation(dv_data)

    dv_prestador.add("B2:B1000")
    dv_motivo.add("C2:C1000")
    dv_data.add("E2:E1000")
    dv_origem.add("F2:F1000")
    dv_validado.add("G2:G1000")

    return ws


def configurar_aba_listas(wb):
    ws = garantir_aba(wb, "Listas")

    ws["A1"] = "Prestadores"
    ws["B1"] = "Motivos"
    ws["C1"] = "Origem"
    ws["D1"] = "Validado"

    listas = {
        "A": PRESTADORES,
        "B": MOTIVOS,
        "C": ORIGENS,
        "D": VALIDADO,
    }

    preenchimento_cabecalho = PatternFill("solid", fgColor="70AD47")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)

    for col in ["A", "B", "C", "D"]:
        ws[f"{col}1"].fill = preenchimento_cabecalho
        ws[f"{col}1"].font = fonte_cabecalho
        ws[f"{col}1"].alignment = Alignment(horizontal="center")

        # Limpa a lista antiga
        for row in range(2, 101):
            ws[f"{col}{row}"] = None

        # Escreve a lista atual
        for idx, valor in enumerate(listas[col], start=2):
            ws[f"{col}{idx}"] = valor

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14

    return ws


def remover_abas_vazias_antigas(wb):
    # Se existir uma aba padrão vazia chamada Sheet, remove.
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        ws = wb["Sheet"]

        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            wb.remove(ws)


def configurar_base():
    wb = criar_ou_abrir_workbook()

    configurar_aba_justificativas(wb)
    configurar_aba_listas(wb)
    remover_abas_vazias_antigas(wb)

    wb.save(ARQUIVO_BASE)

    print("Base de justificativas configurada com sucesso!")
    print(f"Arquivo: {ARQUIVO_BASE}")
    print("")
    print("Agora a aba 'Justificativas' possui listas suspensas em:")
    print("- Prestador")
    print("- Motivo")
    print("- Origem")
    print("- Validado")
    print("")
    print("A coluna Data Registro aceita datas no formato dd/mm/aaaa.")


if __name__ == "__main__":
    configurar_base()