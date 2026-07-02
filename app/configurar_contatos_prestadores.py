from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = Path(__file__).resolve().parent.parent
ARQUIVO_CONTATOS = BASE_DIR / "bases" / "contatos_prestadores.xlsx"

PRESTADORES = [
    "RS-SMART",
    "RS-SMART - CAXIAS DO SUL",
    "RS-SMART - CAPAO DA CANOA",
    "RS-SMART - PASSO FUNDO",
    "RS-SMART - PELOTAS",
    "RS-SMART - SANTA MARIA",
    "RS-SMART - SANTA CRUZ DO SUL",
    "RS-SMART - SANTA VITORIA DO",
    "RS-SMART - SANTANA DO",
    "RS-SMART - SANTO ANGELO",
    "RS-SMART - URUGUAIANA",
    "RS-SMART - VALE DOS SINOS",
    "RS-SMART TAPES",
]

CABECALHOS = [
    "Prestador",
    "Responsável",
    "WhatsApp",
    "Enviar",
    "Observação",
]

ENVIAR = [
    "Sim",
    "Não",
]


def criar_ou_abrir_workbook():
    ARQUIVO_CONTATOS.parent.mkdir(parents=True, exist_ok=True)

    if ARQUIVO_CONTATOS.exists():
        return load_workbook(ARQUIVO_CONTATOS)

    return Workbook()


def garantir_aba(wb, nome):
    if nome in wb.sheetnames:
        return wb[nome]

    return wb.create_sheet(nome)


def configurar_aba_contatos(wb):
    if "Contatos" in wb.sheetnames:
        ws = wb["Contatos"]
    else:
        ws = wb.active
        ws.title = "Contatos"

    for col, cabecalho in enumerate(CABECALHOS, start=1):
        ws.cell(row=1, column=col).value = cabecalho

    preenchimento_cabecalho = PatternFill("solid", fgColor="1F4E78")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)

    borda_fina = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for cell in ws[1]:
        cell.fill = preenchimento_cabecalho
        cell.font = fonte_cabecalho
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda_fina

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:E200"

    larguras = {
        "A": 34,
        "B": 26,
        "C": 22,
        "D": 12,
        "E": 42,
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    ws.row_dimensions[1].height = 24

    for row in range(2, 201):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = borda_fina
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    dv_prestador = DataValidation(
        type="list",
        formula1="=Listas!$A$2:$A$100",
        allow_blank=True
    )

    dv_enviar = DataValidation(
        type="list",
        formula1="=Listas!$B$2:$B$20",
        allow_blank=True
    )

    ws.add_data_validation(dv_prestador)
    ws.add_data_validation(dv_enviar)

    dv_prestador.add("A2:A200")
    dv_enviar.add("D2:D200")

    return ws


def configurar_aba_listas(wb):
    ws = garantir_aba(wb, "Listas")

    ws["A1"] = "Prestadores"
    ws["B1"] = "Enviar"

    preenchimento_cabecalho = PatternFill("solid", fgColor="70AD47")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)

    for col in ["A", "B"]:
        ws[f"{col}1"].fill = preenchimento_cabecalho
        ws[f"{col}1"].font = fonte_cabecalho
        ws[f"{col}1"].alignment = Alignment(horizontal="center")

        for row in range(2, 101):
            ws[f"{col}{row}"] = None

    for idx, prestador in enumerate(PRESTADORES, start=2):
        ws[f"A{idx}"] = prestador

    for idx, valor in enumerate(ENVIAR, start=2):
        ws[f"B{idx}"] = valor

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14

    return ws


def remover_abas_vazias_antigas(wb):
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        ws = wb["Sheet"]

        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            wb.remove(ws)


def configurar_contatos():
    wb = criar_ou_abrir_workbook()

    configurar_aba_contatos(wb)
    configurar_aba_listas(wb)
    remover_abas_vazias_antigas(wb)

    wb.save(ARQUIVO_CONTATOS)

    print("Base de contatos dos prestadores configurada com sucesso!")
    print(f"Arquivo: {ARQUIVO_CONTATOS}")
    print("")
    print("Aba Contatos:")
    print("- Prestador com lista suspensa")
    print("- Responsável livre")
    print("- WhatsApp livre")
    print("- Enviar com Sim/Não")
    print("- Observação livre")


if __name__ == "__main__":
    configurar_contatos()