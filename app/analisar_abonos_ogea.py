import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


BASE_DIR = Path(__file__).resolve().parent.parent
PASTA_SAIDA = BASE_DIR / "outputs" / "abonos_ogea"
PASTA_HISTORICO = PASTA_SAIDA / "historico"
PASTA_LOGS = BASE_DIR / "logs" / "abonos_ogea"
ARQUIVO_ATUAL = PASTA_SAIDA / "analise_abonos_ogea_atual.xlsx"
ARQUIVO_REGRAS = BASE_DIR / "bases" / "regras_abonos_ogea.json"

COR_PRETO = "090909"
COR_PRETO_2 = "151515"
COR_DOURADO = "F4C430"
COR_DOURADO_ESCURO = "B08D17"
COR_BRANCO = "FFFFFF"
COR_CINZA = "D9D9D9"
COR_VERDE = "C6EFCE"
COR_VERDE_TEXTO = "006100"
COR_VERMELHO = "FFC7CE"
COR_VERMELHO_TEXTO = "9C0006"
COR_AMARELO = "FFF2CC"
COR_AMARELO_TEXTO = "7F6000"
COR_AZUL_CLARO = "DDEBF7"

PADROES_ABRE = [
    r"\bSEG(?:UNDA)?(?: FEIRA)?\s+(?:A|ATE)\s+SAB(?:ADO)?\b",
    r"\bDE\s+SEGUNDA(?: FEIRA)?\s+(?:A|ATE)\s+SABADO\b",
    r"(?<!NAO )\bABRE(?: AOS?)? SABADOS?\b",
    r"(?<!NAO )\bFUNCIONA(?: AOS?)? SABADOS?\b",
    r"(?<!NAO )\bATENDE(?: AOS?)? SABADOS?\b",
    r"\bSABADOS?\s*(?:DAS|DE|:|-)?\s*\d{1,2}(?::\d{2}|H)",
    r"\bSAB\s*(?:DAS|DE|:|-)?\s*\d{1,2}(?::\d{2}|H)",
]

# Negativas que falam especificamente que sábado não possui atendimento.
PADROES_NAO_ABRE_SABADO = [
    r"\bNAO\s+(?:ABRE|FUNCIONA|ATENDE)(?: AOS?)? SABADOS?\b",
    r"\bFECHADO(?: AOS?)? SABADOS?\b",
    r"\bSABADOS?\s*(?:NAO ABRE|NAO FUNCIONA|NAO ATENDE|FECHADO|SEM ATENDIMENTO)\b",
    r"\b(?:NAO|SEM|FECHADO|ENCERRADO).{0,24}\bSAB(?:ADO)?\b",
    r"\bSAB(?:ADO)?\b.{0,24}\b(?:NAO|SEM|FECHADO|ENCERRADO)\b",
]

# Indícios de funcionamento restrito aos dias úteis. Só valem quando não há
# nenhuma evidência positiva de sábado em qualquer outro campo da mesma OS.
PADROES_APENAS_DIAS_UTEIS = [
    r"\bSEG(?:UNDA)?(?: FEIRA)?\s+(?:A|ATE)\s+SEX(?:TA)?(?: FEIRA)?\b",
    r"\bDE\s+SEGUNDA(?: FEIRA)?\s+(?:A|ATE)\s+SEXTA(?: FEIRA)?\b",
    r"\bSOMENTE\s+(?:DE\s+)?SEGUNDA(?: FEIRA)?\s+(?:A|ATE)\s+SEXTA(?: FEIRA)?\b",
]

COLUNAS_SAIDA = [
    "Decisão",
    "Confiança",
    "OS",
    "Data Limite",
    "Motivo da decisão",
    "Evidência encontrada",
    "Campos com evidência",
    "Cliente",
    "Cidade",
    "Bairro / Distrito",
    "Endereço",
    "Técnico",
    "Serviço",
    "Status",
    "Aba de origem",
    "Linha original",
    "Validação Manual",
    "Observação Manual",
    "Texto completo analisado",
]

COLUNAS_PAINEL = [
    "Indicador",
    "OS",
    "Resumo",
    "Confiança",
    "Data Limite",
    "Técnico",
    "Cliente",
    "Cidade",
    "Serviço",
]



def normalizar_texto(valor):
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    texto = str(valor).strip()
    if not texto or texto.lower() in {"nan", "none", "null", "nat"}:
        return ""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.upper()
    texto = re.sub(r"[^A-Z0-9:]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def limpar_os(valor):
    texto = "" if valor is None else str(valor).strip()
    if texto.endswith(".0"):
        texto = texto[:-2]
    apenas_numeros = re.sub(r"\D", "", texto)
    return apenas_numeros or texto


def abrir_caminho(caminho):
    caminho = str(Path(caminho).resolve())
    if os.name == "nt":
        os.startfile(caminho)
    elif sys.platform == "darwin":
        subprocess.Popen(["open", caminho])
    else:
        subprocess.Popen(["xdg-open", caminho])


def ler_csv_robusto(caminho):
    tentativas = [
        ("utf-8-sig", None),
        ("cp1252", None),
        ("latin1", None),
        ("utf-8-sig", ";"),
        ("cp1252", ";"),
        ("latin1", ";"),
        ("utf-8-sig", ","),
        ("cp1252", ","),
        ("latin1", ","),
    ]
    ultimo_erro = None
    for encoding, separador in tentativas:
        try:
            kwargs = {
                "dtype": str,
                "encoding": encoding,
                "keep_default_na": False,
                "na_filter": False,
                "low_memory": False,
            }
            if separador is None:
                kwargs.update({"sep": None, "engine": "python"})
            else:
                kwargs["sep"] = separador
            df = pd.read_csv(caminho, **kwargs)
            if len(df.columns) > 1:
                return df
        except Exception as erro:
            ultimo_erro = erro
    raise RuntimeError(f"Não consegui ler o CSV. Último erro: {ultimo_erro}")


def ler_arquivo(caminho):
    caminho = Path(caminho)
    extensao = caminho.suffix.lower()

    if extensao in {".csv", ".txt"}:
        df = ler_csv_robusto(caminho)
        df["_Aba Origem"] = "CSV"
        return df

    if extensao in {".xlsx", ".xlsm", ".xls"}:
        try:
            abas = pd.read_excel(
                caminho,
                sheet_name=None,
                dtype=str,
                keep_default_na=False,
                na_filter=False,
            )
        except Exception as erro:
            raise RuntimeError(f"Não consegui ler a planilha: {erro}") from erro

        frames = []
        for nome_aba, df_aba in abas.items():
            if df_aba.empty and len(df_aba.columns) == 0:
                continue
            df_aba = df_aba.copy()
            df_aba["_Aba Origem"] = str(nome_aba)
            frames.append(df_aba)

        if not frames:
            raise RuntimeError("A planilha não possui nenhuma aba com dados.")

        return pd.concat(frames, ignore_index=True, sort=False)

    raise ValueError(
        "Formato não suportado. Use CSV, TXT, XLSX, XLSM ou XLS."
    )


def encontrar_coluna(df, candidatos, contem=None):
    mapa = {normalizar_texto(coluna): coluna for coluna in df.columns}
    for candidato in candidatos:
        normalizado = normalizar_texto(candidato)
        if normalizado in mapa:
            return mapa[normalizado]

    if contem:
        termos = [normalizar_texto(item) for item in contem]
        for coluna_norm, coluna_original in mapa.items():
            if all(termo in coluna_norm for termo in termos):
                return coluna_original

    return None


def extrair_data(valor):
    if valor is None or str(valor).strip() == "":
        return None
    try:
        convertido = pd.to_datetime(valor, dayfirst=True, errors="coerce")
        if pd.isna(convertido):
            return None
        return convertido.date()
    except Exception:
        return None


def carregar_regras_personalizadas():
    padrao = {
        "frases_abre_sabado": [],
        "frases_nao_abre_sabado": [],
    }
    if not ARQUIVO_REGRAS.exists():
        return padrao
    try:
        dados = json.loads(ARQUIVO_REGRAS.read_text(encoding="utf-8"))
        for chave in padrao:
            valores = dados.get(chave, [])
            if isinstance(valores, list):
                padrao[chave] = [normalizar_texto(v) for v in valores if str(v).strip()]
        return padrao
    except Exception as erro:
        print(f"Aviso: não consegui ler regras personalizadas: {erro}")
        return padrao


def valor_indica_horario_aberto(valor):
    texto = normalizar_texto(valor)
    if not texto:
        return False
    if texto in {"SIM", "ABERTO", "FUNCIONA", "ATENDE"}:
        return True
    horarios = re.findall(r"\b(\d{1,2}):([0-5]\d)\b", texto)
    for hora, minuto in horarios:
        if int(hora) != 0 or int(minuto) != 0:
            return True
    if re.search(r"\b\d{1,2}H(?:\d{2})?\b", texto):
        return True
    return False


def valor_indica_fechado(valor):
    texto = normalizar_texto(valor)
    if not texto:
        return True
    if texto in {"NAO", "N", "FECHADO", "NAO ABRE", "SEM ATENDIMENTO", "N A", "00:00", "00 00"}:
        return True
    horarios = re.findall(r"\b(\d{1,2}):([0-5]\d)\b", texto)
    if horarios and all(int(h) == 0 and int(m) == 0 for h, m in horarios):
        return True
    return False


def capturar_campos_relevantes(linha):
    relevantes = []
    for coluna, valor in linha.items():
        if str(coluna).startswith("_"):
            continue
        valor_txt = str(valor).strip()
        if not valor_txt or valor_txt.lower() in {"nan", "none", "null"}:
            continue
        combinado = normalizar_texto(f"{coluna} {valor_txt}")
        if (
            "SAB" in combinado
            or "SEGUNDA A SEXTA" in combinado
            or "SEG A SEX" in combinado
            or "SEGUNDA A SABADO" in combinado
            or "SEG A SAB" in combinado
        ):
            relevantes.append(f"{coluna}: {valor_txt}")
    return relevantes


def analisar_linha(linha, regras):
    campos = []
    abre_evidencias = []
    fecha_sabado_evidencias = []
    dias_uteis_evidencias = []

    for coluna, valor in linha.items():
        if str(coluna).startswith("_"):
            continue
        valor_txt = str(valor).strip()
        if not valor_txt or valor_txt.lower() in {"nan", "none", "null", "nat"}:
            continue

        campo_original = f"{coluna}: {valor_txt}"
        campos.append(campo_original)
        coluna_norm = normalizar_texto(coluna)
        valor_norm = normalizar_texto(valor_txt)
        campo_norm = normalizar_texto(campo_original)

        negativa_explicita = any(
            re.search(padrao, campo_norm)
            for padrao in PADROES_NAO_ABRE_SABADO
        )

        # Regra operacional principal: qualquer menção positiva a SAB/SÁBADO,
        # inclusive em listas como "SEG, TER, QUA, QUI, SEX, SAB", significa
        # que a OS pode ser atendida no sábado. A exceção é quando o mesmo
        # campo declara explicitamente que sábado não abre/não atende.
        menciona_sabado = bool(re.search(r"\bSAB(?:ADO|ADOS)?\b", valor_norm))

        if negativa_explicita:
            fecha_sabado_evidencias.append(campo_original)
        elif menciona_sabado:
            abre_evidencias.append(campo_original)

        if "SAB" in coluna_norm:
            if valor_indica_horario_aberto(valor_txt) and not negativa_explicita:
                abre_evidencias.append(campo_original)
            elif valor_indica_fechado(valor_txt):
                fecha_sabado_evidencias.append(campo_original)

        for padrao in PADROES_APENAS_DIAS_UTEIS:
            if re.search(padrao, campo_norm):
                dias_uteis_evidencias.append(campo_original)
                break

        for padrao in PADROES_ABRE:
            if re.search(padrao, campo_norm) and not negativa_explicita:
                abre_evidencias.append(campo_original)
                break

    texto_completo = " | ".join(campos)
    texto_norm = normalizar_texto(texto_completo)

    for frase in regras.get("frases_nao_abre_sabado", []):
        if not frase or frase not in texto_norm:
            continue

        # "SEG A SEX" e equivalentes não são uma negativa explícita de
        # sábado. São apenas indícios de dias úteis e perdem prioridade se
        # qualquer outro campo informar atendimento no sábado.
        if "SEX" in frase and "SAB" not in frase:
            dias_uteis_evidencias.append(f"Regra personalizada: {frase}")
        else:
            fecha_sabado_evidencias.append(f"Regra personalizada: {frase}")

    for frase in regras.get("frases_abre_sabado", []):
        if not frase or frase not in texto_norm:
            continue

        # Evita interpretar "NÃO FUNCIONA AOS SÁBADOS" como positiva só
        # porque contém o trecho "FUNCIONA AOS SÁBADOS".
        if f"NAO {frase}" in texto_norm:
            continue

        abre_evidencias.append(f"Regra personalizada: {frase}")

    # Remove repetições sem perder a ordem.
    abre_evidencias = list(dict.fromkeys(abre_evidencias))
    fecha_sabado_evidencias = list(dict.fromkeys(fecha_sabado_evidencias))
    dias_uteis_evidencias = list(dict.fromkeys(dias_uteis_evidencias))
    campos_relevantes = capturar_campos_relevantes(linha)

    # Prioridade correta:
    # 1. se existe evidência positiva de sábado, NÃO ABONAR;
    # 2. só vira conflito se também houver negativa explícita sobre sábado;
    # 3. "segunda a sexta" só abona quando não existe nenhuma evidência
    #    positiva de sábado em qualquer outro campo.
    if abre_evidencias and fecha_sabado_evidencias:
        decisao = "REVISAR MANUALMENTE"
        confianca = "BAIXA"
        motivo = (
            "Foram encontradas informações conflitantes: um campo indica "
            "atendimento aos sábados e outro declara fechamento aos sábados."
        )
        evidencia = (
            "ABRE: " + " ; ".join(abre_evidencias[:4])
            + " | NÃO ABRE: " + " ; ".join(fecha_sabado_evidencias[:4])
        )
    elif abre_evidencias:
        decisao = "NÃO ABONAR"
        confianca = "ALTA"
        motivo = (
            "Existe indicação de que o estabelecimento pode ser atendido "
            "no sábado. Portanto, a OS deve ser atendida e não abonada."
        )
        evidencia = " ; ".join(abre_evidencias[:6])
    elif fecha_sabado_evidencias:
        decisao = "ABONAR"
        confianca = "ALTA"
        motivo = "Existe indicação explícita de que o estabelecimento não atende aos sábados."
        evidencia = " ; ".join(fecha_sabado_evidencias[:6])
    elif dias_uteis_evidencias:
        decisao = "ABONAR"
        confianca = "ALTA"
        motivo = (
            "A OS informa funcionamento somente de segunda a sexta e não "
            "apresenta nenhuma evidência positiva de atendimento no sábado."
        )
        evidencia = " ; ".join(dias_uteis_evidencias[:6])
    else:
        decisao = "ABONAR"
        confianca = "MÉDIA"
        motivo = (
            "Nenhuma indicação de funcionamento aos sábados foi encontrada. "
            "Pela regra operacional informada, deve ser considerado fechado aos sábados."
        )
        evidencia = "Nenhuma evidência de abertura aos sábados encontrada em nenhum campo."

    return {
        "Decisão": decisao,
        "Confiança": confianca,
        "Motivo da decisão": motivo,
        "Evidência encontrada": evidencia,
        "Campos com evidência": " | ".join(campos_relevantes),
        "Texto completo analisado": texto_completo,
    }


def preparar_resultado(df, hoje):
    col_os = encontrar_coluna(
        df,
        ["Código", "Codigo", "OS", "Número da OS", "Numero da OS", "Ordem de Serviço"],
    )
    col_data = encontrar_coluna(
        df,
        ["Data Limite", "Vencimento", "Data de Vencimento", "Prazo", "Data Prazo"],
        contem=["DATA", "LIMITE"],
    )

    avisos = []
    df_trabalho = df.copy()
    df_trabalho["_Linha Original"] = range(2, len(df_trabalho) + 2)

    if col_data:
        datas = df_trabalho[col_data].map(extrair_data)
        df_trabalho = df_trabalho[datas == hoje].copy()
        df_trabalho["_Data Interpretada"] = hoje.strftime("%d/%m/%Y")
    else:
        avisos.append(
            "Não foi possível identificar a coluna de Data Limite. Todas as linhas do relatório foram analisadas."
        )
        df_trabalho["_Data Interpretada"] = ""

    if hoje.weekday() != 5:
        avisos.append(
            "A data de execução não é um sábado. A análise foi realizada, mas a regra operacional deve ser validada."
        )

    regras = carregar_regras_personalizadas()

    campos_chave = {
        "Cliente": encontrar_coluna(df_trabalho, ["Cliente", "Nome Cliente", "Razão Social", "Razao Social"]),
        "Cidade": encontrar_coluna(df_trabalho, ["Cidade", "Município", "Municipio"]),
        "Bairro / Distrito": encontrar_coluna(df_trabalho, ["Distrito", "Bairro", "Bairro / Distrito"]),
        "Endereço": encontrar_coluna(df_trabalho, ["Endereço", "Endereco", "Logradouro"]),
        "Técnico": encontrar_coluna(df_trabalho, ["Técnico", "Tecnico", "Prestador"]),
        "Serviço": encontrar_coluna(df_trabalho, ["Serviço", "Servico", "Tipo de Serviço", "Tipo de Servico"]),
        "Status": encontrar_coluna(df_trabalho, ["Status", "Situação", "Situacao"]),
    }

    linhas_saida = []
    for _, linha in df_trabalho.iterrows():
        analise = analisar_linha(linha, regras)
        registro = {
            "Decisão": analise["Decisão"],
            "Confiança": analise["Confiança"],
            "OS": limpar_os(linha.get(col_os, "")) if col_os else "",
            "Data Limite": (
                str(linha.get(col_data, "")).strip() if col_data else ""
            ),
            "Motivo da decisão": analise["Motivo da decisão"],
            "Evidência encontrada": analise["Evidência encontrada"],
            "Campos com evidência": analise["Campos com evidência"],
            "Aba de origem": str(linha.get("_Aba Origem", "")),
            "Linha original": int(linha.get("_Linha Original", 0)),
            "Validação Manual": "Pendente",
            "Observação Manual": "",
            "Texto completo analisado": analise["Texto completo analisado"],
        }
        for nome_saida, coluna_origem in campos_chave.items():
            registro[nome_saida] = (
                str(linha.get(coluna_origem, "")).strip() if coluna_origem else ""
            )
        linhas_saida.append(registro)

    resultado = pd.DataFrame(linhas_saida, columns=COLUNAS_SAIDA)
    return resultado, avisos, col_data, col_os


def estilizar_titulo(ws, titulo, subtitulo=None):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:S1")
    celula = ws["A1"]
    celula.value = titulo
    celula.fill = PatternFill("solid", fgColor=COR_PRETO)
    celula.font = Font(color=COR_DOURADO, size=20, bold=True)
    celula.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    if subtitulo:
        ws.merge_cells("A2:S2")
        ws["A2"] = subtitulo
        ws["A2"].fill = PatternFill("solid", fgColor=COR_PRETO_2)
        ws["A2"].font = Font(color=COR_BRANCO, size=10, italic=True)
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[2].height = 30


def formatar_tabela(ws, linha_cabecalho, total_linhas, total_colunas):
    fill_header = PatternFill("solid", fgColor=COR_PRETO)
    font_header = Font(color=COR_DOURADO, bold=True)
    borda = Border(
        left=Side(style="thin", color=COR_DOURADO_ESCURO),
        right=Side(style="thin", color=COR_DOURADO_ESCURO),
        top=Side(style="thin", color=COR_DOURADO_ESCURO),
        bottom=Side(style="thin", color=COR_DOURADO_ESCURO),
    )
    fill_par = PatternFill("solid", fgColor="FAFAFA")
    fill_impar = PatternFill("solid", fgColor="F2F2F2")

    for cell in ws[linha_cabecalho]:
        if cell.column > total_colunas:
            break
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda

    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=linha_cabecalho + 1,
            max_row=max(linha_cabecalho + 1, total_linhas),
            min_col=1,
            max_col=total_colunas,
        ),
        start=1,
    ):
        fill_linha = fill_par if row_idx % 2 else fill_impar
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="D9D9D9"))
            if cell.column != 1:
                cell.fill = fill_linha

    ws.freeze_panes = f"A{linha_cabecalho + 1}"
    ws.auto_filter.ref = f"A{linha_cabecalho}:{get_column_letter(total_colunas)}{max(linha_cabecalho, total_linhas)}"

    larguras = {
        "A": 21,
        "B": 12,
        "C": 15,
        "D": 20,
        "E": 42,
        "F": 48,
        "G": 42,
        "H": 28,
        "I": 18,
        "J": 24,
        "K": 34,
        "L": 20,
        "M": 31,
        "N": 18,
        "O": 18,
        "P": 14,
        "Q": 20,
        "R": 30,
        "S": 70,
    }
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    if total_linhas > linha_cabecalho:
        faixa_decisao = f"A{linha_cabecalho + 1}:A{total_linhas}"
        ws.conditional_formatting.add(
            faixa_decisao,
            FormulaRule(
                formula=[f'$A{linha_cabecalho + 1}="ABONAR"'],
                fill=PatternFill("solid", fgColor=COR_VERMELHO),
                font=Font(color=COR_VERMELHO_TEXTO, bold=True),
            ),
        )
        ws.conditional_formatting.add(
            faixa_decisao,
            FormulaRule(
                formula=[f'$A{linha_cabecalho + 1}="NÃO ABONAR"'],
                fill=PatternFill("solid", fgColor=COR_VERDE),
                font=Font(color=COR_VERDE_TEXTO, bold=True),
            ),
        )
        ws.conditional_formatting.add(
            faixa_decisao,
            FormulaRule(
                formula=[f'$A{linha_cabecalho + 1}="REVISAR MANUALMENTE"'],
                fill=PatternFill("solid", fgColor=COR_AMARELO),
                font=Font(color=COR_AMARELO_TEXTO, bold=True),
            ),
        )

        validacao = DataValidation(
            type="list",
            formula1='"Pendente,Correto,Corrigir para ABONAR,Corrigir para NÃO ABONAR,Revisado"',
            allow_blank=True,
        )
        ws.add_data_validation(validacao)
        validacao.add(f"Q{linha_cabecalho + 1}:Q{total_linhas}")



def adicionar_aba_dados(wb, nome, df, subtitulo):
    ws = wb.create_sheet(nome)
    estilizar_titulo(ws, f"Central Visconde | {nome}", subtitulo)
    linha_cabecalho = 4

    for coluna_idx, coluna in enumerate(COLUNAS_SAIDA, start=1):
        ws.cell(linha_cabecalho, coluna_idx, coluna)

    for linha_idx, (_, registro) in enumerate(df.iterrows(), start=linha_cabecalho + 1):
        for coluna_idx, coluna in enumerate(COLUNAS_SAIDA, start=1):
            ws.cell(linha_idx, coluna_idx, registro.get(coluna, ""))

    total_linhas = linha_cabecalho + len(df)
    formatar_tabela(ws, linha_cabecalho, total_linhas, len(COLUNAS_SAIDA))
    return ws



def limitar_texto(texto, limite=110):
    texto = re.sub(r"\s+", " ", str(texto or "")).strip()
    if len(texto) <= limite:
        return texto
    return texto[: limite - 3].rstrip() + "..."


def extrair_evidencia_curta(evidencia, campos_evidencia):
    candidatos = []

    for item in str(campos_evidencia or "").split("|"):
        item = item.strip()
        if item:
            candidatos.append(item)

    if not candidatos:
        for item in re.split(r"[;|]", str(evidencia or "")):
            item = item.strip()
            if not item:
                continue
            if item.startswith("Padrão encontrado:") or item.startswith("Regra personalizada:"):
                continue
            candidatos.append(item)

    if not candidatos:
        return ""

    detalhe = candidatos[0]
    if ":" in detalhe:
        campo, valor = detalhe.split(":", 1)
        campo = campo.strip()
        valor = limitar_texto(valor.strip(), 90)
        return f"{campo}: {valor}" if valor else campo

    return limitar_texto(detalhe, 96)

def resumir_decisao(decisao, confianca, motivo, evidencia, campos_evidencia):
    if decisao == "ABONAR":
        if confianca == "ALTA":
            resumo = "Abonar: há indício de que o local não atende aos sábados."
        else:
            resumo = "Abonar: não foi encontrada evidência de atendimento aos sábados."
    elif decisao == "NÃO ABONAR":
        resumo = "Atender: há indício de que o local funciona aos sábados."
    else:
        resumo = "Revisar: foram encontrados sinais conflitantes sobre sábado."

    detalhe = extrair_evidencia_curta(evidencia, campos_evidencia)
    if detalhe:
        resumo += f" Evidência: {detalhe}."

    return resumo


def montar_painel_operacional(resultado):
    if resultado.empty:
        return pd.DataFrame(columns=COLUNAS_PAINEL)

    painel = pd.DataFrame()
    painel["Indicador"] = resultado["Decisão"].map({
        "ABONAR": "ABONAR",
        "NÃO ABONAR": "ATENDER",
        "REVISAR MANUALMENTE": "REVISAR",
    }).fillna("REVISAR")
    painel["OS"] = resultado["OS"]
    painel["Resumo"] = [
        resumir_decisao(dec, conf, mot, evid, campos)
        for dec, conf, mot, evid, campos in zip(
            resultado["Decisão"],
            resultado["Confiança"],
            resultado["Motivo da decisão"],
            resultado["Evidência encontrada"],
            resultado["Campos com evidência"],
        )
    ]
    painel["Confiança"] = resultado["Confiança"]
    painel["Data Limite"] = resultado["Data Limite"]
    painel["Técnico"] = resultado["Técnico"]
    painel["Cliente"] = resultado["Cliente"]
    painel["Cidade"] = resultado["Cidade"]
    painel["Serviço"] = resultado["Serviço"]

    ordem = {"ABONAR": 0, "ATENDER": 1, "REVISAR": 2}
    painel["__ordem"] = painel["Indicador"].map(ordem).fillna(9)
    painel = painel.sort_values(by=["__ordem", "Técnico", "OS"], kind="stable").drop(columns=["__ordem"])
    return painel[COLUNAS_PAINEL]


def adicionar_aba_painel(wb, resultado):
    ws = wb.create_sheet("Painel")
    estilizar_titulo(
        ws,
        "Central Visconde | Painel Operacional",
        "Visão rápida para o dia a dia: o que ABONAR, o que ATENDER e o que REVISAR manualmente.",
    )

    painel = montar_painel_operacional(resultado)
    qtd_abonar = int((painel["Indicador"] == "ABONAR").sum()) if not painel.empty else 0
    qtd_atender = int((painel["Indicador"] == "ATENDER").sum()) if not painel.empty else 0
    qtd_revisar = int((painel["Indicador"] == "REVISAR").sum()) if not painel.empty else 0
    total = len(painel)

    cards = [
        ("A4", "ABONAR", qtd_abonar, COR_VERMELHO, COR_VERMELHO_TEXTO),
        ("D4", "ATENDER", qtd_atender, COR_VERDE, COR_VERDE_TEXTO),
        ("G4", "REVISAR", qtd_revisar, COR_AMARELO, COR_AMARELO_TEXTO),
        ("J4", "TOTAL", total, COR_AZUL_CLARO, "1F4E78"),
    ]
    for celula, titulo, valor, cor_fill, cor_font in cards:
        coluna = ws[celula].column
        linha = ws[celula].row
        ws.merge_cells(start_row=linha, start_column=coluna, end_row=linha, end_column=coluna + 1)
        ws.cell(linha, coluna, titulo)
        ws.cell(linha, coluna).fill = PatternFill("solid", fgColor=COR_PRETO)
        ws.cell(linha, coluna).font = Font(color=COR_DOURADO, bold=True)
        ws.cell(linha, coluna).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=linha + 1, start_column=coluna, end_row=linha + 2, end_column=coluna + 1)
        ws.cell(linha + 1, coluna, valor)
        ws.cell(linha + 1, coluna).fill = PatternFill("solid", fgColor=cor_fill)
        ws.cell(linha + 1, coluna).font = Font(color=cor_font, bold=True, size=22)
        ws.cell(linha + 1, coluna).alignment = Alignment(horizontal="center", vertical="center")

    ws["A8"] = "Legenda"
    ws["A8"].font = Font(color=COR_DOURADO, bold=True)
    ws["B8"] = "ABONAR = sugerir abono | ATENDER = cliente pode ser atendido sábado | REVISAR = conferir manualmente"
    ws["B8"].alignment = Alignment(wrap_text=True)

    linha_cabecalho = 10
    for coluna_idx, coluna in enumerate(COLUNAS_PAINEL, start=1):
        ws.cell(linha_cabecalho, coluna_idx, coluna)

    fill_header = PatternFill("solid", fgColor=COR_PRETO)
    font_header = Font(color=COR_DOURADO, bold=True)
    borda_header = Border(
        left=Side(style="thin", color=COR_DOURADO_ESCURO),
        right=Side(style="thin", color=COR_DOURADO_ESCURO),
        top=Side(style="thin", color=COR_DOURADO_ESCURO),
        bottom=Side(style="thin", color=COR_DOURADO_ESCURO),
    )
    borda_linha = Border(bottom=Side(style="thin", color="E5E5E5"))
    cores_linha = {
        "ABONAR": ("FFF2F2", "B71C1C"),
        "ATENDER": ("EEF8EE", "1B5E20"),
        "REVISAR": ("FFF9E8", "8A6D1D"),
    }
    cores_conf = {
        "ALTA": ("E2F0D9", "2F6B2F"),
        "MÉDIA": ("FFF2CC", "7F6000"),
        "BAIXA": ("FCE4D6", "A64D00"),
    }

    for cell in ws[linha_cabecalho]:
        if cell.column > len(COLUNAS_PAINEL):
            break
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda_header

    for linha_idx, (_, registro) in enumerate(painel.iterrows(), start=linha_cabecalho + 1):
        indicador = str(registro.get("Indicador", "")).strip() or "REVISAR"
        fill_cor, fonte_cor = cores_linha.get(indicador, ("F3F3F3", "333333"))
        for coluna_idx, coluna in enumerate(COLUNAS_PAINEL, start=1):
            valor = registro.get(coluna, "")
            cell = ws.cell(linha_idx, coluna_idx, valor)
            cell.border = borda_linha
            cell.fill = PatternFill("solid", fgColor=fill_cor)
            if coluna in {"Indicador", "OS", "Confiança", "Data Limite", "Cidade", "Serviço"}:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if coluna == "Indicador":
                cell.font = Font(color=fonte_cor, bold=True, size=12)
            elif coluna == "Confiança":
                conf_fill, conf_font = cores_conf.get(str(valor).strip(), (fill_cor, fonte_cor))
                cell.fill = PatternFill("solid", fgColor=conf_fill)
                cell.font = Font(color=conf_font, bold=True)
            else:
                cell.font = Font(color="202020")
        ws.row_dimensions[linha_idx].height = 40

    total_linhas = linha_cabecalho + len(painel)
    ws.freeze_panes = f"A{linha_cabecalho + 1}"
    ws.auto_filter.ref = f"A{linha_cabecalho}:I{max(linha_cabecalho, total_linhas)}"

    larguras = {
        "A": 15,
        "B": 13,
        "C": 62,
        "D": 12,
        "E": 20,
        "F": 22,
        "G": 34,
        "H": 18,
        "I": 24,
        "J": 12,
        "K": 12,
    }
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    return ws



def criar_planilha(resultado, caminho_entrada, hoje, avisos, col_data, col_os):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    total = len(resultado)
    qtd_abonar = int((resultado["Decisão"] == "ABONAR").sum()) if total else 0
    qtd_nao = int((resultado["Decisão"] == "NÃO ABONAR").sum()) if total else 0
    qtd_revisar = int((resultado["Decisão"] == "REVISAR MANUALMENTE").sum()) if total else 0

    ws.merge_cells("A1:H2")
    ws["A1"] = "CENTRAL VISCONDE | ANALISTA DE ABONOS OGEA"
    ws["A1"].fill = PatternFill("solid", fgColor=COR_PRETO)
    ws["A1"].font = Font(color=COR_DOURADO, size=22, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 20

    dados = [
        ("Data analisada", hoje.strftime("%d/%m/%Y")),
        ("Dia da semana", ["segunda-feira", "terça-feira", "quarta-feira", "quinta-feira", "sexta-feira", "sábado", "domingo"][hoje.weekday()]),
        ("Arquivo de origem", str(Path(caminho_entrada).name)),
        ("Coluna de data identificada", col_data or "Não identificada"),
        ("Coluna de OS identificada", col_os or "Não identificada"),
        ("Total analisado", total),
    ]

    ws["A4"] = "Informações da análise"
    ws["A4"].font = Font(bold=True, color=COR_DOURADO, size=13)
    for idx, (rotulo, valor) in enumerate(dados, start=5):
        ws[f"A{idx}"] = rotulo
        ws[f"A{idx}"].font = Font(bold=True, color=COR_BRANCO)
        ws[f"A{idx}"].fill = PatternFill("solid", fgColor=COR_PRETO_2)
        ws[f"B{idx}"] = valor
        ws[f"B{idx}"].fill = PatternFill("solid", fgColor="F3F3F3")

    cards = [
        ("D4", "ABONAR", qtd_abonar, COR_VERMELHO, COR_VERMELHO_TEXTO),
        ("F4", "NÃO ABONAR", qtd_nao, COR_VERDE, COR_VERDE_TEXTO),
        ("D8", "REVISAR", qtd_revisar, COR_AMARELO, COR_AMARELO_TEXTO),
        ("F8", "TOTAL", total, COR_AZUL_CLARO, "1F4E78"),
    ]
    for celula, titulo, valor, cor_fill, cor_font in cards:
        coluna = ws[celula].column
        linha = ws[celula].row
        ws.merge_cells(start_row=linha, start_column=coluna, end_row=linha, end_column=coluna + 1)
        ws.cell(linha, coluna, titulo)
        ws.cell(linha, coluna).fill = PatternFill("solid", fgColor=COR_PRETO)
        ws.cell(linha, coluna).font = Font(color=COR_DOURADO, bold=True)
        ws.cell(linha, coluna).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=linha + 1, start_column=coluna, end_row=linha + 2, end_column=coluna + 1)
        ws.cell(linha + 1, coluna, valor)
        ws.cell(linha + 1, coluna).fill = PatternFill("solid", fgColor=cor_fill)
        ws.cell(linha + 1, coluna).font = Font(color=cor_font, bold=True, size=24)
        ws.cell(linha + 1, coluna).alignment = Alignment(horizontal="center", vertical="center")

    ws["A13"] = "Regra operacional aplicada"
    ws["A13"].font = Font(bold=True, color=COR_DOURADO, size=13)
    ws.merge_cells("A14:H16")
    ws["A14"] = (
        "A análise procura informações em TODOS os campos da OS, inclusive bairro, endereço, observação, cliente e serviço. "
        "Quando existe indicação clara de abertura aos sábados, classifica como NÃO ABONAR. Quando existe indicação de fechamento, "
        "classifica como ABONAR. Quando não existe nenhuma informação sobre sábado, aplica a regra informada pela operação: considerar que não abre sábado e classificar como ABONAR."
    )
    ws["A14"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A14"].fill = PatternFill("solid", fgColor="FFF8DC")
    ws["A14"].border = Border(left=Side(style="medium", color=COR_DOURADO), right=Side(style="medium", color=COR_DOURADO))

    linha_aviso = 18
    ws[f"A{linha_aviso}"] = "Avisos"
    ws[f"A{linha_aviso}"].font = Font(bold=True, color=COR_DOURADO, size=13)
    if not avisos:
        avisos = ["Nenhum aviso técnico nesta execução."]
    for aviso in avisos:
        linha_aviso += 1
        ws.merge_cells(start_row=linha_aviso, start_column=1, end_row=linha_aviso, end_column=8)
        ws.cell(linha_aviso, 1, f"• {aviso}")
        ws.cell(linha_aviso, 1).alignment = Alignment(wrap_text=True)
        ws.cell(linha_aviso, 1).fill = PatternFill("solid", fgColor=COR_AMARELO)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42
    for col in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 18

    adicionar_aba_painel(wb, resultado)

    adicionar_aba_dados(
        wb,
        "Abonar",
        resultado[resultado["Decisão"] == "ABONAR"],
        "OSs com indicação de fechamento aos sábados ou sem qualquer evidência de abertura.",
    )
    adicionar_aba_dados(
        wb,
        "Não Abonar",
        resultado[resultado["Decisão"] == "NÃO ABONAR"],
        "OSs com indicação explícita de atendimento ou funcionamento aos sábados.",
    )
    adicionar_aba_dados(
        wb,
        "Revisar",
        resultado[resultado["Decisão"] == "REVISAR MANUALMENTE"],
        "OSs com informações conflitantes que exigem conferência manual.",
    )
    adicionar_aba_dados(
        wb,
        "Todos",
        resultado,
        "Base completa da análise. Use Validação Manual e Observação Manual durante a conferência.",
    )

    regras_ws = wb.create_sheet("Regras")
    estilizar_titulo(
        regras_ws,
        "Central Visconde | Regras do Analista",
        "Esta aba documenta a lógica da primeira versão. As validações manuais serão usadas para aprimorar as próximas versões.",
    )
    regras_dados = [
        ["Prioridade", "Situação", "Resultado", "Exemplos"],
        [1, "Qualquer indicação positiva de sábado", "NÃO ABONAR", "SEG, TER, QUA, QUI, SEX, SAB; segunda a sábado; sábado 08:00 às 12:00"],
        [2, "Indicação explícita de que sábado não abre", "ABONAR", "fechado sábado; não abre aos sábados; sábado sem atendimento"],
        [3, "Somente segunda a sexta, sem sábado positivo", "ABONAR", "segunda a sexta; somente de segunda a sexta"],
        [4, "Informações conflitantes sobre sábado", "REVISAR MANUALMENTE", "um campo diz que abre sábado e outro diz fechado sábado"],
        [5, "Nenhuma informação sobre sábado", "ABONAR", "regra padrão informada pela operação"],
    ]
    for r_idx, linha in enumerate(regras_dados, start=4):
        for c_idx, valor in enumerate(linha, start=1):
            regras_ws.cell(r_idx, c_idx, valor)
    formatar_tabela(regras_ws, 4, 9, 4)
    regras_ws.column_dimensions["A"].width = 12
    regras_ws.column_dimensions["B"].width = 44
    regras_ws.column_dimensions["C"].width = 24
    regras_ws.column_dimensions["D"].width = 70

    return wb


def salvar_log(caminho_entrada, resultado, avisos, caminho_saida):
    PASTA_LOGS.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho_log = PASTA_LOGS / f"analise_abonos_{timestamp}.txt"
    linhas = [
        "CENTRAL VISCONDE - ANALISTA DE ABONOS OGEA",
        f"Executado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        f"Arquivo: {caminho_entrada}",
        f"Saída: {caminho_saida}",
        f"Total: {len(resultado)}",
        f"Abonar: {(resultado['Decisão'] == 'ABONAR').sum() if len(resultado) else 0}",
        f"Não abonar: {(resultado['Decisão'] == 'NÃO ABONAR').sum() if len(resultado) else 0}",
        f"Revisar: {(resultado['Decisão'] == 'REVISAR MANUALMENTE').sum() if len(resultado) else 0}",
        "",
        "Avisos:",
    ]
    linhas.extend(f"- {aviso}" for aviso in avisos)
    caminho_log.write_text("\n".join(linhas), encoding="utf-8")
    return caminho_log


def executar(caminho_entrada, abrir_resultado=True):
    caminho_entrada = Path(caminho_entrada).expanduser().resolve()
    if not caminho_entrada.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_entrada}")

    print("=" * 70)
    print("CENTRAL VISCONDE | ANALISTA DE ABONOS OGEA")
    print("=" * 70)
    print(f"Arquivo: {caminho_entrada}")

    hoje = date.today()
    print(f"Data considerada: {hoje.strftime('%d/%m/%Y')}")
    print("Lendo relatório...")
    df = ler_arquivo(caminho_entrada)
    print(f"Linhas recebidas: {len(df)}")
    print(f"Colunas recebidas: {len(df.columns)}")

    resultado, avisos, col_data, col_os = preparar_resultado(df, hoje)
    print(f"Linhas da data analisada: {len(resultado)}")

    qtd_abonar = int((resultado["Decisão"] == "ABONAR").sum()) if len(resultado) else 0
    qtd_nao = int((resultado["Decisão"] == "NÃO ABONAR").sum()) if len(resultado) else 0
    qtd_revisar = int((resultado["Decisão"] == "REVISAR MANUALMENTE").sum()) if len(resultado) else 0
    print(f"ABONAR: {qtd_abonar}")
    print(f"NÃO ABONAR: {qtd_nao}")
    print(f"REVISAR MANUALMENTE: {qtd_revisar}")

    PASTA_SAIDA.mkdir(parents=True, exist_ok=True)
    PASTA_HISTORICO.mkdir(parents=True, exist_ok=True)

    wb = criar_planilha(resultado, caminho_entrada, hoje, avisos, col_data, col_os)
    wb.save(ARQUIVO_ATUAL)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho_historico = PASTA_HISTORICO / f"analise_abonos_ogea_{timestamp}.xlsx"
    shutil.copy2(ARQUIVO_ATUAL, caminho_historico)
    caminho_log = salvar_log(caminho_entrada, resultado, avisos, ARQUIVO_ATUAL)

    print("")
    print("=" * 70)
    print(f"Planilha criada: {ARQUIVO_ATUAL}")
    print(f"Cópia histórica: {caminho_historico}")
    print(f"Log: {caminho_log}")
    print("=" * 70)

    if abrir_resultado:
        abrir_caminho(ARQUIVO_ATUAL)

    return ARQUIVO_ATUAL


def selecionar_arquivo():
    from tkinter import Tk, filedialog

    root = Tk()
    root.withdraw()
    root.update()
    arquivo = filedialog.askopenfilename(
        title="Selecione o relatório da OGEA",
        filetypes=[
            ("Relatórios", "*.csv *.txt *.xlsx *.xlsm *.xls"),
            ("Todos os arquivos", "*.*"),
        ],
    )
    root.destroy()
    return arquivo


def main():
    parser = argparse.ArgumentParser(
        description="Analisa OSs OGEA com vencimento na data atual e identifica possíveis abonos de sábado."
    )
    parser.add_argument("--arquivo", default="", help="Relatório OGEA em CSV ou Excel.")
    parser.add_argument("--sem-abrir", action="store_true", help="Não abre a planilha ao terminar.")
    args = parser.parse_args()

    arquivo = args.arquivo or selecionar_arquivo()
    if not arquivo:
        print("Nenhum arquivo selecionado.")
        return 1

    try:
        executar(arquivo, abrir_resultado=not args.sem_abrir)
        return 0
    except Exception as erro:
        print("")
        print("ERRO AO ANALISAR ABONOS")
        print(str(erro))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
