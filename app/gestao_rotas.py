import difflib
import getpass
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import unicodedata
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, ttk

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


from caminho_base import BASE_DIR, FROZEN, RECURSOS_DIR

ARQUIVO_REGRAS = BASE_DIR / "bases" / "regras_roteirizacao.xlsx"
ARQUIVO_ROTEIRIZACAO = BASE_DIR / "outputs" / "roteirizacao" / "roteirizacao_atual.xlsx"
ARQUIVO_TEMPORARIO = BASE_DIR / "outputs" / "roteirizacao" / "resolucoes_temporarias.json"
SCRIPT_ROTEIRIZACAO = RECURSOS_DIR / "app" / "gerar_roteirizacao.py"
PASTA_BACKUPS = BASE_DIR / "bases" / "backups_roteirizacao"

COR_FUNDO = "#080808"
COR_FUNDO_2 = "#101010"
COR_CARD = "#171717"
COR_BORDA = "#584A18"
COR_DOURADO = "#F4C430"
COR_DOURADO_ESCURO = "#9B7910"
COR_BRANCO = "#F5F5F5"
COR_TEXTO = "#D5D5D5"
COR_TEXTO_FRACO = "#929292"
COR_VERDE = "#2EAD68"
COR_VERMELHO = "#D14949"
COR_AZUL = "#3278C8"
COR_LARANJA = "#D88A24"

CAB_REGRAS = [
    "Ativo", "Prioridade", "Técnico", "Tipo de Regra", "Cidade",
    "Bairro / Localidade Normalizada", "Origem", "Regra Original", "Observação",
]
CAB_ALIASES = [
    "Ativo", "Cidade", "Nome recebido", "Nome considerado", "Técnico", "Observação", "Origem",
]
CAB_HISTORICO = [
    "Data/Hora", "Tipo", "Ação", "Origem", "Cidade", "Chave",
    "Valor novo", "Valor anterior", "Observação", "Usuário",
]


class BotaoVisconde(tk.Label):
    """Botão visual consistente no macOS e Windows.

    O tk.Button nativo do macOS pode ignorar as cores configuradas e deixar
    texto claro sobre fundo claro. Este componente usa um Label clicável para
    manter as cores da identidade Visconde em todos os sistemas.
    """

    def __init__(self, parent, texto, comando, cor, fg, width=None):
        self._comando = comando
        self._habilitado = True
        self._cor_normal = cor
        self._cor_hover = self._calcular_hover(cor)
        self._fg_normal = fg
        self._fg_desabilitado = "#777777"
        self._cor_desabilitado = "#2B2B2B"
        super().__init__(
            parent,
            text=texto,
            bg=cor,
            fg=fg,
            font=("Arial", 9, "bold"),
            padx=12,
            pady=9,
            cursor="hand2",
            anchor="center",
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=cor,
            width=width,
        )
        self.bind("<Button-1>", self._clicar)
        self.bind("<Enter>", self._entrar)
        self.bind("<Leave>", self._sair)
        self.bind("<Return>", self._clicar)
        self.bind("<space>", self._clicar)

    @staticmethod
    def _calcular_hover(cor):
        try:
            valor = cor.lstrip("#")
            r, g, b = (int(valor[i:i + 2], 16) for i in (0, 2, 4))
            r = min(255, int(r * 1.16) + 5)
            g = min(255, int(g * 1.16) + 5)
            b = min(255, int(b * 1.16) + 5)
            return f"#{r:02X}{g:02X}{b:02X}"
        except Exception:
            return cor

    def _clicar(self, event=None):
        if self._habilitado and callable(self._comando):
            self._comando()

    def _entrar(self, event=None):
        if self._habilitado:
            self.config(bg=self._cor_hover, highlightbackground=self._cor_hover)

    def _sair(self, event=None):
        if self._habilitado:
            self.config(bg=self._cor_normal, highlightbackground=self._cor_normal)

    def set_enabled(self, enabled):
        self._habilitado = bool(enabled)
        if self._habilitado:
            self.config(
                bg=self._cor_normal,
                fg=self._fg_normal,
                cursor="hand2",
                highlightbackground=self._cor_normal,
            )
        else:
            self.config(
                bg=self._cor_desabilitado,
                fg=self._fg_desabilitado,
                cursor="arrow",
                highlightbackground="#3A3A3A",
            )


def normalizar_texto(valor):
    if valor is None:
        return ""
    texto = str(valor).strip()
    if texto.lower() in {"", "nan", "none", "null", "nat"}:
        return ""
    texto = texto.replace("Μ", "A").replace("µ", "A")
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.upper()
    texto = re.sub(r"[^A-Z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def ativo(valor):
    return normalizar_texto(valor) in {"SIM", "S", "YES", "Y", "1", "TRUE"}


def expandir_abreviacoes(texto):
    mapa = {
        "S": "SAO", "STO": "SANTO", "STA": "SANTA", "JD": "JARDIM",
        "VL": "VILA", "PQUE": "PARQUE", "PQ": "PARQUE", "CH": "CHACARA",
        "LOT": "LOTEAMENTO", "COND": "CONDOMINIO", "RES": "RESIDENCIAL",
        "DIST": "DISTRITO", "B": "BAIRRO",
    }
    tokens = normalizar_texto(texto).split()
    return " ".join(mapa.get(token, token) for token in tokens)


def similaridade(a, b):
    a1 = expandir_abreviacoes(a)
    b1 = expandir_abreviacoes(b)
    direta = difflib.SequenceMatcher(None, a1, b1).ratio()
    tokens_a = " ".join(sorted(a1.split()))
    tokens_b = " ".join(sorted(b1.split()))
    por_tokens = difflib.SequenceMatcher(None, tokens_a, tokens_b).ratio()
    contem = 0.93 if a1 and b1 and (a1 in b1 or b1 in a1) else 0
    return max(direta, por_tokens, contem)


def obter_python_console():
    exe = Path(sys.executable)
    if os.name == "nt" and exe.name.lower() == "pythonw.exe":
        console = exe.with_name("python.exe")
        if console.exists():
            return console
    return exe


def abrir_caminho(caminho):
    caminho = Path(caminho)
    if os.name == "nt":
        os.startfile(str(caminho))
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(caminho)])
    else:
        subprocess.Popen(["xdg-open", str(caminho)])


def _copiar_estilo_cabecalho(origem, destino):
    if origem is None:
        return
    try:
        from copy import copy
        destino.font = copy(origem.font)
        destino.fill = copy(origem.fill)
        destino.border = copy(origem.border)
        destino.alignment = copy(origem.alignment)
        destino.number_format = origem.number_format
    except Exception:
        destino.font = Font(bold=True, color="FFFFFF")
        destino.fill = PatternFill("solid", fgColor="1F4E78")
        destino.alignment = Alignment(horizontal="center", vertical="center")


def garantir_estrutura_base(caminho=ARQUIVO_REGRAS, criar_backup=False):
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Base de rotas não encontrada: {caminho}")

    if criar_backup:
        criar_backup_base("migracao")

    wb = load_workbook(caminho)
    alterou = False

    if "Regras" not in wb.sheetnames:
        ws = wb.create_sheet("Regras")
        ws.append(CAB_REGRAS)
        alterou = True
    else:
        ws = wb["Regras"]
        cab = [str(c.value or "").strip() for c in ws[1]]
        for nome in CAB_REGRAS:
            if nome not in cab:
                col = ws.max_column + 1
                cell = ws.cell(1, col, nome)
                _copiar_estilo_cabecalho(ws.cell(1, max(1, col - 1)), cell)
                cab.append(nome)
                alterou = True

    if "Aliases" not in wb.sheetnames:
        ws = wb.create_sheet("Aliases")
        ws.append(CAB_ALIASES)
        alterou = True
    else:
        ws = wb["Aliases"]
        cab = [str(c.value or "").strip() for c in ws[1]]
        for nome in CAB_ALIASES:
            if nome not in cab:
                col = ws.max_column + 1
                cell = ws.cell(1, col, nome)
                _copiar_estilo_cabecalho(ws.cell(1, max(1, col - 1)), cell)
                if nome == "Origem":
                    for linha in range(2, ws.max_row + 1):
                        ws.cell(linha, col, "AMBOS")
                cab.append(nome)
                alterou = True

    if "Histórico" not in wb.sheetnames:
        ws = wb.create_sheet("Histórico")
        ws.append(CAB_HISTORICO)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        alterou = True
    else:
        ws = wb["Histórico"]
        cab = [str(c.value or "").strip() for c in ws[1]]
        for nome in CAB_HISTORICO:
            if nome not in cab:
                col = ws.max_column + 1
                cell = ws.cell(1, col, nome)
                _copiar_estilo_cabecalho(ws.cell(1, max(1, col - 1)), cell)
                cab.append(nome)
                alterou = True

    if alterou:
        wb.save(caminho)
    else:
        wb.close()
    return alterou


def criar_backup_base(motivo="alteracao"):
    if not ARQUIVO_REGRAS.exists():
        return None
    PASTA_BACKUPS.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    destino = PASTA_BACKUPS / f"regras_roteirizacao_{motivo}_{stamp}.xlsx"
    shutil.copy2(ARQUIVO_REGRAS, destino)
    return destino


def _cabecalhos(ws):
    return {str(cell.value or "").strip(): idx for idx, cell in enumerate(ws[1], start=1)}


def _linha_dict(ws, row):
    cab = _cabecalhos(ws)
    return {nome: ws.cell(row, col).value for nome, col in cab.items()}


def _registrar_historico(wb, tipo, acao, origem="", cidade="", chave="", novo="", anterior="", observacao=""):
    if "Histórico" not in wb.sheetnames:
        ws = wb.create_sheet("Histórico")
        ws.append(CAB_HISTORICO)
    ws = wb["Histórico"]
    cab = _cabecalhos(ws)
    valores = {
        "Data/Hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "Tipo": tipo,
        "Ação": acao,
        "Origem": origem,
        "Cidade": cidade,
        "Chave": chave,
        "Valor novo": novo,
        "Valor anterior": anterior,
        "Observação": observacao,
        "Usuário": getpass.getuser(),
    }
    linha = ws.max_row + 1
    for nome, valor in valores.items():
        col = cab.get(nome)
        if col:
            ws.cell(linha, col, valor)


class RepositorioRotas:
    def __init__(self, caminho=ARQUIVO_REGRAS):
        self.caminho = Path(caminho)
        garantir_estrutura_base(self.caminho)

    def carregar(self):
        garantir_estrutura_base(self.caminho)
        wb = load_workbook(self.caminho, data_only=False)
        dados = {}
        for nome in ["Regras", "Aliases", "Técnicos", "Histórico"]:
            if nome not in wb.sheetnames:
                dados[nome] = []
                continue
            ws = wb[nome]
            cab = _cabecalhos(ws)
            linhas = []
            for row in range(2, ws.max_row + 1):
                registro = {k: ws.cell(row, c).value for k, c in cab.items()}
                if not any(v not in (None, "") for v in registro.values()):
                    continue
                registro["_linha"] = row
                linhas.append(registro)
            dados[nome] = linhas
        wb.close()
        return dados

    def tecnicos_ativos(self):
        dados = self.carregar()
        tecnicos = [
            str(r.get("Técnico") or "").strip()
            for r in dados.get("Técnicos", [])
            if ativo(r.get("Ativo")) and str(r.get("Técnico") or "").strip()
        ]
        if not tecnicos:
            tecnicos = sorted({
                str(r.get("Técnico") or "").strip()
                for r in dados.get("Regras", [])
                if str(r.get("Técnico") or "").strip()
            })
        return sorted(set(tecnicos))

    def salvar_regra(self, registro, linha=None):
        criar_backup_base("antes_regra")
        garantir_estrutura_base(self.caminho)
        wb = load_workbook(self.caminho)
        ws = wb["Regras"]
        cab = _cabecalhos(ws)

        cidade_norm = normalizar_texto(registro.get("Cidade"))
        bairro_norm = normalizar_texto(registro.get("Bairro / Localidade Normalizada"))
        origem = normalizar_texto(registro.get("Origem")) or "AMBOS"
        tipo = registro.get("Tipo de Regra") or "Cidade + bairro"

        # Evita duplicidade: atualiza uma regra com a mesma chave lógica.
        linha_encontrada = None
        for row in range(2, ws.max_row + 1):
            atual = _linha_dict(ws, row)
            if (
                normalizar_texto(atual.get("Cidade")) == cidade_norm
                and normalizar_texto(atual.get("Bairro / Localidade Normalizada")) == bairro_norm
                and normalizar_texto(atual.get("Origem")) == origem
                and normalizar_texto(atual.get("Tipo de Regra")) == normalizar_texto(tipo)
            ):
                linha_encontrada = row
                break

        destino = int(linha or linha_encontrada or (ws.max_row + 1))
        anterior = _linha_dict(ws, destino) if destino <= ws.max_row else {}
        registro = dict(registro)
        registro["Ativo"] = registro.get("Ativo") or "Sim"
        registro["Prioridade"] = registro.get("Prioridade") or (0 if origem != "AMBOS" else 1)
        registro["Cidade"] = str(registro.get("Cidade") or "").strip()
        registro["Bairro / Localidade Normalizada"] = bairro_norm
        registro["Origem"] = origem
        registro["Tipo de Regra"] = tipo
        registro["Regra Original"] = registro.get("Regra Original") or str(registro.get("Bairro / Localidade Normalizada") or "")

        for nome in CAB_REGRAS:
            col = cab.get(nome)
            if col:
                ws.cell(destino, col, registro.get(nome, ""))

        _registrar_historico(
            wb, "REGRA", "EDITAR" if anterior else "CRIAR", origem,
            registro.get("Cidade", ""), bairro_norm,
            registro.get("Técnico", ""), anterior.get("Técnico", "") if anterior else "",
            registro.get("Observação", ""),
        )
        wb.save(self.caminho)
        return destino

    def salvar_alias(self, registro, linha=None):
        criar_backup_base("antes_alias")
        garantir_estrutura_base(self.caminho)
        wb = load_workbook(self.caminho)
        ws = wb["Aliases"]
        cab = _cabecalhos(ws)

        cidade_norm = normalizar_texto(registro.get("Cidade"))
        recebido_norm = normalizar_texto(registro.get("Nome recebido"))
        origem = normalizar_texto(registro.get("Origem")) or "AMBOS"

        linha_encontrada = None
        for row in range(2, ws.max_row + 1):
            atual = _linha_dict(ws, row)
            if (
                normalizar_texto(atual.get("Cidade")) == cidade_norm
                and normalizar_texto(atual.get("Nome recebido")) == recebido_norm
                and (normalizar_texto(atual.get("Origem")) or "AMBOS") == origem
            ):
                linha_encontrada = row
                break

        destino = int(linha or linha_encontrada or (ws.max_row + 1))
        anterior = _linha_dict(ws, destino) if destino <= ws.max_row else {}
        registro = dict(registro)
        registro["Ativo"] = registro.get("Ativo") or "Sim"
        registro["Cidade"] = str(registro.get("Cidade") or "").strip()
        registro["Nome recebido"] = recebido_norm
        registro["Nome considerado"] = normalizar_texto(registro.get("Nome considerado"))
        registro["Origem"] = origem

        for nome in CAB_ALIASES:
            col = cab.get(nome)
            if col:
                ws.cell(destino, col, registro.get(nome, ""))

        _registrar_historico(
            wb, "ALIAS", "EDITAR" if anterior else "CRIAR", origem,
            registro.get("Cidade", ""), recebido_norm,
            registro.get("Nome considerado", ""), anterior.get("Nome considerado", "") if anterior else "",
            registro.get("Observação", ""),
        )
        wb.save(self.caminho)
        return destino

    def alternar_ativo(self, aba, linha):
        if aba not in {"Regras", "Aliases"}:
            return
        criar_backup_base("antes_status")
        wb = load_workbook(self.caminho)
        ws = wb[aba]
        cab = _cabecalhos(ws)
        col_ativo = cab.get("Ativo")
        if not col_ativo:
            wb.close()
            return
        anterior = str(ws.cell(linha, col_ativo).value or "")
        novo = "Não" if ativo(anterior) else "Sim"
        ws.cell(linha, col_ativo, novo)
        registro = _linha_dict(ws, linha)
        chave = registro.get("Bairro / Localidade Normalizada") or registro.get("Nome recebido") or ""
        _registrar_historico(
            wb, "REGRA" if aba == "Regras" else "ALIAS", "ATIVAR" if novo == "Sim" else "DESATIVAR",
            registro.get("Origem", ""), registro.get("Cidade", ""), chave, novo, anterior,
            "Alteração realizada pela Gestão de Rotas",
        )
        wb.save(self.caminho)


def ler_pendencias():
    colunas = [
        "Resultado", "Origem", "Cidade", "Bairro / Distrito", "Quantidade",
        "OSs", "Regra Aplicada", "Candidatos",
    ]
    if not ARQUIVO_ROTEIRIZACAO.exists():
        return pd.DataFrame(columns=colunas)

    partes = []
    for aba in ["Sem Rota", "Conflitos"]:
        try:
            df = pd.read_excel(ARQUIVO_ROTEIRIZACAO, sheet_name=aba, dtype=str).fillna("")
            if not df.empty:
                partes.append(df)
        except Exception:
            continue
    if not partes:
        return pd.DataFrame(columns=colunas)

    base = pd.concat(partes, ignore_index=True)
    for col in ["Resultado", "Origem", "Cidade", "Bairro / Distrito", "Regra Aplicada", "Candidatos", "OS"]:
        if col not in base.columns:
            base[col] = ""

    agrupado = (
        base.groupby(
            ["Resultado", "Origem", "Cidade", "Bairro / Distrito", "Regra Aplicada", "Candidatos"],
            dropna=False,
        )
        .agg(
            Quantidade=("OS", "size"),
            OSs=("OS", lambda x: ", ".join(dict.fromkeys(str(v) for v in x if str(v).strip()))),
        )
        .reset_index()
    )
    return agrupado[colunas]


def obter_resumo_rotas():
    resumo = {"sem_rota": 0, "conflitos": 0, "regras": 0, "aliases": 0}
    try:
        pend = ler_pendencias()
        if not pend.empty:
            resumo["sem_rota"] = int(pend.loc[pend["Resultado"] == "SEM ROTA", "Quantidade"].astype(int).sum())
            resumo["conflitos"] = int(pend.loc[pend["Resultado"] == "CONFLITO", "Quantidade"].astype(int).sum())
    except Exception:
        pass
    try:
        repo = RepositorioRotas()
        dados = repo.carregar()
        resumo["regras"] = sum(1 for r in dados.get("Regras", []) if ativo(r.get("Ativo")))
        resumo["aliases"] = sum(1 for r in dados.get("Aliases", []) if ativo(r.get("Ativo")))
    except Exception:
        pass
    return resumo


class GestaoRotasWindow:
    def __init__(self, parent, on_change=None):
        self.parent = parent
        self.on_change = on_change
        self.repo = RepositorioRotas()
        self.dados = {}
        self.pendencias = pd.DataFrame()
        self.sugestoes_atual = []
        self._processando = False

        self.win = tk.Toplevel(parent)
        self.win.title("Gestão Inteligente de Rotas — Central Visconde")
        self.win.configure(bg=COR_FUNDO)
        self.win.geometry("1260x790")
        self.win.minsize(1080, 680)
        self.win.transient(parent)

        self._configurar_estilo()
        self._montar()
        self.recarregar_tudo()

    def _configurar_estilo(self):
        style = ttk.Style(self.win)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Visconde.TNotebook", background=COR_FUNDO, borderwidth=0)
        style.configure(
            "Visconde.TNotebook.Tab", background=COR_CARD, foreground=COR_TEXTO,
            padding=(16, 9), font=("Arial", 10, "bold"),
        )
        style.map(
            "Visconde.TNotebook.Tab",
            background=[("selected", COR_DOURADO_ESCURO)],
            foreground=[("selected", COR_BRANCO)],
        )
        style.configure(
            "Visconde.Treeview", background="#111111", fieldbackground="#111111",
            foreground=COR_TEXTO, rowheight=27, bordercolor=COR_BORDA,
            font=("Arial", 9),
        )
        style.map("Visconde.Treeview", background=[("selected", "#5A4812")], foreground=[("selected", COR_BRANCO)])
        style.configure(
            "Visconde.Treeview.Heading", background="#252525", foreground=COR_DOURADO,
            font=("Arial", 9, "bold"), relief="flat",
        )
        style.map("Visconde.Treeview.Heading", background=[("active", "#303030")])
        style.configure("Visconde.TCombobox", fieldbackground="#1B1B1B", background="#1B1B1B", foreground="#111111")

    def _montar(self):
        topo = tk.Frame(self.win, bg=COR_FUNDO)
        topo.pack(fill="x", padx=22, pady=(18, 8))
        tk.Label(
            topo, text="GESTÃO INTELIGENTE DE ROTAS", bg=COR_FUNDO, fg=COR_DOURADO,
            font=("Arial", 20, "bold"), anchor="w",
        ).pack(side="left")
        self._botao(
            topo, "Atualizar dados", self.recarregar_tudo, "#2A2A2A", COR_BRANCO
        ).pack(side="right")

        tk.Label(
            self.win,
            text="Resolva bairros, aliases, conflitos e regras sem abrir a planilha de base.",
            bg=COR_FUNDO, fg=COR_TEXTO_FRACO, font=("Arial", 10), anchor="w",
        ).pack(fill="x", padx=22, pady=(0, 10))

        self.notebook = ttk.Notebook(self.win, style="Visconde.TNotebook")
        self.notebook.pack(fill="both", expand=True, padx=22, pady=(0, 14))

        self.tab_pendencias = tk.Frame(self.notebook, bg=COR_FUNDO_2)
        self.tab_regras = tk.Frame(self.notebook, bg=COR_FUNDO_2)
        self.tab_aliases = tk.Frame(self.notebook, bg=COR_FUNDO_2)
        self.tab_historico = tk.Frame(self.notebook, bg=COR_FUNDO_2)
        self.notebook.add(self.tab_pendencias, text="Pendências")
        self.notebook.add(self.tab_regras, text="Regras")
        self.notebook.add(self.tab_aliases, text="Aliases")
        self.notebook.add(self.tab_historico, text="Histórico")

        self._montar_pendencias()
        self._montar_regras()
        self._montar_aliases()
        self._montar_historico()

        rodape = tk.Frame(self.win, bg=COR_FUNDO)
        rodape.pack(fill="x", padx=22, pady=(0, 14))
        self.status = tk.Label(
            rodape, text="Aguardando ação.", bg=COR_FUNDO, fg=COR_TEXTO_FRACO,
            font=("Arial", 9), anchor="w",
        )
        self.status.pack(side="left", fill="x", expand=True)
        self._botao(
            rodape, "Abrir base de regras", lambda: abrir_caminho(ARQUIVO_REGRAS), "#2A2A2A", COR_BRANCO
        ).pack(side="right", padx=(6, 0))
        self._botao(
            rodape, "Abrir roteirização", self._abrir_roteirizacao, "#2A2A2A", COR_BRANCO
        ).pack(side="right")

    def _tree(self, parent, columns, widths):
        container = tk.Frame(parent, bg=COR_CARD, highlightbackground=COR_BORDA, highlightthickness=1)
        tree = ttk.Treeview(container, columns=columns, show="headings", style="Visconde.Treeview")
        for col, width in zip(columns, widths):
            tree.heading(col, text=col)
            tree.column(col, width=width, minwidth=70, anchor="w")
        sy = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
        sx = ttk.Scrollbar(container, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        return container, tree

    def _botao(self, parent, texto, comando, cor=COR_DOURADO, fg="#111111", width=None):
        return BotaoVisconde(parent, texto, comando, cor, fg, width=width)

    def _label(self, parent, texto):
        return tk.Label(parent, text=texto, bg=COR_CARD, fg=COR_TEXTO, font=("Arial", 9), anchor="w")

    def _montar_pendencias(self):
        paned = tk.PanedWindow(self.tab_pendencias, orient="horizontal", bg=COR_FUNDO_2, sashwidth=6, bd=0)
        paned.pack(fill="both", expand=True, padx=12, pady=12)

        esquerda = tk.Frame(paned, bg=COR_FUNDO_2)
        direita = tk.Frame(paned, bg=COR_CARD, highlightbackground=COR_BORDA, highlightthickness=1)
        paned.add(esquerda, minsize=650)
        paned.add(direita, minsize=350)

        barra = tk.Frame(esquerda, bg=COR_FUNDO_2)
        barra.pack(fill="x", pady=(0, 8))
        tk.Label(barra, text="Casos agrupados do último roteiro", bg=COR_FUNDO_2, fg=COR_BRANCO, font=("Arial", 11, "bold")).pack(side="left")
        self.lbl_qtd_pendencias = tk.Label(barra, text="0 casos", bg=COR_FUNDO_2, fg=COR_DOURADO, font=("Arial", 10, "bold"))
        self.lbl_qtd_pendencias.pack(side="right")

        cols = ["Resultado", "Origem", "Cidade", "Bairro", "Qtd.", "Sugestão", "%"]
        cont, self.tree_pendencias = self._tree(esquerda, cols, [95, 75, 125, 190, 52, 190, 48])
        cont.pack(fill="both", expand=True)
        self.tree_pendencias.bind("<<TreeviewSelect>>", self._selecionar_pendencia)

        tk.Label(direita, text="Resolver pendência", bg=COR_CARD, fg=COR_DOURADO, font=("Arial", 14, "bold"), anchor="w").pack(fill="x", padx=16, pady=(16, 8))
        self.lbl_detalhes = tk.Label(
            direita, text="Selecione um caso à esquerda.", bg=COR_CARD, fg=COR_TEXTO,
            font=("Arial", 9), justify="left", anchor="nw", wraplength=360,
        )
        self.lbl_detalhes.pack(fill="x", padx=16, pady=(0, 12))

        self._label(direita, "Aplicar para").pack(fill="x", padx=16)
        self.combo_origem = ttk.Combobox(direita, values=["MOBYAN", "OGEA", "AMBOS"], state="readonly", style="Visconde.TCombobox")
        self.combo_origem.pack(fill="x", padx=16, pady=(3, 10))

        self._label(direita, "Possível bairro existente").pack(fill="x", padx=16)
        self.combo_sugestao = ttk.Combobox(direita, state="readonly", style="Visconde.TCombobox")
        self.combo_sugestao.pack(fill="x", padx=16, pady=(3, 10))
        self.combo_sugestao.bind("<<ComboboxSelected>>", self._sugestao_mudou)

        self._label(direita, "Técnico responsável").pack(fill="x", padx=16)
        self.combo_tecnico = ttk.Combobox(direita, state="readonly", style="Visconde.TCombobox")
        self.combo_tecnico.pack(fill="x", padx=16, pady=(3, 12))

        self.btn_salvar_alias = self._botao(
            direita, "Salvar como alias e gerar novamente", self.salvar_alias_pendencia, "#1F6698", COR_BRANCO
        )
        self.btn_salvar_alias.pack(fill="x", padx=16, pady=4)
        self.btn_cadastrar_rota = self._botao(
            direita, "Cadastrar rota e gerar novamente", self.salvar_rota_pendencia, "#E0B21B", "#111111"
        )
        self.btn_cadastrar_rota.pack(fill="x", padx=16, pady=4)
        self.btn_aplicar_temporario = self._botao(
            direita, "Aplicar somente neste roteiro", self.aplicar_temporariamente, "#B76517", COR_BRANCO
        )
        self.btn_aplicar_temporario.pack(fill="x", padx=16, pady=4)

        tk.Label(
            direita,
            text="A Central cria um backup antes de cada alteração. Casos repetidos são resolvidos de uma só vez.",
            bg=COR_CARD, fg=COR_TEXTO_FRACO, font=("Arial", 8), justify="left", wraplength=360,
        ).pack(fill="x", padx=16, pady=(12, 16))

    def _montar_regras(self):
        topo = tk.Frame(self.tab_regras, bg=COR_FUNDO_2)
        topo.pack(fill="x", padx=12, pady=(12, 8))
        tk.Label(topo, text="Regras territoriais", bg=COR_FUNDO_2, fg=COR_BRANCO, font=("Arial", 11, "bold")).pack(side="left")
        self.var_filtro_regras = tk.StringVar()
        entrada = tk.Entry(topo, textvariable=self.var_filtro_regras, bg="#1B1B1B", fg=COR_BRANCO, insertbackground=COR_DOURADO, relief="flat")
        entrada.pack(side="left", fill="x", expand=True, padx=12, ipady=6)
        entrada.bind("<KeyRelease>", lambda e: self._carregar_tree_regras())
        self._botao(topo, "Nova regra", lambda: self._dialog_regra(), COR_DOURADO).pack(side="right")

        cols = ["Ativo", "Prior.", "Técnico", "Tipo", "Cidade", "Bairro / Localidade", "Origem", "Observação"]
        cont, self.tree_regras = self._tree(self.tab_regras, cols, [55, 55, 100, 120, 125, 220, 70, 260])
        cont.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        self.tree_regras.bind("<Double-1>", lambda e: self.editar_regra())

        botoes = tk.Frame(self.tab_regras, bg=COR_FUNDO_2)
        botoes.pack(fill="x", padx=12, pady=(0, 12))
        self._botao(botoes, "Editar selecionada", self.editar_regra, COR_AZUL, COR_BRANCO).pack(side="left")
        self._botao(botoes, "Ativar / desativar", lambda: self.alternar_item("Regras"), COR_LARANJA, COR_BRANCO).pack(side="left", padx=8)

    def _montar_aliases(self):
        topo = tk.Frame(self.tab_aliases, bg=COR_FUNDO_2)
        topo.pack(fill="x", padx=12, pady=(12, 8))
        tk.Label(topo, text="Aliases de bairros e localidades", bg=COR_FUNDO_2, fg=COR_BRANCO, font=("Arial", 11, "bold")).pack(side="left")
        self.var_filtro_aliases = tk.StringVar()
        entrada = tk.Entry(topo, textvariable=self.var_filtro_aliases, bg="#1B1B1B", fg=COR_BRANCO, insertbackground=COR_DOURADO, relief="flat")
        entrada.pack(side="left", fill="x", expand=True, padx=12, ipady=6)
        entrada.bind("<KeyRelease>", lambda e: self._carregar_tree_aliases())
        self._botao(topo, "Novo alias", lambda: self._dialog_alias(), COR_DOURADO).pack(side="right")

        cols = ["Ativo", "Cidade", "Nome recebido", "Nome considerado", "Técnico", "Origem", "Observação"]
        cont, self.tree_aliases = self._tree(self.tab_aliases, cols, [55, 130, 190, 190, 120, 75, 270])
        cont.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        self.tree_aliases.bind("<Double-1>", lambda e: self.editar_alias())

        botoes = tk.Frame(self.tab_aliases, bg=COR_FUNDO_2)
        botoes.pack(fill="x", padx=12, pady=(0, 12))
        self._botao(botoes, "Editar selecionado", self.editar_alias, COR_AZUL, COR_BRANCO).pack(side="left")
        self._botao(botoes, "Ativar / desativar", lambda: self.alternar_item("Aliases"), COR_LARANJA, COR_BRANCO).pack(side="left", padx=8)

    def _montar_historico(self):
        topo = tk.Frame(self.tab_historico, bg=COR_FUNDO_2)
        topo.pack(fill="x", padx=12, pady=(12, 8))
        tk.Label(topo, text="Últimas alterações realizadas pela Central", bg=COR_FUNDO_2, fg=COR_BRANCO, font=("Arial", 11, "bold")).pack(side="left")
        self._botao(topo, "Abrir pasta de backups", lambda: abrir_caminho(PASTA_BACKUPS), COR_CARD, COR_BRANCO).pack(side="right")
        cols = ["Data/Hora", "Tipo", "Ação", "Origem", "Cidade", "Chave", "Valor novo", "Valor anterior", "Usuário"]
        cont, self.tree_historico = self._tree(self.tab_historico, cols, [130, 70, 80, 70, 120, 190, 150, 150, 90])
        cont.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    def recarregar_tudo(self):
        try:
            self.dados = self.repo.carregar()
            self.pendencias = ler_pendencias()
            self.combo_tecnico["values"] = self.repo.tecnicos_ativos()
            self._carregar_tree_pendencias()
            self._carregar_tree_regras()
            self._carregar_tree_aliases()
            self._carregar_tree_historico()
            self.status.config(text="Dados atualizados.", fg=COR_VERDE)
            if self.on_change:
                self.on_change()
        except Exception as erro:
            messagebox.showerror("Gestão de Rotas", f"Não consegui carregar os dados:\n\n{erro}", parent=self.win)

    def _limpar_tree(self, tree):
        for item in tree.get_children():
            tree.delete(item)

    def _origem_compativel(self, regra_origem, origem):
        regra = normalizar_texto(regra_origem) or "AMBOS"
        origem = normalizar_texto(origem)
        return regra in {"AMBOS", origem}

    def _sugerir(self, cidade, bairro, origem):
        cidade_n = normalizar_texto(cidade)
        bairro_n = normalizar_texto(bairro)
        candidatos = {}
        for regra in self.dados.get("Regras", []):
            if not ativo(regra.get("Ativo")):
                continue
            if normalizar_texto(regra.get("Cidade")) != cidade_n:
                continue
            if not self._origem_compativel(regra.get("Origem"), origem):
                continue
            alvo = normalizar_texto(regra.get("Bairro / Localidade Normalizada"))
            if not alvo:
                continue
            score = similaridade(bairro_n, alvo)
            tecnico = str(regra.get("Técnico") or "").strip()
            anterior = candidatos.get(alvo)
            if anterior is None or score > anterior[0]:
                candidatos[alvo] = (score, tecnico)
        lista = [(alvo, score, tecnico) for alvo, (score, tecnico) in candidatos.items()]
        lista.sort(key=lambda x: x[1], reverse=True)
        return lista[:8]

    def _carregar_tree_pendencias(self):
        self._limpar_tree(self.tree_pendencias)
        if self.pendencias.empty:
            self.lbl_qtd_pendencias.config(text="0 casos")
            return
        total = int(self.pendencias["Quantidade"].astype(int).sum())
        self.lbl_qtd_pendencias.config(text=f"{len(self.pendencias)} casos • {total} OSs")
        for idx, row in self.pendencias.reset_index(drop=True).iterrows():
            sugestoes = self._sugerir(row["Cidade"], row["Bairro / Distrito"], row["Origem"])
            sug = sugestoes[0][0] if sugestoes else ""
            pct = f"{int(sugestoes[0][1] * 100)}%" if sugestoes else ""
            tag = "erro" if row["Resultado"] == "SEM ROTA" else "conflito"
            self.tree_pendencias.insert(
                "", "end", iid=str(idx),
                values=(row["Resultado"], row["Origem"], row["Cidade"], row["Bairro / Distrito"], row["Quantidade"], sug, pct),
                tags=(tag,),
            )
        self.tree_pendencias.tag_configure("erro", foreground="#F19A9A")
        self.tree_pendencias.tag_configure("conflito", foreground="#F4C76D")

        itens = self.tree_pendencias.get_children()
        if itens:
            primeiro = itens[0]
            self.tree_pendencias.selection_set(primeiro)
            self.tree_pendencias.focus(primeiro)
            self.tree_pendencias.see(primeiro)
            self._selecionar_pendencia()

    def _selecionar_pendencia(self, event=None):
        sel = self.tree_pendencias.selection()
        if not sel:
            return
        row = self.pendencias.reset_index(drop=True).iloc[int(sel[0])]
        self.pendencia_atual = row.to_dict()
        self.combo_origem.set(str(row["Origem"] or "AMBOS"))
        self.lbl_detalhes.config(
            text=(
                f"Resultado: {row['Resultado']}\n"
                f"Cidade: {row['Cidade']}\n"
                f"Bairro recebido: {row['Bairro / Distrito']}\n"
                f"OSs afetadas: {row['Quantidade']}\n"
                f"Números: {row['OSs']}\n"
                f"Motivo: {row['Regra Aplicada']}"
                + (f"\nCandidatos: {row['Candidatos']}" if row.get("Candidatos") else "")
            )
        )
        self.sugestoes_atual = self._sugerir(row["Cidade"], row["Bairro / Distrito"], row["Origem"])
        valores = [f"{alvo} | {int(score * 100)}% | {tecnico or 'sem técnico'}" for alvo, score, tecnico in self.sugestoes_atual]
        self.combo_sugestao["values"] = valores
        if valores:
            self.combo_sugestao.current(0)
            self._sugestao_mudou()
        else:
            self.combo_sugestao.set("")
            self.combo_tecnico.set("")

    def _sugestao_mudou(self, event=None):
        idx = self.combo_sugestao.current()
        if idx >= 0 and idx < len(self.sugestoes_atual):
            tecnico = self.sugestoes_atual[idx][2]
            if tecnico:
                self.combo_tecnico.set(tecnico)

    def _obter_pendencia(self):
        if not hasattr(self, "pendencia_atual"):
            messagebox.showwarning("Gestão de Rotas", "Selecione uma pendência primeiro.", parent=self.win)
            return None
        return self.pendencia_atual

    def salvar_alias_pendencia(self):
        pend = self._obter_pendencia()
        if not pend:
            return
        idx = self.combo_sugestao.current()
        if idx < 0 or idx >= len(self.sugestoes_atual):
            messagebox.showwarning("Alias", "Selecione o bairro existente que representa esse nome.", parent=self.win)
            return
        alvo, score, tecnico_sug = self.sugestoes_atual[idx]
        origem = self.combo_origem.get() or pend["Origem"] or "AMBOS"
        registro = {
            "Ativo": "Sim",
            "Cidade": pend["Cidade"],
            "Nome recebido": pend["Bairro / Distrito"],
            "Nome considerado": alvo,
            "Técnico": self.combo_tecnico.get() or tecnico_sug,
            "Observação": f"Criado pela Central a partir de sugestão de {int(score * 100)}%",
            "Origem": origem,
        }
        if not messagebox.askyesno(
            "Confirmar alias",
            f"Salvar o alias abaixo?\n\n{pend['Bairro / Distrito']} → {alvo}\nCidade: {pend['Cidade']}\nOrigem: {origem}\n\nDepois a Central gerará novamente o roteiro.",
            parent=self.win,
        ):
            return
        try:
            self.repo.salvar_alias(registro)
            self._reprocessar(False)
        except Exception as erro:
            messagebox.showerror("Alias", f"Não consegui salvar o alias:\n\n{erro}", parent=self.win)

    def salvar_rota_pendencia(self):
        pend = self._obter_pendencia()
        if not pend:
            return
        tecnico = self.combo_tecnico.get().strip()
        if not tecnico:
            messagebox.showwarning("Nova rota", "Selecione o técnico responsável.", parent=self.win)
            return
        origem = self.combo_origem.get() or pend["Origem"] or "AMBOS"
        registro = {
            "Ativo": "Sim",
            "Prioridade": 0 if origem != "AMBOS" else 1,
            "Técnico": tecnico,
            "Tipo de Regra": "Cidade + bairro",
            "Cidade": pend["Cidade"],
            "Bairro / Localidade Normalizada": pend["Bairro / Distrito"],
            "Origem": origem,
            "Regra Original": pend["Bairro / Distrito"],
            "Observação": "Cadastrada pela Gestão Inteligente de Rotas",
        }
        if not messagebox.askyesno(
            "Confirmar nova rota",
            f"Cadastrar esta rota?\n\n{pend['Cidade']} + {pend['Bairro / Distrito']}\nTécnico: {tecnico}\nOrigem: {origem}\n\nDepois a Central gerará novamente o roteiro.",
            parent=self.win,
        ):
            return
        try:
            self.repo.salvar_regra(registro)
            self._reprocessar(False)
        except Exception as erro:
            messagebox.showerror("Nova rota", f"Não consegui salvar a rota:\n\n{erro}", parent=self.win)

    def aplicar_temporariamente(self):
        pend = self._obter_pendencia()
        if not pend:
            return
        tecnico = self.combo_tecnico.get().strip()
        if not tecnico:
            messagebox.showwarning("Resolução temporária", "Selecione o técnico que receberá essas OSs.", parent=self.win)
            return
        origem = pend["Origem"]
        registro = {
            "origem": normalizar_texto(origem),
            "cidade": normalizar_texto(pend["Cidade"]),
            "bairro": normalizar_texto(pend["Bairro / Distrito"]),
            "tecnico": tecnico,
        }
        ARQUIVO_TEMPORARIO.parent.mkdir(parents=True, exist_ok=True)
        existentes = []
        if ARQUIVO_TEMPORARIO.exists():
            try:
                existentes = json.loads(ARQUIVO_TEMPORARIO.read_text(encoding="utf-8"))
            except Exception:
                existentes = []
        existentes = [
            item for item in existentes
            if not (
                item.get("origem") == registro["origem"]
                and item.get("cidade") == registro["cidade"]
                and item.get("bairro") == registro["bairro"]
            )
        ]
        existentes.append(registro)
        ARQUIVO_TEMPORARIO.write_text(json.dumps(existentes, ensure_ascii=False, indent=2), encoding="utf-8")
        self._reprocessar(True)

    def _reprocessar(self, usar_temporarias):
        if self._processando:
            return
        if not SCRIPT_ROTEIRIZACAO.exists():
            messagebox.showerror("Roteirização", f"Script não encontrado:\n{SCRIPT_ROTEIRIZACAO}", parent=self.win)
            return
        self._processando = True
        self.status.config(text="Gerando novamente o roteiro com as alterações...", fg=COR_DOURADO)
        progresso = tk.Toplevel(self.win)
        progresso.title("Gerando roteiro")
        progresso.geometry("700x390")
        progresso.configure(bg=COR_FUNDO)
        progresso.transient(self.win)
        texto = tk.Text(progresso, bg="#050505", fg=COR_TEXTO, insertbackground=COR_DOURADO, font=(("Consolas" if os.name == "nt" else "Menlo"), 9), wrap="word")
        texto.pack(fill="both", expand=True, padx=12, pady=12)
        texto.insert("end", "Reprocessando os relatórios que já foram baixados...\n\n")
        texto.config(state="disabled")

        def log(linha):
            def inserir():
                texto.config(state="normal")
                texto.insert("end", linha)
                texto.see("end")
                texto.config(state="disabled")
            try:
                progresso.after(0, inserir)
            except Exception:
                pass

        def executar():
            if FROZEN:
                comando = [
                    str(obter_python_console()), "--rodar-automacao", SCRIPT_ROTEIRIZACAO.stem,
                    "--reprocessar-atual", "--sem-abrir",
                ]
            else:
                comando = [str(obter_python_console()), str(SCRIPT_ROTEIRIZACAO), "--reprocessar-atual", "--sem-abrir"]
            if usar_temporarias:
                comando.append("--usar-resolucoes-temporarias")
            kwargs = {}
            if os.name == "nt":
                kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
            linhas_log = []
            try:
                processo = subprocess.Popen(
                    comando, cwd=str(BASE_DIR), stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, encoding="utf-8", errors="replace", bufsize=1, **kwargs,
                )
                for linha in processo.stdout:
                    linhas_log.append(linha.rstrip())
                    log(linha)
                processo.wait()
                sucesso = processo.returncode == 0
                if sucesso:
                    erro = ""
                else:
                    detalhes = "\n".join(linhas_log[-18:]).strip()
                    erro = f"O processo terminou com código {processo.returncode}."
                    if detalhes:
                        erro += f"\n\nDetalhes finais do processamento:\n{detalhes}"
            except Exception as exc:
                sucesso = False
                erro = str(exc)

            def finalizar():
                self._processando = False
                if sucesso:
                    self.status.config(text="Roteiro atualizado com sucesso.", fg=COR_VERDE)
                    self.recarregar_tudo()
                    restantes = int(len(self.pendencias)) if not self.pendencias.empty else 0
                    mensagem = "Alteração salva e roteiro gerado novamente com sucesso."
                    if restantes == 0:
                        mensagem += "\n\nTodas as pendências de rota foram resolvidas."
                    else:
                        mensagem += f"\n\nAinda existem {restantes} pendência(s) agrupada(s) para revisar."
                    messagebox.showinfo("Gestão de Rotas", mensagem, parent=self.win)
                    try:
                        progresso.destroy()
                    except Exception:
                        pass
                else:
                    self.status.config(text="Erro ao gerar novamente o roteiro.", fg=COR_VERMELHO)
                    messagebox.showerror("Gestão de Rotas", f"Não consegui gerar o roteiro novamente.\n\n{erro}\n\nVeja o log da janela.", parent=progresso)
            try:
                progresso.after(0, finalizar)
            except Exception:
                pass

        threading.Thread(target=executar, daemon=True).start()

    def _carregar_tree_regras(self):
        self._limpar_tree(self.tree_regras)
        filtro = normalizar_texto(self.var_filtro_regras.get())
        for r in self.dados.get("Regras", []):
            texto = normalizar_texto(" ".join(str(r.get(k) or "") for k in CAB_REGRAS))
            if filtro and filtro not in texto:
                continue
            self.tree_regras.insert(
                "", "end", iid=str(r["_linha"]),
                values=(
                    r.get("Ativo", ""), r.get("Prioridade", ""), r.get("Técnico", ""),
                    r.get("Tipo de Regra", ""), r.get("Cidade", ""),
                    r.get("Bairro / Localidade Normalizada", ""), r.get("Origem", ""), r.get("Observação", ""),
                ),
            )

    def _carregar_tree_aliases(self):
        self._limpar_tree(self.tree_aliases)
        filtro = normalizar_texto(self.var_filtro_aliases.get())
        for r in self.dados.get("Aliases", []):
            texto = normalizar_texto(" ".join(str(r.get(k) or "") for k in CAB_ALIASES))
            if filtro and filtro not in texto:
                continue
            self.tree_aliases.insert(
                "", "end", iid=str(r["_linha"]),
                values=(
                    r.get("Ativo", ""), r.get("Cidade", ""), r.get("Nome recebido", ""),
                    r.get("Nome considerado", ""), r.get("Técnico", ""), r.get("Origem", "AMBOS"), r.get("Observação", ""),
                ),
            )

    def _carregar_tree_historico(self):
        self._limpar_tree(self.tree_historico)
        historico = list(reversed(self.dados.get("Histórico", [])))[:500]
        for r in historico:
            self.tree_historico.insert(
                "", "end",
                values=(
                    r.get("Data/Hora", ""), r.get("Tipo", ""), r.get("Ação", ""), r.get("Origem", ""),
                    r.get("Cidade", ""), r.get("Chave", ""), r.get("Valor novo", ""),
                    r.get("Valor anterior", ""), r.get("Usuário", ""),
                ),
            )

    def _form_dialog(self, titulo, campos, valores, ao_salvar):
        janela = tk.Toplevel(self.win)
        janela.title(titulo)
        janela.configure(bg=COR_FUNDO)
        janela.geometry("590x610")
        janela.transient(self.win)
        janela.grab_set()
        vars_ = {}
        frame = tk.Frame(janela, bg=COR_FUNDO)
        frame.pack(fill="both", expand=True, padx=24, pady=20)
        tk.Label(frame, text=titulo, bg=COR_FUNDO, fg=COR_DOURADO, font=("Arial", 16, "bold")).pack(anchor="w", pady=(0, 14))
        for nome, tipo, opcoes in campos:
            tk.Label(frame, text=nome, bg=COR_FUNDO, fg=COR_TEXTO, font=("Arial", 9)).pack(fill="x", pady=(6, 2))
            var = tk.StringVar(value=str(valores.get(nome, "") or ""))
            vars_[nome] = var
            if tipo == "combo":
                widget = ttk.Combobox(frame, textvariable=var, values=opcoes, state="readonly", style="Visconde.TCombobox")
            else:
                widget = tk.Entry(frame, textvariable=var, bg="#1B1B1B", fg=COR_BRANCO, insertbackground=COR_DOURADO, relief="flat")
            widget.pack(fill="x", ipady=5)

        def salvar():
            dados = {nome: var.get().strip() for nome, var in vars_.items()}
            try:
                ao_salvar(dados)
                janela.destroy()
                self.recarregar_tudo()
            except Exception as erro:
                messagebox.showerror(titulo, str(erro), parent=janela)

        botoes = tk.Frame(frame, bg=COR_FUNDO)
        botoes.pack(fill="x", pady=(18, 0))
        self._botao(botoes, "Salvar", salvar, COR_DOURADO).pack(side="right")
        self._botao(botoes, "Cancelar", janela.destroy, COR_CARD, COR_BRANCO).pack(side="right", padx=8)

    def _dialog_regra(self, registro=None):
        registro = registro or {}
        campos = [
            ("Ativo", "combo", ["Sim", "Não"]),
            ("Prioridade", "text", None),
            ("Técnico", "combo", self.repo.tecnicos_ativos()),
            ("Tipo de Regra", "combo", ["Cidade + bairro", "Cidade inteira"]),
            ("Cidade", "text", None),
            ("Bairro / Localidade Normalizada", "text", None),
            ("Origem", "combo", ["AMBOS", "MOBYAN", "OGEA"]),
            ("Regra Original", "text", None),
            ("Observação", "text", None),
        ]
        valores = {nome: registro.get(nome, "") for nome, _, _ in campos}
        valores.setdefault("Ativo", "Sim")
        valores.setdefault("Prioridade", "1")
        valores.setdefault("Origem", "AMBOS")
        linha = registro.get("_linha")

        def salvar(dados):
            if not dados["Técnico"] or not dados["Cidade"]:
                raise ValueError("Informe pelo menos a cidade e o técnico.")
            if dados["Tipo de Regra"] != "Cidade inteira" and not dados["Bairro / Localidade Normalizada"]:
                raise ValueError("Informe o bairro/localidade ou selecione Cidade inteira.")
            try:
                dados["Prioridade"] = int(dados["Prioridade"] or 1)
            except ValueError:
                raise ValueError("A prioridade deve ser um número inteiro.")
            self.repo.salvar_regra(dados, linha=linha)

        self._form_dialog("Regra de roteirização", campos, valores, salvar)

    def _dialog_alias(self, registro=None):
        registro = registro or {}
        campos = [
            ("Ativo", "combo", ["Sim", "Não"]),
            ("Cidade", "text", None),
            ("Nome recebido", "text", None),
            ("Nome considerado", "text", None),
            ("Técnico", "combo", [""] + self.repo.tecnicos_ativos()),
            ("Origem", "combo", ["AMBOS", "MOBYAN", "OGEA"]),
            ("Observação", "text", None),
        ]
        valores = {nome: registro.get(nome, "") for nome, _, _ in campos}
        valores.setdefault("Ativo", "Sim")
        valores.setdefault("Origem", "AMBOS")
        linha = registro.get("_linha")

        def salvar(dados):
            if not dados["Cidade"] or not dados["Nome recebido"] or not dados["Nome considerado"]:
                raise ValueError("Informe cidade, nome recebido e nome considerado.")
            self.repo.salvar_alias(dados, linha=linha)

        self._form_dialog("Alias de bairro/localidade", campos, valores, salvar)

    def editar_regra(self):
        sel = self.tree_regras.selection()
        if not sel:
            messagebox.showwarning("Regras", "Selecione uma regra.", parent=self.win)
            return
        linha = int(sel[0])
        registro = next((r for r in self.dados.get("Regras", []) if r.get("_linha") == linha), None)
        if registro:
            self._dialog_regra(registro)

    def editar_alias(self):
        sel = self.tree_aliases.selection()
        if not sel:
            messagebox.showwarning("Aliases", "Selecione um alias.", parent=self.win)
            return
        linha = int(sel[0])
        registro = next((r for r in self.dados.get("Aliases", []) if r.get("_linha") == linha), None)
        if registro:
            self._dialog_alias(registro)

    def alternar_item(self, aba):
        tree = self.tree_regras if aba == "Regras" else self.tree_aliases
        sel = tree.selection()
        if not sel:
            messagebox.showwarning(aba, "Selecione um item.", parent=self.win)
            return
        try:
            self.repo.alternar_ativo(aba, int(sel[0]))
            self.recarregar_tudo()
        except Exception as erro:
            messagebox.showerror(aba, f"Não consegui alterar o status:\n\n{erro}", parent=self.win)

    def _abrir_roteirizacao(self):
        if ARQUIVO_ROTEIRIZACAO.exists():
            abrir_caminho(ARQUIVO_ROTEIRIZACAO)
        else:
            messagebox.showwarning("Roteirização", "Ainda não existe uma roteirização atual.", parent=self.win)


def abrir_gestao_rotas(parent, on_change=None):
    return GestaoRotasWindow(parent, on_change=on_change)
