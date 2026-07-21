"""Tela de configuração dos contatos de WhatsApp por base (prestador).

Antes disso, a única forma de preencher bases/contatos_prestadores.xlsx era
rodando configurar_contatos_prestadores.py no terminal — que só funciona com
Python instalado. Numa instalação empacotada (sem Python), não havia nenhuma
forma de cadastrar contato nenhum, e por isso "Enviar WhatsApp" ficava sempre
vazio numa conta nova. Esta tela substitui esse script pra quem usa o
executável.
"""

import os
import re
from pathlib import Path

import tkinter as tk
from tkinter import messagebox, ttk

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from caminho_base import BASE_DIR
from estilo_visconde import (
    COR_BORDA,
    COR_BRANCO,
    COR_CARD,
    COR_DOURADO,
    COR_FUNDO,
    COR_LARANJA,
    COR_TEXTO_FRACO,
    COR_TEXTO_SECUNDARIO,
)
from gestao_rotas import BotaoVisconde, normalizar_texto

load_dotenv(BASE_DIR / ".env")

ARQUIVO_CONTATOS = BASE_DIR / "bases" / "contatos_prestadores.xlsx"
CAB_CONTATOS = ["Prestador", "Responsável", "WhatsApp", "Enviar", "Observação"]

# Preservada só como fallback pra instalação atual (RS-SMART); contas novas
# configuram a lista real de bases via MOBYAN_PRESTADORES (mesma variável que
# exportador_mobyan.py já usa).
_PRESTADORES_PADRAO = [
    "RS-SMART",
    "RS-SMART - CAXIAS DO SUL",
    "RS-SMART - CAPAO DA CANOA",
    "RS-SMART - PASSO FUNDO",
    "RS-SMART - PELOTAS",
    "RS-SMART - SANTA MARIA",
    "RS-SMART - SANTA CRUZ DO SUL",
    "RS-SMART - SANTA VITORIA DO",
    "RS-SMART - SANTANA DO",
    "RS-SMART - SANTO ANGELO",
    "RS-SMART - URUGUAIANA",
    "RS-SMART - VALE DOS SINOS",
    "RS-SMART TAPES",
]


def listar_prestadores():
    valor = os.getenv("MOBYAN_PRESTADORES", "").strip()
    if not valor:
        return list(_PRESTADORES_PADRAO)
    return [item.strip() for item in valor.split(",") if item.strip()]


def normalizar_whatsapp(valor):
    digitos = re.sub(r"\D", "", str(valor or ""))
    # Aceita com ou sem o código do país (55) na frente — o envio real
    # (enviar_whatsapp.py) precisa do número completo com código do país pra
    # abrir a conversa certa no WhatsApp Web, então sempre normaliza pra esse
    # formato final, adicionando o 55 automaticamente se faltar.
    if len(digitos) in (10, 11):
        digitos = "55" + digitos
    if len(digitos) not in (12, 13) or not digitos.startswith("55"):
        raise ValueError(
            "WhatsApp inválido — informe DDD + número, com ou sem o 55 na frente, "
            "ex.: 51999998888 ou 5551999998888."
        )
    return digitos


def formatar_whatsapp(valor):
    digitos = re.sub(r"\D", "", str(valor or ""))
    resto = digitos[2:] if digitos.startswith("55") and len(digitos) in (12, 13) else digitos
    if len(resto) == 11:
        return f"+55 ({resto[:2]}) {resto[2:7]}-{resto[7:]}"
    if len(resto) == 10:
        return f"+55 ({resto[:2]}) {resto[2:6]}-{resto[6:]}"
    return digitos


def _cabecalhos(ws):
    return {str(cell.value or "").strip(): idx for idx, cell in enumerate(ws[1], start=1)}


def garantir_estrutura_contatos(caminho=ARQUIVO_CONTATOS):
    caminho = Path(caminho)
    if not caminho.exists():
        caminho.parent.mkdir(parents=True, exist_ok=True)
        Workbook().save(caminho)

    wb = load_workbook(caminho)
    alterou = False

    if "Contatos" not in wb.sheetnames:
        ws = wb.create_sheet("Contatos")
        ws.append(CAB_CONTATOS)
        alterou = True
    else:
        ws = wb["Contatos"]
        cab = [str(c.value or "").strip() for c in ws[1]]
        for nome in CAB_CONTATOS:
            if nome not in cab:
                ws.cell(1, ws.max_column + 1, nome)
                cab.append(nome)
                alterou = True

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])
        alterou = True

    if alterou:
        wb.save(caminho)
    else:
        wb.close()
    return alterou


class RepositorioContatos:
    def __init__(self, caminho=ARQUIVO_CONTATOS):
        self.caminho = Path(caminho)
        garantir_estrutura_contatos(self.caminho)

    def carregar(self):
        garantir_estrutura_contatos(self.caminho)
        wb = load_workbook(self.caminho)
        ws = wb["Contatos"]
        cab = _cabecalhos(ws)
        linhas = []
        for row in range(2, ws.max_row + 1):
            registro = {nome: ws.cell(row, col).value for nome, col in cab.items()}
            if not any(v not in (None, "") for v in registro.values()):
                continue
            registro["_linha"] = row
            linhas.append(registro)
        wb.close()
        return linhas

    def salvar_contato(self, registro, linha=None):
        garantir_estrutura_contatos(self.caminho)
        wb = load_workbook(self.caminho)
        ws = wb["Contatos"]
        cab = _cabecalhos(ws)

        prestador_norm = normalizar_texto(registro.get("Prestador"))
        linha_encontrada = None
        for row in range(2, ws.max_row + 1):
            if normalizar_texto(ws.cell(row, cab["Prestador"]).value) == prestador_norm:
                linha_encontrada = row
                break

        destino = int(linha or linha_encontrada or (ws.max_row + 1))
        registro = dict(registro)
        registro["Enviar"] = registro.get("Enviar") or "Sim"
        registro["WhatsApp"] = normalizar_whatsapp(registro.get("WhatsApp"))

        for nome in CAB_CONTATOS:
            col = cab.get(nome)
            if col:
                ws.cell(destino, col, registro.get(nome, ""))

        wb.save(self.caminho)
        return destino

    def alternar_enviar(self, linha):
        wb = load_workbook(self.caminho)
        ws = wb["Contatos"]
        cab = _cabecalhos(ws)
        col = cab.get("Enviar")
        if not col:
            wb.close()
            return
        anterior = str(ws.cell(linha, col).value or "")
        novo = "Não" if normalizar_texto(anterior) in {"SIM", "S", "YES", "Y"} else "Sim"
        ws.cell(linha, col, novo)
        wb.save(self.caminho)


class JanelaConfigurarWhatsapp:
    def __init__(self, parent):
        self.repo = RepositorioContatos()
        self.dados = []

        self.win = tk.Toplevel(parent)
        self.win.title("Configurar WhatsApp — ViscondeApp")
        self.win.configure(bg=COR_FUNDO)
        self.win.geometry("920x560")
        self.win.minsize(760, 460)
        self.win.transient(parent)

        self._configurar_estilo()
        self._montar()
        self.recarregar()

    def _configurar_estilo(self):
        style = ttk.Style(self.win)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure(
            "Visconde.Treeview", background="#111111", fieldbackground="#111111",
            foreground=COR_TEXTO_SECUNDARIO, rowheight=27, bordercolor=COR_BORDA,
            font=("Arial", 9),
        )
        style.map("Visconde.Treeview", background=[("selected", "#5A4812")], foreground=[("selected", COR_BRANCO)])
        style.configure(
            "Visconde.Treeview.Heading", background="#252525", foreground=COR_DOURADO,
            font=("Arial", 9, "bold"), relief="flat",
        )
        style.configure("Visconde.TCombobox", fieldbackground="#1B1B1B", background="#1B1B1B", foreground=COR_BRANCO)
        style.map(
            "Visconde.TCombobox",
            fieldbackground=[("readonly", "#1B1B1B")],
            foreground=[("readonly", COR_BRANCO)],
        )

    def _botao(self, parent, texto, comando, cor, fg=COR_BRANCO):
        return BotaoVisconde(parent, texto, comando, cor, fg)

    def _montar(self):
        topo = tk.Frame(self.win, bg=COR_FUNDO)
        topo.pack(fill="x", padx=22, pady=(18, 8))
        tk.Label(
            topo, text="CONFIGURAR WHATSAPP", bg=COR_FUNDO, fg=COR_DOURADO,
            font=("Arial", 18, "bold"), anchor="w",
        ).pack(side="left")
        self._botao(topo, "Novo contato", lambda: self._dialog_contato(), COR_DOURADO).pack(side="right")

        tk.Label(
            self.win,
            text="Cadastre o WhatsApp de quem recebe as pendências de cada base.",
            bg=COR_FUNDO, fg=COR_TEXTO_FRACO, font=("Arial", 10), anchor="w",
        ).pack(fill="x", padx=22, pady=(0, 10))

        cols = ["Prestador", "Responsável", "WhatsApp", "Enviar", "Observação"]
        container = tk.Frame(self.win, bg=COR_CARD, highlightbackground=COR_BORDA, highlightthickness=1)
        container.pack(fill="both", expand=True, padx=22, pady=(0, 14))
        self.tree = ttk.Treeview(container, columns=cols, show="headings", style="Visconde.Treeview")
        for col, largura in zip(cols, [280, 200, 150, 80, 260]):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=largura, minwidth=70, anchor="w")
        sy = ttk.Scrollbar(container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sy.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        self.tree.bind("<Double-1>", lambda e: self.editar_selecionado())

        botoes = tk.Frame(self.win, bg=COR_FUNDO)
        botoes.pack(fill="x", padx=22, pady=(0, 18))
        self._botao(botoes, "Editar selecionado", self.editar_selecionado, "#2A2A2A", COR_BRANCO).pack(side="left")
        self._botao(botoes, "Ativar / desativar envio", self.alternar_selecionado, COR_LARANJA, COR_BRANCO).pack(side="left", padx=8)

    def recarregar(self):
        self.dados = self.repo.carregar()
        for item in self.tree.get_children():
            self.tree.delete(item)
        for r in self.dados:
            self.tree.insert(
                "", "end", iid=str(r["_linha"]),
                values=(
                    r.get("Prestador", ""), r.get("Responsável", ""),
                    formatar_whatsapp(r.get("WhatsApp", "")), r.get("Enviar", ""),
                    r.get("Observação", ""),
                ),
            )

    def editar_selecionado(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("WhatsApp", "Selecione um contato.", parent=self.win)
            return
        linha = int(sel[0])
        registro = next((r for r in self.dados if r.get("_linha") == linha), None)
        if registro:
            self._dialog_contato(registro)

    def alternar_selecionado(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("WhatsApp", "Selecione um contato.", parent=self.win)
            return
        self.repo.alternar_enviar(int(sel[0]))
        self.recarregar()

    def _dialog_contato(self, registro=None):
        registro = registro or {}
        janela = tk.Toplevel(self.win)
        janela.title("Contato de WhatsApp")
        janela.configure(bg=COR_FUNDO)
        janela.geometry("480x480")
        janela.transient(self.win)
        janela.grab_set()

        frame = tk.Frame(janela, bg=COR_FUNDO)
        frame.pack(fill="both", expand=True, padx=24, pady=20)
        tk.Label(
            frame, text="Contato de WhatsApp", bg=COR_FUNDO, fg=COR_DOURADO,
            font=("Arial", 16, "bold"),
        ).pack(anchor="w", pady=(0, 14))

        def rotulo(texto):
            tk.Label(frame, text=texto, bg=COR_FUNDO, fg=COR_TEXTO_SECUNDARIO, font=("Arial", 9)).pack(fill="x", pady=(6, 2))

        def campo_texto(valor_inicial):
            var = tk.StringVar(value=valor_inicial)
            tk.Entry(
                frame, textvariable=var, bg="#1B1B1B", fg=COR_BRANCO,
                insertbackground=COR_DOURADO, relief="flat",
            ).pack(fill="x", ipady=5)
            return var

        rotulo("Base (prestador)")
        var_prestador = tk.StringVar(value=str(registro.get("Prestador", "") or ""))
        combo_prestador = ttk.Combobox(
            frame, textvariable=var_prestador, values=listar_prestadores(),
            state="readonly", style="Visconde.TCombobox",
        )
        combo_prestador.pack(fill="x", ipady=5)

        rotulo("Responsável")
        var_responsavel = campo_texto(str(registro.get("Responsável", "") or ""))

        rotulo("WhatsApp (com DDD)")
        valor_whatsapp = registro.get("WhatsApp", "")
        var_whatsapp = campo_texto(formatar_whatsapp(valor_whatsapp) if valor_whatsapp else "")
        tk.Label(
            frame, text="Ex.: 51 99999-8888 — só números com DDD também funciona.",
            bg=COR_FUNDO, fg=COR_TEXTO_FRACO, font=("Arial", 8),
        ).pack(fill="x", pady=(2, 0))

        rotulo("Enviar")
        var_enviar = tk.StringVar(value=str(registro.get("Enviar", "") or "Sim"))
        ttk.Combobox(
            frame, textvariable=var_enviar, values=["Sim", "Não"],
            state="readonly", style="Visconde.TCombobox",
        ).pack(fill="x", ipady=5)

        rotulo("Observação")
        var_obs = campo_texto(str(registro.get("Observação", "") or ""))

        linha = registro.get("_linha")

        def salvar():
            if not var_prestador.get().strip():
                messagebox.showerror("Contato de WhatsApp", "Selecione a base (prestador).", parent=janela)
                return
            dados = {
                "Prestador": var_prestador.get().strip(),
                "Responsável": var_responsavel.get().strip(),
                "WhatsApp": var_whatsapp.get().strip(),
                "Enviar": var_enviar.get().strip() or "Sim",
                "Observação": var_obs.get().strip(),
            }
            try:
                self.repo.salvar_contato(dados, linha=linha)
            except ValueError as erro:
                messagebox.showerror("Contato de WhatsApp", str(erro), parent=janela)
                return
            janela.destroy()
            self.recarregar()

        botoes = tk.Frame(frame, bg=COR_FUNDO)
        botoes.pack(fill="x", pady=(18, 0))
        self._botao(botoes, "Salvar", salvar, COR_DOURADO).pack(side="right")
        self._botao(botoes, "Cancelar", janela.destroy, "#2A2A2A", COR_BRANCO).pack(side="right", padx=8)


def abrir_configurar_whatsapp(parent):
    return JanelaConfigurarWhatsapp(parent)
