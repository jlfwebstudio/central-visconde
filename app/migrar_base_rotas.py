import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent.parent
ARQUIVO = BASE_DIR / "bases" / "regras_roteirizacao.xlsx"
PASTA_BACKUPS = BASE_DIR / "bases" / "backups_roteirizacao"
CAB_HISTORICO = [
    "Data/Hora", "Tipo", "Ação", "Origem", "Cidade", "Chave",
    "Valor novo", "Valor anterior", "Observação", "Usuário",
]


def copiar_estilo(origem, destino):
    try:
        from copy import copy
        destino.font = copy(origem.font)
        destino.fill = copy(origem.fill)
        destino.border = copy(origem.border)
        destino.alignment = copy(origem.alignment)
        destino.number_format = origem.number_format
    except Exception:
        destino.font = Font(bold=True, color="FFFFFF")
        destino.fill = PatternFill("solid", fgColor="1F4E78")
        destino.alignment = Alignment(horizontal="center", vertical="center")


def main():
    if not ARQUIVO.exists():
        print(f"ERRO: base não encontrada: {ARQUIVO}")
        return 1

    PASTA_BACKUPS.mkdir(parents=True, exist_ok=True)
    backup = PASTA_BACKUPS / f"regras_roteirizacao_antes_gestao_rotas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(ARQUIVO, backup)
    print(f"Backup criado: {backup}")

    wb = load_workbook(ARQUIVO)
    alteracoes = []

    if "Aliases" not in wb.sheetnames:
        ws = wb.create_sheet("Aliases")
        ws.append(["Ativo", "Cidade", "Nome recebido", "Nome considerado", "Técnico", "Observação", "Origem"])
        alteracoes.append("Aba Aliases criada")
    else:
        ws = wb["Aliases"]
        cab = [str(c.value or "").strip() for c in ws[1]]
        if "Origem" not in cab:
            col = ws.max_column + 1
            destino = ws.cell(1, col, "Origem")
            copiar_estilo(ws.cell(1, max(1, col - 1)), destino)
            for linha in range(2, ws.max_row + 1):
                if any(ws.cell(linha, c).value not in (None, "") for c in range(1, col)):
                    ws.cell(linha, col, "AMBOS")
            alteracoes.append("Coluna Origem adicionada à aba Aliases")

    if "Histórico" not in wb.sheetnames:
        ws = wb.create_sheet("Histórico")
        ws.append(CAB_HISTORICO)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        larguras = [20, 12, 14, 12, 20, 30, 25, 25, 45, 18]
        for idx, largura in enumerate(larguras, start=1):
            ws.column_dimensions[chr(64 + idx)].width = largura
        alteracoes.append("Aba Histórico criada")

    if "Leia-me" in wb.sheetnames:
        ws = wb["Leia-me"]
        textos = [str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1)]
        aviso = "A manutenção de bairros, aliases e regras agora pode ser feita pela Gestão Inteligente de Rotas da Central Visconde."
        if aviso not in textos:
            ws.cell(ws.max_row + 2, 1, "Gestão pela Central")
            ws.cell(ws.max_row + 1, 1, aviso)
            alteracoes.append("Leia-me atualizado")

    try:
        wb.save(ARQUIVO)
    except PermissionError:
        print("ERRO: feche a planilha regras_roteirizacao.xlsx e execute o instalador novamente.")
        return 2

    if alteracoes:
        print("Migração concluída:")
        for item in alteracoes:
            print(f"- {item}")
    else:
        print("A base já estava atualizada. Nenhuma mudança estrutural foi necessária.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
