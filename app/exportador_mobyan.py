import os
import re
import shutil
import subprocess
import sys
from datetime import datetime, date
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageDraw, ImageFont
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

from caminho_base import BASE_DIR
load_dotenv(BASE_DIR / ".env")

MOBYAN_URL = os.getenv("MOBYAN_URL")
MOBYAN_USUARIO = os.getenv("MOBYAN_USUARIO")
MOBYAN_SENHA = os.getenv("MOBYAN_SENHA")


def _lista_a_partir_do_env(nome_var, padrao):
    valor = os.getenv(nome_var, "").strip()
    if not valor:
        return padrao
    return [item.strip() for item in valor.split(",") if item.strip()]


# Lista padrão preservada para não quebrar a instalação atual (RS-SMART): contas
# novas configuram os próprios prestadores/estado via MOBYAN_PRESTADORES/MOBYAN_ESTADO.
_PRESTADORES_RS_SMART_PADRAO = [
    "RS-SMART",
    "RS-SMART - CAXIAS DO SUL",
    "RS-SMART - CAPAO DA CANOA",
    "RS-SMART - PASSO FUNDO",
    "RS-SMART - PELOTAS",
    "RS-SMART - SANTA MARIA",
    "RS-SMART - SANTA CRUZ DO SUL",
    "RS-SMART - SANTA VITORIA DO",
    "RS-SMART - SANTANA DO",
    "RS-SMART - SANTO ANGELO",
    "RS-SMART - URUGUAIANA",
    "RS-SMART - VALE DOS SINOS",
    "RS-SMART TAPES",
]

MOBYAN_PRESTADORES = _lista_a_partir_do_env("MOBYAN_PRESTADORES", _PRESTADORES_RS_SMART_PADRAO)
MOBYAN_ESTADO = os.getenv("MOBYAN_ESTADO", "RS").strip() or "RS"

PASTA_DOWNLOADS = BASE_DIR / "downloads" / "relatorios_completos"
PASTA_PENDENCIAS = BASE_DIR / "outputs" / "pendencias_do_dia"
PASTA_IMAGENS_PRESTADOR = BASE_DIR / "outputs" / "por_prestador_imagens"
PASTA_PRESTADOR_ANTIGA = BASE_DIR / "outputs" / "por_prestador"
PASTA_LOGS = BASE_DIR / "logs"

ARQUIVO_BASE_JUSTIFICATIVAS = BASE_DIR / "bases" / "base_justificativas.xlsx"
ARQUIVO_CONTATOS_PRESTADORES = BASE_DIR / "bases" / "contatos_prestadores.xlsx"
ARQUIVO_FINAL_PENDENCIAS = PASTA_PENDENCIAS / "pendencias_do_dia_atual.xlsx"
PASTA_BACKUPS_PENDENCIAS = PASTA_PENDENCIAS / "backups"
ARQUIVO_BACKUP_ULTIMO = PASTA_BACKUPS_PENDENCIAS / "pendencias_do_dia_backup_ultimo.xlsx"

COLUNAS_IMAGEM = [
    "SITUAÇÃO",
    "Chamado",
    "Numero Referencia",
    "Contratante",
    "Serviço",
    "Status",
    "Data Limite",
    "Cliente",
    "CNPJ / CPF",
    "Cidade",
]

LARGURAS_IMAGEM = {
    "SITUAÇÃO": 110,
    "Chamado": 110,
    "Numero Referencia": 150,
    "Contratante": 120,
    "Serviço": 180,
    "Status": 150,
    "Data Limite": 170,
    "Cliente": 300,
    "CNPJ / CPF": 170,
    "Cidade": 180,
}

ALTURA_CABECALHO_IMAGEM = 34
ALTURA_LINHA_IMAGEM = 32
MARGEM_IMAGEM = 16

COR_CABECALHO = "#1F4E78"
COR_TEXTO_CABECALHO = "#FFFFFF"
COR_LINHA_VENCIDA = "#F4CCCC"
COR_LINHA_HOJE = "#FFF2CC"
COR_LINHA_FUTURA = "#D9EAD3"
COR_SITUACAO_VENCIDA = "#CC0000"
COR_SITUACAO_HOJE = "#F1C232"
COR_SITUACAO_FUTURA = "#70AD47"
COR_TEXTO = "#000000"
COR_TEXTO_BRANCO = "#FFFFFF"
COR_BORDA = "#D9E2F3"
COR_FUNDO = "#FFFFFF"


def abrir_caminho(caminho):
    caminho = Path(caminho).resolve()

    if os.name == "nt":
        os.startfile(str(caminho))
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(caminho)])
    else:
        subprocess.Popen(["xdg-open", str(caminho)])


def validar_configuracoes():
    if not MOBYAN_URL:
        raise ValueError("MOBYAN_URL não foi encontrada no arquivo .env")
    if not MOBYAN_USUARIO:
        raise ValueError("MOBYAN_USUARIO não foi encontrado no arquivo .env")
    if not MOBYAN_SENHA:
        raise ValueError("MOBYAN_SENHA não foi encontrada no arquivo .env")


def fazer_login(page):
    print("Verificando tela de login...")

    try:
        campo_usuario = page.get_by_role("textbox", name="Usuário:")
        campo_senha = page.get_by_role("textbox", name="Senha:")

        campo_usuario.wait_for(timeout=10000)

        print("Preenchendo usuário e senha...")
        campo_usuario.fill(MOBYAN_USUARIO)
        campo_senha.fill(MOBYAN_SENHA)

        try:
            page.get_by_role("button", name="Entrar").click(timeout=3000)
        except Exception:
            campo_senha.press("Enter")

        print("Login enviado.")
        page.wait_for_timeout(3000)

    except PlaywrightTimeoutError:
        print("Tela de login não apareceu. Continuando...")


def finalizar_sessao_anterior_se_aparecer(page):
    try:
        print("Verificando se existe sessão anterior em uso...")

        botao_yes = page.get_by_role("button", name="Yes")

        if botao_yes.is_visible(timeout=5000):
            print("Sessão anterior encontrada. Clicando em Yes...")
            botao_yes.click()
            page.wait_for_timeout(3000)

            print("Após clicar em Yes, tentando fazer login novamente se necessário...")
            fazer_login(page)
        else:
            print("Nenhuma sessão anterior encontrada.")

    except Exception:
        print("Tela de sessão anterior não apareceu. Continuando...")


def abrir_relatorio_ordem_servico_sla(page):
    print("Abrindo OPPAY MENU...")

    PASTA_LOGS.mkdir(parents=True, exist_ok=True)

    page.wait_for_timeout(3000)
    page.mouse.click(55, 748)
    page.wait_for_timeout(1500)

    print("Procurando menu Relatórios...")

    try:
        page.get_by_text("Relatórios", exact=True).click(timeout=4000)
        print("Menu Relatórios aberto.")
    except Exception:
        caminho_print = PASTA_LOGS / "erro_menu_relatorios.png"
        page.screenshot(path=str(caminho_print), full_page=True)
        raise Exception(
            f"Falha ao abrir o menu de relatórios. Screenshot salvo em: {caminho_print}"
        )

    page.wait_for_timeout(1000)

    print("Abrindo Relatórios Ordem de Serviço - SLA...")

    try:
        page.get_by_text(
            "Relatórios Ordem de Serviço - SLA",
            exact=True
        ).click(timeout=5000)

        print("Relatório Ordem de Serviço - SLA aberto.")

    except Exception:
        try:
            page.get_by_text("Ordem de Serviço - SLA").click(timeout=5000)
            print("Relatório Ordem de Serviço - SLA aberto.")

        except Exception:
            caminho_print = PASTA_LOGS / "erro_item_relatorio_sla.png"
            page.screenshot(path=str(caminho_print), full_page=True)
            raise Exception(
                f"Falha ao abrir Relatórios Ordem de Serviço - SLA. Screenshot salvo em: {caminho_print}"
            )


def selecionar_status(frame_relatorio):
    print("Selecionando status...")

    frame_relatorio.get_by_role(
        "cell",
        name="Status Visualize a situação"
    ).get_by_role("button").click()

    frame_relatorio.locator("label").filter(has_text="EM CAMPO").click()
    frame_relatorio.locator("span").filter(has_text="EM TRANSFERÊNCIA").click()
    frame_relatorio.get_by_role("checkbox", name="ENCAMINHADA").check()
    frame_relatorio.locator("span").filter(has_text="PROCED. TÉCNICO").click()
    frame_relatorio.locator("span").filter(has_text="RECEBIDO PA").click()
    frame_relatorio.locator("label").filter(has_text="REENCAMINHADO").click()
    frame_relatorio.locator("label").filter(has_text="ROLLBACK").click()

    frame_relatorio.locator("html").click()


def selecionar_prestadores(frame_relatorio, prestadores=None):
    prestadores = prestadores if prestadores is not None else MOBYAN_PRESTADORES

    if not prestadores:
        print("Nenhum prestador configurado para filtrar — pulando seleção de prestador.")
        return

    principal, *filiais = prestadores

    print(f"Selecionando prestadores: {', '.join(prestadores)}")

    frame_relatorio.get_by_role("cell", name="Prestador").get_by_role("button").click()

    try:
        frame_relatorio.get_by_role(
            "checkbox",
            name=principal,
            exact=True
        ).check(timeout=3000)

        print(f"Prestador marcado: {principal}")

    except Exception:
        try:
            frame_relatorio.locator("label").filter(
                has_text=re.compile(rf"^{re.escape(principal)}$")
            ).click(timeout=3000)

            print(f"Prestador marcado por label: {principal}")

        except Exception as erro:
            print(f"Aviso: não consegui marcar o prestador {principal}. Erro: {erro}")

    for prestador in filiais:
        try:
            frame_relatorio.get_by_role(
                "checkbox",
                name=prestador
            ).check(timeout=3000)

            print(f"Prestador marcado: {prestador}")

        except PlaywrightTimeoutError:
            print(f"Aviso: não encontrei o checkbox do prestador: {prestador}")

        except Exception as erro:
            print(f"Aviso: não consegui marcar o prestador {prestador}. Erro: {erro}")

    frame_relatorio.locator("html").click()


def selecionar_estado_rs(frame_relatorio, estado=None):
    estado = estado if estado is not None else MOBYAN_ESTADO
    print(f"Selecionando estado {estado}...")

    frame_relatorio.locator("#TD_STATE_NAME").get_by_role("button").filter(
        has_text=re.compile(r"^$")
    ).click()

    frame_relatorio.get_by_role("checkbox", name=estado).check()
    frame_relatorio.locator("html").click()


def ler_relatorio_csv(caminho_csv):
    tentativas = [
        {"encoding": "utf-8-sig", "sep": ";"},
        {"encoding": "latin1", "sep": ";"},
        {"encoding": "cp1252", "sep": ";"},
        {"encoding": "utf-8-sig", "sep": ","},
        {"encoding": "latin1", "sep": ","},
        {"encoding": "cp1252", "sep": ","},
    ]

    ultimo_erro = None

    for tentativa in tentativas:
        try:
            df = pd.read_csv(caminho_csv, dtype=str, **tentativa)

            if len(df.columns) > 1:
                return df

        except Exception as erro:
            ultimo_erro = erro

    raise Exception(f"Não consegui ler o CSV exportado. Último erro: {ultimo_erro}")


def encontrar_coluna_data_limite(df):
    colunas = list(df.columns)

    for coluna in colunas:
        nome_normalizado = str(coluna).strip().lower()

        if "data" in nome_normalizado and "limite" in nome_normalizado:
            return coluna

    raise Exception(
        "Não encontrei a coluna de Data Limite no relatório. "
        f"Colunas encontradas: {colunas}"
    )


def encontrar_coluna_por_alias(df, aliases):
    colunas = list(df.columns)

    for alias in aliases:
        alias_normalizado = alias.strip().lower()

        for coluna in colunas:
            coluna_normalizada = str(coluna).strip().lower()

            if coluna_normalizada == alias_normalizado:
                return coluna

    return None


def padronizar_colunas_pendencias(df):
    mapa_colunas = {
        "Chamado": [
            "Chamado",
            "OS",
            "Ordem de Serviço",
            "Ordem de Servico",
            "Numero OS",
            "Número OS",
            "Nº OS",
            "N OS",
        ],
        "Numero Referencia": [
            "Numero Referencia",
            "Número Referência",
            "Referencia",
            "Referência",
            "Nº Referência",
            "N Referencia",
        ],
        "Contratante": ["Contratante"],
        "Serviço": ["Serviço", "Servico", "Tipo Serviço", "Tipo Servico"],
        "Status": ["Status", "Situação", "Situacao"],
        "Data Limite": ["Data Limite", "Data Limite SLA", "Limite", "Prazo"],
        "Cliente": [
            "Cliente",
            "Nome Cliente",
            "Nome do Cliente",
            "Cliente Nome",
            "Razão Social",
            "Razao Social",
            "Nome Fantasia",
            "Estabelecimento",
            "EC",
            "Nome EC",
            "Nome do EC",
            "Ponto de Atendimento",
            "PA",
            "Nome PA",
        ],
        "CNPJ / CPF": [
            "CNPJ / CPF",
            "CNPJ/CPF",
            "CNPJ CPF",
            "CPF / CNPJ",
            "CPF/CNPJ",
            "Documento",
            "Documento Cliente",
            "CPF",
            "CNPJ",
        ],
        "Cidade": ["Cidade", "Município", "Municipio"],
        "Técnico": ["Técnico", "Tecnico", "Nome Técnico", "Nome Tecnico"],
        "Prestador": ["Prestador", "Base", "Fornecedor"],
        "Justificativa do Abono": [
            "Justificativa do Abono",
            "Justificativa Abono",
            "Abono",
            "Motivo Abono",
            "Motivo do Abono",
        ],
    }

    for coluna_padrao, aliases in mapa_colunas.items():
        coluna_encontrada = encontrar_coluna_por_alias(df, aliases)

        if coluna_encontrada and coluna_encontrada != coluna_padrao:
            df[coluna_padrao] = df[coluna_encontrada]

    return df


def normalizar_chamado(valor):
    if pd.isna(valor):
        return ""

    texto = str(valor).strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    texto = re.sub(r"\D", "", texto)

    return texto


def valor_vazio(valor):
    if pd.isna(valor):
        return True

    texto = str(valor).strip()

    if texto == "":
        return True

    if texto.lower() in ["nan", "none", "null", "nat"]:
        return True

    return False


def limpar_valor(valor):
    if pd.isna(valor):
        return ""

    texto = str(valor).strip()

    if texto.lower() in ["nan", "none", "null", "nat"]:
        return ""

    if texto.startswith('="') and texto.endswith('"'):
        texto = texto[2:-1].strip()

    if texto.startswith('"') and texto.endswith('"'):
        texto = texto[1:-1].strip()

    return texto


def limpar_nome_arquivo(texto):
    texto = "" if texto is None else str(texto).strip()

    substituicoes = {
        "/": "-",
        "\\": "-",
        ":": "-",
        "*": "",
        "?": "",
        '"': "",
        "<": "",
        ">": "",
        "|": "",
    }

    for antigo, novo in substituicoes.items():
        texto = texto.replace(antigo, novo)

    texto = re.sub(r"\s+", " ", texto).strip()

    if texto == "":
        texto = "SEM_PRESTADOR"

    return texto


def ler_base_justificativas():
    if not ARQUIVO_BASE_JUSTIFICATIVAS.exists():
        print("Base de justificativas não encontrada. Continuando sem cruzamento.")
        return {}

    try:
        df_base = pd.read_excel(
            ARQUIVO_BASE_JUSTIFICATIVAS,
            sheet_name="Justificativas",
            dtype=str,
            keep_default_na=False,
            na_filter=False,
        )
    except Exception:
        try:
            df_base = pd.read_excel(
                ARQUIVO_BASE_JUSTIFICATIVAS,
                dtype=str,
                keep_default_na=False,
                na_filter=False,
            )
        except Exception as erro:
            print(f"Não consegui ler a base de justificativas. Erro: {erro}")
            return {}

    colunas_obrigatorias = ["Chamado", "Motivo", "Observação", "Validado"]

    for coluna in colunas_obrigatorias:
        if coluna not in df_base.columns:
            print(f"Base de justificativas sem coluna obrigatória: {coluna}")
            return {}

    df_base["Chamado_Normalizado"] = df_base["Chamado"].apply(normalizar_chamado)

    df_base["Validado_Normalizado"] = (
        df_base["Validado"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.lower()
    )

    df_base = df_base[
        (df_base["Chamado_Normalizado"] != "")
        & (df_base["Validado_Normalizado"].isin(["sim", "s", "yes", "y"]))
    ].copy()

    df_base = df_base.drop_duplicates(subset=["Chamado_Normalizado"], keep="last")

    base = {}

    for _, linha in df_base.iterrows():
        chamado = linha["Chamado_Normalizado"]

        base[chamado] = {
            "motivo": "" if pd.isna(linha.get("Motivo")) else str(linha.get("Motivo")).strip(),
            "observacao": "" if pd.isna(linha.get("Observação")) else str(linha.get("Observação")).strip(),
            "prestador": "" if pd.isna(linha.get("Prestador")) else str(linha.get("Prestador")).strip(),
            "origem": "" if pd.isna(linha.get("Origem")) else str(linha.get("Origem")).strip(),
            "data_registro": "" if pd.isna(linha.get("Data Registro")) else str(linha.get("Data Registro")).strip(),
        }

    print(f"Base de justificativas carregada: {len(base)} chamado(s) validado(s).")

    return base


def aplicar_justificativas_base(df, base_justificativas):
    df["Motivo Base"] = ""
    df["Observação Base"] = ""

    for indice, linha in df.iterrows():
        chamado = normalizar_chamado(linha.get("Chamado", ""))
        dados_base = base_justificativas.get(chamado)

        if dados_base:
            motivo_base = dados_base.get("motivo", "")
            observacao_base = dados_base.get("observacao", "")

            # Importante:
            # Motivo Base e Observação Base são informações internas da planilha de justificativas.
            # Elas NÃO devem sobrescrever a coluna Justificativa do Abono, que representa
            # somente o que veio do sistema Mobyan.
            df.at[indice, "Motivo Base"] = motivo_base
            df.at[indice, "Observação Base"] = observacao_base

    return df

def formatar_data_registro_base(valor):
    texto = limpar_valor(valor)

    if texto == "":
        return date.today().strftime("%d/%m/%Y")

    match_iso = re.search(r"(\d{4})-(\d{2})-(\d{2})", texto)
    if match_iso:
        ano, mes, dia = match_iso.groups()
        return f"{dia}/{mes}/{ano}"

    match_br = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", texto)
    if match_br:
        dia, mes, ano = match_br.groups()
        return f"{int(dia):02d}/{int(mes):02d}/{ano}"

    try:
        data = pd.to_datetime(texto, dayfirst=True, errors="coerce")
        if pd.notna(data):
            return data.strftime("%d/%m/%Y")
    except Exception:
        pass

    return date.today().strftime("%d/%m/%Y")


def atualizar_base_justificativas_com_pendencias(df_pendencias):
    print("Atualizando base de justificativas com as OSs pendentes atuais...")

    ARQUIVO_BASE_JUSTIFICATIVAS.parent.mkdir(parents=True, exist_ok=True)

    colunas_base = [
        "Chamado",
        "Prestador",
        "Cliente",
        "Cidade",
        "Data Limite",
        "Situação",
        "Motivo",
        "Observação",
        "Data Registro",
        "Origem",
        "Validado",
    ]

    hoje_texto = date.today().strftime("%d/%m/%Y")

    if ARQUIVO_BASE_JUSTIFICATIVAS.exists():
        try:
            df_base_antiga = pd.read_excel(
                ARQUIVO_BASE_JUSTIFICATIVAS,
                sheet_name="Justificativas",
                dtype=str,
                keep_default_na=False,
                na_filter=False,
            )
        except Exception:
            try:
                df_base_antiga = pd.read_excel(
                    ARQUIVO_BASE_JUSTIFICATIVAS,
                    dtype=str,
                    keep_default_na=False,
                    na_filter=False,
                )
            except Exception as erro:
                print(f"Aviso: não consegui ler a base antiga de justificativas. Erro: {erro}")
                df_base_antiga = pd.DataFrame(columns=colunas_base)
    else:
        df_base_antiga = pd.DataFrame(columns=colunas_base)

    for coluna in colunas_base:
        if coluna not in df_base_antiga.columns:
            df_base_antiga[coluna] = ""

    for coluna in df_base_antiga.columns:
        df_base_antiga[coluna] = (
            df_base_antiga[coluna]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace(["nan", "NaN", "None", "NULL", "null", "NaT"], "")
        )

    df_base_antiga["Chamado_Normalizado"] = df_base_antiga["Chamado"].apply(normalizar_chamado)
    df_base_antiga = df_base_antiga.drop_duplicates(
        subset=["Chamado_Normalizado"],
        keep="last"
    )

    justificativas_antigas = {}

    for _, linha in df_base_antiga.iterrows():
        chamado_normalizado = linha.get("Chamado_Normalizado", "")

        if chamado_normalizado == "":
            continue

        justificativas_antigas[chamado_normalizado] = {
            "Motivo": limpar_valor(linha.get("Motivo", "")),
            "Observação": limpar_valor(linha.get("Observação", "")),
            "Data Registro": formatar_data_registro_base(linha.get("Data Registro", "")),
            "Origem": limpar_valor(linha.get("Origem", "")),
            "Validado": limpar_valor(linha.get("Validado", "")),
        }

    novas_linhas = []
    chamados_adicionados = set()

    for _, linha in df_pendencias.iterrows():
        chamado = limpar_valor(linha.get("Chamado", ""))
        chamado_normalizado = normalizar_chamado(chamado)

        if chamado_normalizado == "":
            continue

        if chamado_normalizado in chamados_adicionados:
            continue

        chamados_adicionados.add(chamado_normalizado)

        dados_antigos = justificativas_antigas.get(chamado_normalizado, {})

        motivo = dados_antigos.get("Motivo", "")
        observacao = dados_antigos.get("Observação", "")
        data_registro = dados_antigos.get("Data Registro", "")
        origem = dados_antigos.get("Origem", "")
        validado = dados_antigos.get("Validado", "")

        if valor_vazio(data_registro):
            data_registro = hoje_texto
        else:
            data_registro = formatar_data_registro_base(data_registro)

        if valor_vazio(origem):
            origem = "Manual"

        if valor_vazio(validado):
            validado = "Não"

        novas_linhas.append({
            "Chamado": chamado,
            "Prestador": limpar_valor(linha.get("Prestador", "")),
            "Cliente": limpar_valor(linha.get("Cliente", "")),
            "Cidade": limpar_valor(linha.get("Cidade", "")),
            "Data Limite": limpar_valor(linha.get("Data Limite", "")),
            "Situação": limpar_valor(linha.get("SITUAÇÃO", "")),
            "Motivo": motivo,
            "Observação": observacao,
            "Data Registro": data_registro,
            "Origem": origem,
            "Validado": validado,
        })

    df_base_nova = pd.DataFrame(novas_linhas, columns=colunas_base)

    if not df_base_nova.empty:
        ordem_situacao = {
            "VENCIDA": 1,
            "VENCE HOJE": 2,
            "FUTURA": 3,
            "SEM DATA": 4,
        }

        df_base_nova["_ORDEM"] = (
            df_base_nova["Situação"]
            .fillna("")
            .astype(str)
            .str.strip()
            .map(ordem_situacao)
            .fillna(99)
        )

        df_base_nova = df_base_nova.sort_values(
            by=["_ORDEM", "Prestador", "Data Limite", "Chamado"],
            ascending=True
        )

        df_base_nova = df_base_nova.drop(columns=["_ORDEM"])

    try:
        with pd.ExcelWriter(ARQUIVO_BASE_JUSTIFICATIVAS, engine="openpyxl") as writer:
            df_base_nova.to_excel(writer, index=False, sheet_name="Justificativas")

        formatar_base_justificativas(ARQUIVO_BASE_JUSTIFICATIVAS)

        print(
            "Base de justificativas recriada com base nas pendências atuais: "
            f"{len(df_base_nova)} OS(s)."
        )

    except Exception as erro:
        print(
            "Aviso: não consegui salvar a base de justificativas. "
            "Verifique se ela está aberta no Excel. "
            f"Erro: {erro}"
        )


def formatar_base_justificativas(caminho_xlsx):
    try:
        wb = load_workbook(caminho_xlsx)

        if "Justificativas" in wb.sheetnames:
            ws = wb["Justificativas"]
        else:
            ws = wb.active
            ws.title = "Justificativas"

        if "Listas" in wb.sheetnames:
            del wb["Listas"]

        ws_listas = wb.create_sheet("Listas")
        ws_listas.sheet_state = "hidden"

        motivos = [
            "Erro de roteirização",
            "Falta de Equipamento",
            "Falta de Bobina",
            "Falta de POS",
            "Falta de POS avançado",
            "Falta de Rota",
            "Rota viagem",
            "Reversão de Insucesso",
            "Equipamento chegou Fora do Prazo",
            "Falha Operacional-Alteração de PA",
            "Falha Operacional-Gestão da rota",
            "Falha na sincronização do encerramento",
            "Instabilidade Sistema Contratante",
            "Instabilidade Sistema Mobyan",
            "Problemas meteorológicos",
        ]

        origens = [
            "Manual",
            "Técnico",
            "Base",
        ]

        validados = [
            "Sim",
            "Não",
        ]

        ws_listas["A1"] = "Motivos"
        for indice, motivo in enumerate(motivos, start=2):
            ws_listas.cell(row=indice, column=1).value = motivo

        ws_listas["B1"] = "Origens"
        for indice, origem in enumerate(origens, start=2):
            ws_listas.cell(row=indice, column=2).value = origem

        ws_listas["C1"] = "Validados"
        for indice, validado in enumerate(validados, start=2):
            ws_listas.cell(row=indice, column=3).value = validado

        max_row = ws.max_row
        max_col = ws.max_column

        preenchimento_cabecalho = PatternFill("solid", fgColor="1F4E78")
        fonte_cabecalho = Font(color="FFFFFF", bold=True)

        preenchimento_vencida = PatternFill("solid", fgColor="F4CCCC")
        preenchimento_hoje = PatternFill("solid", fgColor="FFF2CC")
        preenchimento_validado = PatternFill("solid", fgColor="D9EAD3")
        preenchimento_nao_validado = PatternFill("solid", fgColor="FCE4D6")

        borda_fina = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )

        for cell in ws[1]:
            cell.fill = preenchimento_cabecalho
            cell.font = fonte_cabecalho
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borda_fina

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        headers = [cell.value for cell in ws[1]]

        col_situacao = headers.index("Situação") + 1 if "Situação" in headers else None
        col_validado = headers.index("Validado") + 1 if "Validado" in headers else None
        col_motivo = headers.index("Motivo") + 1 if "Motivo" in headers else None
        col_observacao = headers.index("Observação") + 1 if "Observação" in headers else None
        col_data_registro = headers.index("Data Registro") + 1 if "Data Registro" in headers else None
        col_origem = headers.index("Origem") + 1 if "Origem" in headers else None

        for row in range(2, max_row + 1):
            situacao = ws.cell(row=row, column=col_situacao).value if col_situacao else ""
            validado = ws.cell(row=row, column=col_validado).value if col_validado else ""

            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = borda_fina
                cell.alignment = Alignment(vertical="center", wrap_text=False)

            if situacao == "VENCIDA":
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = preenchimento_vencida

            elif situacao == "VENCE HOJE":
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = preenchimento_hoje

            if col_validado:
                cell_validado = ws.cell(row=row, column=col_validado)

                if str(validado).strip().lower() in ["sim", "s", "yes", "y"]:
                    cell_validado.fill = preenchimento_validado
                    cell_validado.font = Font(bold=True)
                else:
                    cell_validado.fill = preenchimento_nao_validado

            if col_motivo:
                ws.cell(row=row, column=col_motivo).alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

            if col_observacao:
                ws.cell(row=row, column=col_observacao).alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

            if col_data_registro:
                valor_data = ws.cell(row=row, column=col_data_registro).value
                ws.cell(row=row, column=col_data_registro).value = formatar_data_registro_base(valor_data)
                ws.cell(row=row, column=col_data_registro).number_format = "dd/mm/yyyy"

        max_row_validacao = max(max_row + 200, 500)

        if col_motivo:
            dv_motivo = DataValidation(
                type="list",
                formula1=f"=Listas!$A$2:$A${len(motivos) + 1}",
                allow_blank=True,
            )
            ws.add_data_validation(dv_motivo)
            dv_motivo.add(
                f"{get_column_letter(col_motivo)}2:{get_column_letter(col_motivo)}{max_row_validacao}"
            )

        if col_origem:
            dv_origem = DataValidation(
                type="list",
                formula1=f"=Listas!$B$2:$B${len(origens) + 1}",
                allow_blank=True,
            )
            ws.add_data_validation(dv_origem)
            dv_origem.add(
                f"{get_column_letter(col_origem)}2:{get_column_letter(col_origem)}{max_row_validacao}"
            )

        if col_validado:
            dv_validado = DataValidation(
                type="list",
                formula1=f"=Listas!$C$2:$C${len(validados) + 1}",
                allow_blank=True,
            )
            ws.add_data_validation(dv_validado)
            dv_validado.add(
                f"{get_column_letter(col_validado)}2:{get_column_letter(col_validado)}{max_row_validacao}"
            )

        larguras = {
            "Chamado": 14,
            "Prestador": 30,
            "Cliente": 36,
            "Cidade": 20,
            "Data Limite": 14,
            "Situação": 14,
            "Motivo": 34,
            "Observação": 52,
            "Data Registro": 16,
            "Origem": 16,
            "Validado": 12,
        }

        for col in range(1, max_col + 1):
            letra = get_column_letter(col)
            cabecalho = ws.cell(row=1, column=col).value

            if cabecalho in larguras:
                ws.column_dimensions[letra].width = larguras[cabecalho]
            else:
                ws.column_dimensions[letra].width = 18

        ws.row_dimensions[1].height = 24

        for row in range(2, max_row + 1):
            ws.row_dimensions[row].height = 22

        wb.save(caminho_xlsx)

    except Exception as erro:
        print(f"Aviso: não consegui formatar a base de justificativas. Erro: {erro}")


def classificar_situacao(data_limite, hoje):
    if pd.isna(data_limite):
        return "SEM DATA"

    if data_limite < hoje:
        return "VENCIDA"

    if data_limite == hoje:
        return "VENCE HOJE"

    return "FUTURA"


def obter_top(df, coluna, total_pendencias, limite=5, ignorar_vazio=True):
    if coluna not in df.columns or total_pendencias == 0:
        return pd.DataFrame(columns=[coluna, "Quantidade", "%"])

    serie = df[coluna].fillna("").astype(str).str.strip()

    if ignorar_vazio:
        serie = serie[serie != ""]

    if len(serie) == 0:
        return pd.DataFrame(columns=[coluna, "Quantidade", "%"])

    resumo = serie.value_counts().reset_index()
    resumo.columns = [coluna, "Quantidade"]
    resumo["%"] = resumo["Quantidade"].apply(lambda qtd: qtd / total_pendencias)

    return resumo.head(limite)


# ATENÇÃO: a lógica de desenho das imagens por prestador daqui até
# gerar_imagens_por_prestador() tem uma cópia paralela e intencional em
# gerar_imagens_prestadores.py, usada para regerar as imagens manualmente a
# partir da planilha já salva, sem re-raspar a Mobyan. Se mudar o layout ou
# as cores aqui, replique lá também.
def carregar_fonte_imagem(tamanho=14, negrito=False):
    caminhos = []

    pasta_windows = Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts"

    if negrito:
        caminhos.extend([
            str(pasta_windows / "arialbd.ttf"),
            str(pasta_windows / "calibrib.ttf"),
            "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
            "/System/Library/Fonts/Supplemental/Arial Bold Unicode.ttf",
            "/Library/Fonts/Arial Bold.ttf",
        ])
    else:
        caminhos.extend([
            str(pasta_windows / "arial.ttf"),
            str(pasta_windows / "calibri.ttf"),
            "/System/Library/Fonts/Supplemental/Arial.ttf",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
            "/Library/Fonts/Arial.ttf",
        ])

    for caminho in caminhos:
        try:
            return ImageFont.truetype(caminho, tamanho)
        except Exception:
            pass

    return ImageFont.load_default()


FONTE_CABECALHO_IMAGEM = carregar_fonte_imagem(14, negrito=True)
FONTE_TEXTO_IMAGEM = carregar_fonte_imagem(13, negrito=False)
FONTE_TEXTO_NEGRITO_IMAGEM = carregar_fonte_imagem(13, negrito=True)


def texto_curto_imagem(texto, fonte, largura_maxima):
    texto = limpar_valor(texto)

    if texto == "":
        return ""

    draw_temp = ImageDraw.Draw(Image.new("RGB", (1, 1)))

    if draw_temp.textlength(texto, font=fonte) <= largura_maxima:
        return texto

    sufixo = "..."
    texto_base = texto

    while len(texto_base) > 0:
        tentativa = texto_base + sufixo

        if draw_temp.textlength(tentativa, font=fonte) <= largura_maxima:
            return tentativa

        texto_base = texto_base[:-1]

    return sufixo


def desenhar_texto_centralizado(draw, caixa, texto, fonte, cor):
    x1, y1, x2, y2 = caixa
    texto = limpar_valor(texto)

    bbox = draw.textbbox((0, 0), texto, font=fonte)
    largura_texto = bbox[2] - bbox[0]
    altura_texto = bbox[3] - bbox[1]

    x = x1 + ((x2 - x1) - largura_texto) / 2
    y = y1 + ((y2 - y1) - altura_texto) / 2 - 1

    draw.text((x, y), texto, font=fonte, fill=cor)


def desenhar_texto_esquerda(draw, caixa, texto, fonte, cor):
    x1, y1, x2, y2 = caixa
    texto = texto_curto_imagem(texto, fonte, (x2 - x1) - 10)

    bbox = draw.textbbox((0, 0), texto, font=fonte)
    altura_texto = bbox[3] - bbox[1]

    x = x1 + 5
    y = y1 + ((y2 - y1) - altura_texto) / 2 - 1

    draw.text((x, y), texto, font=fonte, fill=cor)


def cor_linha_por_situacao(situacao):
    situacao = limpar_valor(situacao).upper()

    if situacao == "VENCIDA":
        return COR_LINHA_VENCIDA

    if situacao == "VENCE HOJE":
        return COR_LINHA_HOJE

    return COR_LINHA_FUTURA


def cor_situacao(situacao):
    situacao = limpar_valor(situacao).upper()

    if situacao == "VENCIDA":
        return COR_SITUACAO_VENCIDA, COR_TEXTO_BRANCO

    if situacao == "VENCE HOJE":
        return COR_SITUACAO_HOJE, COR_TEXTO

    return COR_SITUACAO_FUTURA, COR_TEXTO_BRANCO


def limpar_imagens_antigas():
    PASTA_IMAGENS_PRESTADOR.mkdir(parents=True, exist_ok=True)

    for arquivo in PASTA_IMAGENS_PRESTADOR.glob("*.png"):
        try:
            arquivo.unlink()
        except Exception as erro:
            print(f"Aviso: não consegui apagar imagem antiga {arquivo}. Erro: {erro}")


def limpar_pasta_excel_prestador_antiga():
    if not PASTA_PRESTADOR_ANTIGA.exists():
        return

    for arquivo in PASTA_PRESTADOR_ANTIGA.glob("*.xlsx"):
        try:
            arquivo.unlink()
        except Exception as erro:
            print(f"Aviso: não consegui apagar arquivo antigo por prestador {arquivo}. Erro: {erro}")


def gerar_imagem_prestador(prestador, df_prestador):
    df_img = df_prestador[COLUNAS_IMAGEM].copy()

    total_largura_tabela = sum(LARGURAS_IMAGEM[coluna] for coluna in COLUNAS_IMAGEM)
    total_altura_tabela = ALTURA_CABECALHO_IMAGEM + (len(df_img) * ALTURA_LINHA_IMAGEM)

    largura_imagem = total_largura_tabela + (MARGEM_IMAGEM * 2)
    altura_imagem = total_altura_tabela + (MARGEM_IMAGEM * 2)

    imagem = Image.new("RGB", (largura_imagem, altura_imagem), COR_FUNDO)
    draw = ImageDraw.Draw(imagem)

    x = MARGEM_IMAGEM
    y = MARGEM_IMAGEM

    for coluna in COLUNAS_IMAGEM:
        largura = LARGURAS_IMAGEM[coluna]
        caixa = (x, y, x + largura, y + ALTURA_CABECALHO_IMAGEM)

        draw.rectangle(caixa, fill=COR_CABECALHO, outline=COR_BORDA)

        desenhar_texto_centralizado(
            draw,
            caixa,
            coluna,
            FONTE_CABECALHO_IMAGEM,
            COR_TEXTO_CABECALHO,
        )

        x += largura

    y += ALTURA_CABECALHO_IMAGEM

    for _, linha in df_img.iterrows():
        situacao = linha.get("SITUAÇÃO", "")
        cor_linha = cor_linha_por_situacao(situacao)

        x = MARGEM_IMAGEM

        for coluna in COLUNAS_IMAGEM:
            largura = LARGURAS_IMAGEM[coluna]
            caixa = (x, y, x + largura, y + ALTURA_LINHA_IMAGEM)

            if coluna == "SITUAÇÃO":
                fill, cor_texto = cor_situacao(situacao)
                draw.rectangle(caixa, fill=fill, outline=COR_BORDA)

                desenhar_texto_centralizado(
                    draw,
                    caixa,
                    linha.get(coluna, ""),
                    FONTE_TEXTO_NEGRITO_IMAGEM,
                    cor_texto,
                )
            else:
                draw.rectangle(caixa, fill=cor_linha, outline=COR_BORDA)

                if coluna in [
                    "Chamado",
                    "Numero Referencia",
                    "Contratante",
                    "Status",
                    "Data Limite",
                    "CNPJ / CPF",
                    "Cidade",
                ]:
                    desenhar_texto_centralizado(
                        draw,
                        caixa,
                        linha.get(coluna, ""),
                        FONTE_TEXTO_IMAGEM,
                        COR_TEXTO,
                    )
                else:
                    desenhar_texto_esquerda(
                        draw,
                        caixa,
                        linha.get(coluna, ""),
                        FONTE_TEXTO_IMAGEM,
                        COR_TEXTO,
                    )

            x += largura

        y += ALTURA_LINHA_IMAGEM

    nome_arquivo = limpar_nome_arquivo(prestador) + ".png"
    caminho_saida = PASTA_IMAGENS_PRESTADOR / nome_arquivo

    imagem.save(caminho_saida, "PNG")

    return caminho_saida


def gerar_imagens_por_prestador(df_pendencias):
    print("Gerando imagens por prestador...")

    PASTA_IMAGENS_PRESTADOR.mkdir(parents=True, exist_ok=True)
    limpar_imagens_antigas()
    limpar_pasta_excel_prestador_antiga()

    if "Prestador" not in df_pendencias.columns:
        print("Coluna Prestador não encontrada. Pulando geração de imagens.")
        return {}

    for coluna in COLUNAS_IMAGEM:
        if coluna not in df_pendencias.columns:
            df_pendencias[coluna] = ""

    serie_prestadores = (
        df_pendencias["Prestador"]
        .fillna("SEM_PRESTADOR")
        .astype(str)
        .str.strip()
        .replace("", "SEM_PRESTADOR")
    )

    prestadores = sorted(serie_prestadores.unique())

    imagens = {}
    total = 0

    for prestador in prestadores:
        df_prestador = df_pendencias[serie_prestadores == prestador].copy()

        if df_prestador.empty:
            continue

        caminho = gerar_imagem_prestador(prestador, df_prestador)
        imagens[prestador] = caminho

        print(f"Imagem gerada: {caminho}")
        total += 1

    print(f"Total de imagens geradas: {total}")

    return imagens



def criar_aba_resumo(writer, df_pendencias, df_acompanhamento=None):
    total_pendencias = len(df_pendencias)

    total_vencidas = (
        len(df_pendencias[df_pendencias["SITUAÇÃO"] == "VENCIDA"])
        if "SITUAÇÃO" in df_pendencias.columns
        else 0
    )

    total_vence_hoje = (
        len(df_pendencias[df_pendencias["SITUAÇÃO"] == "VENCE HOJE"])
        if "SITUAÇÃO" in df_pendencias.columns
        else 0
    )

    total_falta_abonar = (
        len(
            df_pendencias[
                df_pendencias["Justificativa do Abono"]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.upper()
                == "FALTA ABONAR"
            ]
        )
        if "Justificativa do Abono" in df_pendencias.columns
        else 0
    )

    # Base própria do Acompanhamento (exclui OSs de bobina) — usada para as %
    # de "Em Risco"/"Controladas", que não pertencem ao universo de total_pendencias.
    total_acompanhamento = len(df_acompanhamento) if df_acompanhamento is not None else 0

    if df_acompanhamento is None or df_acompanhamento.empty:
        total_em_risco = 0
        total_controladas = 0
    else:
        if "Status Operacional" in df_acompanhamento.columns:
            status_operacional = df_acompanhamento["Status Operacional"].fillna("").astype(str).str.strip()
        else:
            status_operacional = pd.Series([""] * len(df_acompanhamento), index=df_acompanhamento.index)

        # Compatibilidade com versões antigas que juntavam "Risco" e "Ação
        # Necessária" em uma única coluna "Risco / Ação".
        if "Risco" in df_acompanhamento.columns:
            risco = df_acompanhamento["Risco"].fillna("").astype(str).str.strip().str.upper()
        elif "Risco / Ação" in df_acompanhamento.columns:
            risco = (
                df_acompanhamento["Risco / Ação"]
                .fillna("")
                .astype(str)
                .str.split("|", n=1)
                .str[0]
                .str.strip()
                .str.upper()
            )
        else:
            risco = pd.Series([""] * len(df_acompanhamento), index=df_acompanhamento.index)

        total_em_risco = len(df_acompanhamento[risco.isin(["CRÍTICO", "ALTO RISCO", "ABONAR"])])
        total_controladas = len(
            df_acompanhamento[
                status_operacional.isin([
                    "Atendimento confirmado",
                    "Em rota",
                    "Em atendimento",
                    "Atendido, aguardando baixa",
                ])
            ]
        )

    def pct(valor, base):
        if base == 0:
            return 0
        return valor / base

    cards = pd.DataFrame([
        ["Total de Pendências", total_pendencias, 1 if total_pendencias > 0 else 0],
        ["Vencidas", total_vencidas, pct(total_vencidas, total_pendencias)],
        ["Vence Hoje", total_vence_hoje, pct(total_vence_hoje, total_pendencias)],
        ["Falta Abonar", total_falta_abonar, pct(total_falta_abonar, total_pendencias)],
        ["Em Risco", total_em_risco, pct(total_em_risco, total_acompanhamento)],
        ["Controladas", total_controladas, pct(total_controladas, total_acompanhamento)],
    ], columns=["Indicador", "Quantidade", "%"])

    top_prestador = obter_top(df_pendencias, "Prestador", total_pendencias, limite=5)
    top_cidade = obter_top(df_pendencias, "Cidade", total_pendencias, limite=5)

    sheet_name = "Resumo"

    cards.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2, startcol=0)

    pd.DataFrame([["Top 5 Prestadores"]]).to_excel(
        writer, index=False, header=False, sheet_name=sheet_name, startrow=11, startcol=0
    )
    top_prestador.to_excel(writer, index=False, sheet_name=sheet_name, startrow=12, startcol=0)

    pd.DataFrame([["Top 5 Cidades"]]).to_excel(
        writer, index=False, header=False, sheet_name=sheet_name, startrow=11, startcol=5
    )
    top_cidade.to_excel(writer, index=False, sheet_name=sheet_name, startrow=12, startcol=5)

def preparar_dataframe_final(df, hoje):
    colunas_desejadas = [
        "Chamado",
        "Numero Referencia",
        "Contratante",
        "Serviço",
        "Status",
        "Data Limite",
        "Cliente",
        "CNPJ / CPF",
        "Cidade",
        "Técnico",
        "Prestador",
        "Justificativa do Abono",
    ]

    for coluna in colunas_desejadas:
        if coluna not in df.columns:
            df[coluna] = ""

    datas_limite = pd.to_datetime(
        df["Data Limite"],
        dayfirst=True,
        errors="coerce"
    )

    df["_DATA_LIMITE_FORMATADA"] = datas_limite.dt.date

    df["Data Limite"] = datas_limite.dt.strftime("%d/%m/%Y")
    df["Data Limite"] = df["Data Limite"].fillna("")

    df = df[colunas_desejadas + ["_DATA_LIMITE_FORMATADA"]].copy()

    df.insert(
        0,
        "SITUAÇÃO",
        df["_DATA_LIMITE_FORMATADA"].apply(
            lambda data_limite: classificar_situacao(data_limite, hoje)
        )
    )

    for indice, linha in df.iterrows():
        situacao = str(linha.get("SITUAÇÃO", "")).strip().upper()
        justificativa = linha.get("Justificativa do Abono", "")

        # A coluna "Justificativa do Abono" representa somente o que veio do sistema Mobyan.
        # Se a OS está vencida e veio sem justificativa do sistema, ela precisa aparecer como FALTA ABONAR.
        if situacao == "VENCIDA" and valor_vazio(justificativa):
            df.at[indice, "Justificativa do Abono"] = "FALTA ABONAR"

    df = df.sort_values(
        by="_DATA_LIMITE_FORMATADA",
        ascending=True,
        na_position="last"
    )

    df.drop(columns=["_DATA_LIMITE_FORMATADA"], inplace=True)

    colunas_finais = [
        "SITUAÇÃO",
        "Chamado",
        "Numero Referencia",
        "Contratante",
        "Serviço",
        "Status",
        "Data Limite",
        "Cliente",
        "CNPJ / CPF",
        "Cidade",
        "Técnico",
        "Prestador",
        "Justificativa do Abono",
    ]

    for coluna in colunas_finais:
        if coluna not in df.columns:
            df[coluna] = ""

    for coluna in df.columns:
        df[coluna] = (
            df[coluna]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace(["nan", "NaN", "None", "NULL", "null", "NaT"], "")
        )

    return df[colunas_finais]

def limpar_csv_temporario(caminho_csv):
    try:
        if caminho_csv.exists():
            caminho_csv.unlink()
            print(f"CSV temporário apagado: {caminho_csv}")
    except Exception as erro:
        print(f"Aviso: não consegui apagar o CSV temporário. Erro: {erro}")


def ler_contatos_prestadores():
    if not ARQUIVO_CONTATOS_PRESTADORES.exists():
        print("Base de contatos dos prestadores não encontrada.")
        return pd.DataFrame(columns=["Prestador", "Responsável", "WhatsApp", "Enviar", "Observação"])

    try:
        df = pd.read_excel(ARQUIVO_CONTATOS_PRESTADORES, dtype=str)
    except Exception as erro:
        print(f"Não consegui ler a base de contatos. Erro: {erro}")
        return pd.DataFrame(columns=["Prestador", "Responsável", "WhatsApp", "Enviar", "Observação"])

    colunas = ["Prestador", "Responsável", "WhatsApp", "Enviar", "Observação"]

    for coluna in colunas:
        if coluna not in df.columns:
            df[coluna] = ""

    df = df[colunas].copy()

    df["Prestador"] = df["Prestador"].fillna("").astype(str).str.strip()
    df["Responsável"] = df["Responsável"].fillna("").astype(str).str.strip()
    df["WhatsApp"] = df["WhatsApp"].fillna("").astype(str).str.replace(r"\D", "", regex=True)
    df["Enviar"] = df["Enviar"].fillna("").astype(str).str.strip()
    df["Observação"] = df["Observação"].fillna("").astype(str).str.strip()

    df = df[df["Prestador"] != ""].copy()

    return df



def criar_mensagem_envio(responsavel, prestador, total_pendencias, vencidas, vence_hoje):
    nome = responsavel.strip() if not valor_vazio(responsavel) else ""

    if nome:
        saudacao = f"Bom dia, {nome}!"
    else:
        saudacao = "Bom dia!"

    if total_pendencias <= 0:
        return ""

    partes = []

    if vencidas > 0:
        partes.append(f"{vencidas} vencida(s)")

    if vence_hoje > 0:
        partes.append(f"{vence_hoje} com prazo para hoje")

    resumo = " e ".join(partes) if partes else f"{total_pendencias} pendente(s)"

    mensagem = (
        f"{saudacao}\n\n"
        f"Segue pendências da base {prestador}.\n\n"
        f"Temos {resumo}. Preciso do retorno operacional para evitar atraso no SLA.\n\n"
        "Por favor, responda cada OS com uma das opções abaixo:\n"
        "1 - Em rota / será atendida hoje\n"
        "2 - Atendimento confirmado com horário\n"
        "3 - Em atendimento\n"
        "4 - Atendido, aguardando baixa no sistema\n"
        "5 - Cliente indisponível\n"
        "6 - Sem técnico\n"
        "7 - Sem equipamento/material\n"
        "8 - Vai atrasar\n"
        "9 - Precisa abonar\n\n"
        "Pode responder assim:\n"
        "OS 123456 - opção 1\n"
        "OS 123457 - opção 2, previsão 14h\n"
        "OS 123458 - opção 8, previsão amanhã"
    )

    return mensagem


def criar_mensagem_acompanhamento(responsavel, prestador, total_pendencias, vencidas, vence_hoje):
    nome = responsavel.strip() if not valor_vazio(responsavel) else ""

    if nome:
        saudacao = f"Boa tarde, {nome}!"
    else:
        saudacao = "Boa tarde!"

    if total_pendencias <= 0:
        return ""

    partes = []

    if vencidas > 0:
        partes.append(f"{vencidas} vencida(s)")

    if vence_hoje > 0:
        partes.append(f"{vence_hoje} com prazo para hoje")

    resumo = " e ".join(partes) if partes else f"{total_pendencias} pendente(s)"

    mensagem = (
        f"{saudacao}\n\n"
        f"Segue atualização das pendências da base {prestador}.\n\n"
        f"Ainda temos {resumo}.\n\n"
        "Me atualiza, por favor, quais OSs já estão controladas e quais têm risco de atraso.\n\n"
        "Use as opções abaixo para facilitar:\n"
        "1 - Em rota / será atendida hoje\n"
        "2 - Atendimento confirmado com horário\n"
        "3 - Em atendimento\n"
        "4 - Atendido, aguardando baixa no sistema\n"
        "5 - Cliente indisponível\n"
        "6 - Sem técnico\n"
        "7 - Sem equipamento/material\n"
        "8 - Vai atrasar\n"
        "9 - Precisa abonar\n\n"
        "Se tiver OS que vai atrasar, já me manda a previsão e o motivo."
    )

    return mensagem

def criar_dataframe_envios(df_pendencias, imagens_por_prestador):
    print("Criando fila de envios...")

    contatos = ler_contatos_prestadores()

    colunas_envios = [
        "Prestador",
        "Responsável",
        "WhatsApp",
        "Enviar",
        "Imagem",
        "Total Pendências",
        "Vencidas",
        "Vence Hoje",
        "Status Envio",
        "Mensagem",
        "Mensagem Manhã",
        "Mensagem Acompanhamento",
        "Observação",
    ]

    if contatos.empty:
        print("Nenhum contato encontrado para envio.")
        return pd.DataFrame(columns=colunas_envios)

    linhas = []

    for _, contato in contatos.iterrows():
        prestador = contato["Prestador"]
        responsavel = contato["Responsável"]
        whatsapp = contato["WhatsApp"]
        enviar = contato["Enviar"]
        observacao = contato["Observação"]

        enviar_normalizado = enviar.strip().lower()

        df_prestador = df_pendencias[
            df_pendencias["Prestador"]
            .fillna("")
            .astype(str)
            .str.strip()
            == prestador
        ].copy()

        total_pendencias = len(df_prestador)
        vencidas = len(df_prestador[df_prestador["SITUAÇÃO"] == "VENCIDA"])
        vence_hoje = len(df_prestador[df_prestador["SITUAÇÃO"] == "VENCE HOJE"])

        caminho_imagem = imagens_por_prestador.get(prestador, "")
        caminho_imagem_texto = str(caminho_imagem) if caminho_imagem else ""

        mensagem_manha = ""
        mensagem_acompanhamento = ""

        if enviar_normalizado not in ["sim", "s", "yes", "y"]:
            status = "Não marcado para envio"
        elif whatsapp == "":
            status = "WhatsApp não informado"
        elif total_pendencias == 0:
            status = "Sem pendências"
        elif not caminho_imagem or not Path(caminho_imagem).exists():
            status = "Imagem não encontrada"
        else:
            status = "Pronto para envio"

            mensagem_manha = criar_mensagem_envio(
                responsavel=responsavel,
                prestador=prestador,
                total_pendencias=total_pendencias,
                vencidas=vencidas,
                vence_hoje=vence_hoje
            )

            mensagem_acompanhamento = criar_mensagem_acompanhamento(
                responsavel=responsavel,
                prestador=prestador,
                total_pendencias=total_pendencias,
                vencidas=vencidas,
                vence_hoje=vence_hoje
            )

        linhas.append({
            "Prestador": prestador,
            "Responsável": responsavel,
            "WhatsApp": whatsapp,
            "Enviar": enviar,
            "Imagem": caminho_imagem_texto,
            "Total Pendências": total_pendencias,
            "Vencidas": vencidas,
            "Vence Hoje": vence_hoje,
            "Status Envio": status,
            "Mensagem": mensagem_manha,
            "Mensagem Manhã": mensagem_manha,
            "Mensagem Acompanhamento": mensagem_acompanhamento,
            "Observação": observacao,
        })

    df_envios = pd.DataFrame(linhas, columns=colunas_envios)

    print(f"Fila de envios criada: {len(df_envios)} registro(s).")

    return df_envios



STATUS_OPERACIONAIS = [
    "Sem retorno",
    "Atendimento confirmado",
    "Em rota",
    "Em atendimento",
    "Atendido, aguardando baixa",
    "Cliente indisponível",
    "Sem técnico",
    "Sem equipamento",
    "Em tratativa no email",
    "Aguardando cancelamento",
    "Reversão de Insucesso",
    "Vai atrasar",
    "Precisa abonar",
]

STATUS_CONTROLADOS = [
    "Atendimento confirmado",
    "Em rota",
    "Em atendimento",
    "Atendido, aguardando baixa",
]

STATUS_MONITORAR = [
    "Em tratativa no email",
    "Aguardando cancelamento",
    "Reversão de Insucesso",
]

STATUS_CRITICOS = [
    "Cliente indisponível",
    "Sem técnico",
    "Sem equipamento",
    "Vai atrasar",
]

OPCOES_PREVISAO = [
    "Hoje até 10h",
    "Hoje até 12h",
    "Hoje até 15h",
    "Hoje até 17h",
    "Hoje fim do dia",
    "Amanhã manhã",
    "Amanhã tarde",
    "Sem previsão",
    "Aguardar retorno",
    "Não se aplica",
]

OPCOES_ULTIMO_RETORNO = [
    "Cobrado no WhatsApp",
    "Cobrado por ligação",
    "Base confirmou",
    "Técnico em rota",
    "Técnico em atendimento",
    "Aguardando técnico",
    "Aguardando cliente",
    "Aguardando equipamento",
    "Em tratativa no email",
    "Cancelamento solicitado",
    "Reversão solicitada",
    "Sem resposta",
]

# Aba Acompanhamento em formato de trabalho.
# A planilha completa continua nas abas Pendências/Chamados.
# Aqui deixamos as colunas essenciais separadas, mas com técnico encurtado
# para manter a leitura confortável em 100%.
COLUNAS_ACOMPANHAMENTO = [
    "SITUAÇÃO",
    "Chamado",
    "Prestador",
    "Cliente",
    "Cidade",
    "Serviço",
    "Status Mobyan",
    "Data Limite",
    "Técnico",
    "Status Operacional",
    "Previsão",
    "Último Retorno",
    "Observação",
    "Risco",
    "Ação Necessária",
]


def obter_primeiro_nome_tecnico(valor):
    texto = limpar_valor(valor)

    if texto == "":
        return ""

    # Remove prefixos como (TA), que ocupam espaço e não ajudam na análise diária.
    texto = re.sub(r"^\s*\(\s*TA\s*\)\s*", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"^\s*TA\s+", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"\s+", " ", texto).strip()

    if texto == "":
        return ""

    return texto.split(" ")[0]


def criar_backup_planilha_atual():
    if not ARQUIVO_FINAL_PENDENCIAS.exists():
        return

    try:
        PASTA_BACKUPS_PENDENCIAS.mkdir(parents=True, exist_ok=True)

        # Mantém somente o backup mais recente, usado para preservar os
        # preenchimentos manuais do Acompanhamento.
        for arquivo in PASTA_BACKUPS_PENDENCIAS.glob(
            "pendencias_do_dia_backup_*.xlsx"
        ):
            if arquivo != ARQUIVO_BACKUP_ULTIMO:
                try:
                    arquivo.unlink()
                except Exception:
                    pass

        shutil.copy2(ARQUIVO_FINAL_PENDENCIAS, ARQUIVO_BACKUP_ULTIMO)
        print(f"Backup anterior atualizado em: {ARQUIVO_BACKUP_ULTIMO}")

    except Exception as erro:
        print(f"Aviso: não consegui criar backup da planilha anterior. Erro: {erro}")


def ler_acompanhamento_anterior():
    candidatos = []

    if ARQUIVO_FINAL_PENDENCIAS.exists():
        candidatos.append(ARQUIVO_FINAL_PENDENCIAS)

    if ARQUIVO_BACKUP_ULTIMO.exists():
        candidatos.append(ARQUIVO_BACKUP_ULTIMO)

    if PASTA_BACKUPS_PENDENCIAS.exists():
        for arquivo in sorted(PASTA_BACKUPS_PENDENCIAS.glob("pendencias_do_dia_backup_*.xlsx"), reverse=True):
            if arquivo not in candidatos:
                candidatos.append(arquivo)

    if not candidatos:
        return {}

    ultimo_erro = None

    for caminho in candidatos:
        try:
            df_anterior = pd.read_excel(
                caminho,
                sheet_name="Acompanhamento",
                dtype=str,
                keep_default_na=False,
                na_filter=False,
            )
        except Exception as erro:
            ultimo_erro = erro
            continue

        if "Chamado" not in df_anterior.columns:
            continue

        for coluna in df_anterior.columns:
            df_anterior[coluna] = (
                df_anterior[coluna]
                .fillna("")
                .astype(str)
                .str.strip()
                .replace(["nan", "NaN", "None", "NULL", "null", "NaT"], "")
            )

        df_anterior["Chamado_Normalizado"] = df_anterior["Chamado"].apply(normalizar_chamado)
        df_anterior = df_anterior[df_anterior["Chamado_Normalizado"] != ""].copy()
        df_anterior = df_anterior.drop_duplicates(subset=["Chamado_Normalizado"], keep="last")

        acompanhamento = {}

        for _, linha in df_anterior.iterrows():
            chamado = linha["Chamado_Normalizado"]

            # Compatibilidade entre as versões antigas/novas do acompanhamento.
            observacao = limpar_valor(linha.get("Observação", ""))
            if valor_vazio(observacao):
                observacao = limpar_valor(linha.get("Observação Operacional", ""))

            acompanhamento[chamado] = {
                "Status Operacional": limpar_valor(linha.get("Status Operacional", "")),
                "Previsão": limpar_valor(linha.get("Previsão", "")),
                "Último Retorno": limpar_valor(linha.get("Último Retorno", "")),
                "Observação": observacao,
            }

        print(f"Acompanhamento anterior carregado: {len(acompanhamento)} OS(s).")
        print(f"Origem do acompanhamento anterior: {caminho}")

        return acompanhamento

    if ultimo_erro:
        print(f"Aviso: não consegui ler acompanhamento anterior. Último erro: {ultimo_erro}")

    return {}


def ler_layout_acompanhamento_anterior():
    """
    Lê o layout visual da aba Acompanhamento existente antes de gerar uma nova planilha.
    Isso permite preservar as larguras que foram ajustadas manualmente no Excel.
    """
    candidatos = []

    if ARQUIVO_FINAL_PENDENCIAS.exists():
        candidatos.append(ARQUIVO_FINAL_PENDENCIAS)

    if ARQUIVO_BACKUP_ULTIMO.exists():
        candidatos.append(ARQUIVO_BACKUP_ULTIMO)

    if PASTA_BACKUPS_PENDENCIAS.exists():
        for arquivo in sorted(PASTA_BACKUPS_PENDENCIAS.glob("pendencias_do_dia_backup_*.xlsx"), reverse=True):
            if arquivo not in candidatos:
                candidatos.append(arquivo)

    for caminho in candidatos:
        try:
            wb = load_workbook(caminho)

            if "Acompanhamento" not in wb.sheetnames:
                continue

            ws = wb["Acompanhamento"]
            headers = [cell.value for cell in ws[1]]

            larguras = {}
            for indice, cabecalho in enumerate(headers, start=1):
                cabecalho = limpar_valor(cabecalho)
                if cabecalho == "":
                    continue

                letra = get_column_letter(indice)
                largura = ws.column_dimensions[letra].width

                if largura:
                    larguras[cabecalho] = float(largura)

            zoom = ws.sheet_view.zoomScale or 120

            altura_linhas = None
            if ws.max_row >= 2 and ws.row_dimensions[2].height:
                altura_linhas = float(ws.row_dimensions[2].height)

            if larguras:
                print(f"Layout anterior do acompanhamento carregado: {len(larguras)} coluna(s).")
                print(f"Origem do layout anterior: {caminho}")

                return {
                    "larguras": larguras,
                    "zoom": zoom,
                    "altura_linhas": altura_linhas,
                }

        except Exception as erro:
            print(f"Aviso: não consegui ler layout do acompanhamento em {caminho}. Erro: {erro}")

    return {}

def classificar_risco_operacional(situacao, justificativa_abono, status_operacional):
    situacao = limpar_valor(situacao).upper()
    justificativa_abono = limpar_valor(justificativa_abono).upper()
    status_operacional = limpar_valor(status_operacional)

    if status_operacional in STATUS_CONTROLADOS:
        return "CONTROLADO"

    if status_operacional == "Precisa abonar":
        return "ABONAR"

    if status_operacional in STATUS_MONITORAR:
        return "MONITORAR"

    if status_operacional in STATUS_CRITICOS:
        return "CRÍTICO"

    if situacao == "VENCIDA" and justificativa_abono == "FALTA ABONAR":
        return "ABONAR"

    if situacao == "VENCIDA":
        return "CRÍTICO"

    if situacao == "VENCE HOJE" and status_operacional == "Sem retorno":
        return "ALTO RISCO"

    if situacao == "VENCE HOJE":
        return "MONITORAR"

    return "MONITORAR"


def definir_acao_necessaria(situacao, justificativa_abono, status_operacional, risco):
    situacao = limpar_valor(situacao).upper()
    justificativa_abono = limpar_valor(justificativa_abono).upper()
    status_operacional = limpar_valor(status_operacional)
    risco = limpar_valor(risco).upper()

    if status_operacional in STATUS_CONTROLADOS:
        return "Monitorar baixa/encerramento"

    if status_operacional == "Em tratativa no email":
        return "Acompanhar tratativa e-mail"

    if status_operacional == "Aguardando cancelamento":
        return "Acompanhar cancelamento"

    if status_operacional == "Reversão de Insucesso":
        return "Acompanhar reversão/encerramento"

    if status_operacional == "Precisa abonar" or justificativa_abono == "FALTA ABONAR":
        return "Verificar abono no Mobyan"

    if status_operacional == "Sem retorno":
        if situacao == "VENCIDA":
            return "Cobrar base agora"
        if situacao == "VENCE HOJE":
            return "Cobrar previsão hoje"
        return "Acompanhar"

    if risco == "CRÍTICO":
        return "Acionar base e pedir plano"

    if risco == "ALTO RISCO":
        return "Cobrar retorno/previsão"

    return "Acompanhar"


def criar_dataframe_acompanhamento(df_pendencias, acompanhamento_anterior):
    print("Criando aba de acompanhamento operacional...")

    linhas = []

    for _, linha in df_pendencias.iterrows():
        servico = limpar_valor(linha.get("Serviço", ""))

        # OSs de bobina continuam aparecendo nas abas Pendências/Chamados,
        # mas não entram no Acompanhamento Operacional porque não impactam o controle de SLA do dia.
        if "BOBINA" in servico.upper():
            continue

        chamado = limpar_valor(linha.get("Chamado", ""))
        chamado_normalizado = normalizar_chamado(chamado)

        dados_anteriores = acompanhamento_anterior.get(chamado_normalizado, {})

        status_operacional = dados_anteriores.get("Status Operacional", "")
        previsao = dados_anteriores.get("Previsão", "")
        ultimo_retorno = dados_anteriores.get("Último Retorno", "")
        observacao_operacional = dados_anteriores.get("Observação", "")

        if valor_vazio(status_operacional):
            status_operacional = "Sem retorno"

        situacao = limpar_valor(linha.get("SITUAÇÃO", ""))
        justificativa_abono = limpar_valor(linha.get("Justificativa do Abono", ""))

        risco = classificar_risco_operacional(
            situacao=situacao,
            justificativa_abono=justificativa_abono,
            status_operacional=status_operacional,
        )

        acao = definir_acao_necessaria(
            situacao=situacao,
            justificativa_abono=justificativa_abono,
            status_operacional=status_operacional,
            risco=risco,
        )

        status_mobyan = limpar_valor(linha.get("Status", ""))
        tecnico_curto = obter_primeiro_nome_tecnico(linha.get("Técnico", ""))

        linhas.append({
            "SITUAÇÃO": situacao,
            "Chamado": chamado,
            "Prestador": limpar_valor(linha.get("Prestador", "")),
            "Cliente": limpar_valor(linha.get("Cliente", "")),
            "Cidade": limpar_valor(linha.get("Cidade", "")),
            "Serviço": servico,
            "Status Mobyan": status_mobyan,
            "Data Limite": limpar_valor(linha.get("Data Limite", "")),
            "Técnico": tecnico_curto,
            "Status Operacional": status_operacional,
            "Previsão": previsao,
            "Último Retorno": ultimo_retorno,
            "Observação": observacao_operacional,
            "Risco": risco,
            "Ação Necessária": acao,
        })

    df_acompanhamento = pd.DataFrame(linhas, columns=COLUNAS_ACOMPANHAMENTO)

    if not df_acompanhamento.empty:
        ordem_risco = {
            "ABONAR": 1,
            "CRÍTICO": 2,
            "ALTO RISCO": 3,
            "MONITORAR": 4,
            "CONTROLADO": 5,
        }

        ordem_situacao = {
            "VENCIDA": 1,
            "VENCE HOJE": 2,
            "FUTURA": 3,
            "SEM DATA": 4,
        }

        df_acompanhamento["_ORDEM_RISCO"] = (
            df_acompanhamento["Risco"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
            .map(ordem_risco)
            .fillna(99)
        )
        df_acompanhamento["_ORDEM_SITUACAO"] = (
            df_acompanhamento["SITUAÇÃO"].map(ordem_situacao).fillna(99)
        )

        # Ordenação operacional principal:
        # sempre da Data Limite mais antiga para a mais atual.
        # Assim as OSs mais vencidas aparecem primeiro no acompanhamento do dia.
        df_acompanhamento["_DATA_LIMITE_ORDENACAO"] = pd.to_datetime(
            df_acompanhamento["Data Limite"],
            dayfirst=True,
            errors="coerce",
        )

        df_acompanhamento = df_acompanhamento.sort_values(
            by=[
                "_DATA_LIMITE_ORDENACAO",
                "_ORDEM_SITUACAO",
                "_ORDEM_RISCO",
                "Prestador",
                "Chamado",
            ],
            ascending=True,
            na_position="last",
        ).drop(
            columns=[
                "_ORDEM_RISCO",
                "_RISCO_ORDENACAO",
                "_ORDEM_SITUACAO",
                "_DATA_LIMITE_ORDENACAO",
            ],
            errors="ignore",
        )

    print(f"Acompanhamento criado: {len(df_acompanhamento)} OS(s).")

    return df_acompanhamento


def formatar_aba_acompanhamento(caminho_xlsx, layout_anterior=None):
    wb = load_workbook(caminho_xlsx)

    if "Acompanhamento" not in wb.sheetnames:
        wb.save(caminho_xlsx)
        return

    ws = wb["Acompanhamento"]

    if layout_anterior is None:
        layout_anterior = {}

    if "Listas_Acompanhamento" in wb.sheetnames:
        del wb["Listas_Acompanhamento"]

    ws_listas = wb.create_sheet("Listas_Acompanhamento")
    ws_listas.sheet_state = "hidden"

    ws_listas["A1"] = "Status Operacional"
    for indice, status in enumerate(STATUS_OPERACIONAIS, start=2):
        ws_listas.cell(row=indice, column=1).value = status

    ws_listas["B1"] = "Previsão"
    for indice, previsao in enumerate(OPCOES_PREVISAO, start=2):
        ws_listas.cell(row=indice, column=2).value = previsao

    ws_listas["C1"] = "Último Retorno"
    for indice, retorno in enumerate(OPCOES_ULTIMO_RETORNO, start=2):
        ws_listas.cell(row=indice, column=3).value = retorno

    max_row = ws.max_row
    max_col = ws.max_column

    if max_row <= 1:
        wb.save(caminho_xlsx)
        return

    preenchimento_cabecalho = PatternFill("solid", fgColor="1F4E78")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)

    preenchimento_vencida = PatternFill("solid", fgColor="F4CCCC")
    preenchimento_vence_hoje = PatternFill("solid", fgColor="FFF2CC")
    preenchimento_controlado = PatternFill("solid", fgColor="D9EAD3")
    preenchimento_monitorar = PatternFill("solid", fgColor="FFF2CC")
    preenchimento_risco = PatternFill("solid", fgColor="F4CCCC")
    preenchimento_abonar = PatternFill("solid", fgColor="FF0000")
    preenchimento_manual = PatternFill("solid", fgColor="D9EAF7")

    fonte_branca_negrito = Font(color="FFFFFF", bold=True)
    fonte_preta_negrito = Font(color="000000", bold=True)

    borda_fina = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for cell in ws[1]:
        cell.fill = preenchimento_cabecalho
        cell.font = fonte_cabecalho
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda_fina

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    headers = [cell.value for cell in ws[1]]

    col_situacao = headers.index("SITUAÇÃO") + 1 if "SITUAÇÃO" in headers else None
    col_status_operacional = headers.index("Status Operacional") + 1 if "Status Operacional" in headers else None
    col_previsao = headers.index("Previsão") + 1 if "Previsão" in headers else None
    col_ultimo_retorno = headers.index("Último Retorno") + 1 if "Último Retorno" in headers else None
    col_observacao = headers.index("Observação") + 1 if "Observação" in headers else None
    col_risco = None
    if "Risco" in headers:
        col_risco = headers.index("Risco") + 1
    elif "Risco / Ação" in headers:
        col_risco = headers.index("Risco / Ação") + 1

    col_acao = headers.index("Ação Necessária") + 1 if "Ação Necessária" in headers else None
    col_justificativa = headers.index("Justificativa do Abono") + 1 if "Justificativa do Abono" in headers else None

    colunas_manuais = [
        col_status_operacional,
        col_previsao,
        col_ultimo_retorno,
        col_observacao,
    ]

    for row in range(2, max_row + 1):
        situacao = ws.cell(row=row, column=col_situacao).value if col_situacao else ""
        risco = ws.cell(row=row, column=col_risco).value if col_risco else ""
        risco = str(risco).split("|", 1)[0].strip()

        if situacao == "VENCIDA":
            preenchimento_linha = preenchimento_vencida
        elif situacao == "VENCE HOJE":
            preenchimento_linha = preenchimento_vence_hoje
        else:
            preenchimento_linha = PatternFill("solid", fgColor="FFFFFF")

        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = preenchimento_linha
            cell.border = borda_fina
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        for col in colunas_manuais:
            if col:
                cell_manual = ws.cell(row=row, column=col)
                cell_manual.fill = preenchimento_manual
                cell_manual.alignment = Alignment(vertical="top", wrap_text=True)

        if col_risco:
            cell_risco = ws.cell(row=row, column=col_risco)
            risco_texto = str(risco).split("|", 1)[0].strip().upper()

            if risco_texto == "CONTROLADO":
                cell_risco.fill = preenchimento_controlado
                cell_risco.font = fonte_preta_negrito
            elif risco_texto in ["CRÍTICO", "ALTO RISCO"]:
                cell_risco.fill = preenchimento_risco
                cell_risco.font = fonte_preta_negrito
            elif risco_texto == "ABONAR":
                cell_risco.fill = preenchimento_abonar
                cell_risco.font = fonte_branca_negrito
            else:
                cell_risco.fill = preenchimento_monitorar
                cell_risco.font = fonte_preta_negrito

            cell_risco.alignment = Alignment(horizontal="center", vertical="center")

        if col_justificativa:
            cell_just = ws.cell(row=row, column=col_justificativa)
            if str(cell_just.value).strip().upper() == "FALTA ABONAR":
                cell_just.fill = preenchimento_abonar
                cell_just.font = fonte_branca_negrito
                cell_just.alignment = Alignment(horizontal="center", vertical="center")

        if col_acao:
            ws.cell(row=row, column=col_acao).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    # SITUAÇÃO, Risco e Ação Necessária já ficam visíveis pela cor da linha/célula
    # aplicada acima (vermelho/amarelo/verde), então o texto vira ruído na leitura
    # diária. Removemos as colunas só depois de usar os valores para colorir.
    colunas_para_remover = sorted(
        {indice for indice in (col_acao, col_risco, col_situacao) if indice},
        reverse=True,
    )
    for indice in colunas_para_remover:
        ws.delete_cols(indice)

    max_col = ws.max_column
    headers = [cell.value for cell in ws[1]]
    ws.auto_filter.ref = ws.dimensions

    col_status_operacional = headers.index("Status Operacional") + 1 if "Status Operacional" in headers else None
    col_previsao = headers.index("Previsão") + 1 if "Previsão" in headers else None
    col_ultimo_retorno = headers.index("Último Retorno") + 1 if "Último Retorno" in headers else None

    max_row_validacao = max(max_row + 300, 500)

    if col_status_operacional:
        dv_status = DataValidation(
            type="list",
            formula1=f"=Listas_Acompanhamento!$A$2:$A${len(STATUS_OPERACIONAIS) + 1}",
            allow_blank=True,
        )
        ws.add_data_validation(dv_status)
        dv_status.add(
            f"{get_column_letter(col_status_operacional)}2:{get_column_letter(col_status_operacional)}{max_row_validacao}"
        )

    if col_previsao:
        dv_previsao = DataValidation(
            type="list",
            formula1=f"=Listas_Acompanhamento!$B$2:$B${len(OPCOES_PREVISAO) + 1}",
            allow_blank=True,
            showErrorMessage=False,
        )
        ws.add_data_validation(dv_previsao)
        dv_previsao.add(
            f"{get_column_letter(col_previsao)}2:{get_column_letter(col_previsao)}{max_row_validacao}"
        )

    if col_ultimo_retorno:
        dv_retorno = DataValidation(
            type="list",
            formula1=f"=Listas_Acompanhamento!$C$2:$C${len(OPCOES_ULTIMO_RETORNO) + 1}",
            allow_blank=True,
            showErrorMessage=False,
        )
        ws.add_data_validation(dv_retorno)
        dv_retorno.add(
            f"{get_column_letter(col_ultimo_retorno)}2:{get_column_letter(col_ultimo_retorno)}{max_row_validacao}"
        )

    # Layout operacional v5.4:
    # leitura confortável em 120%, técnico um pouco maior e textos centralizados.
    larguras_operacionais = {
        "SITUAÇÃO": 8,
        "Chamado": 10,
        "Prestador": 18,
        "Cliente": 30,
        "Cidade": 13,
        "Serviço": 16,
        "Status Mobyan": 13,
        "Data Limite": 11,
        "Técnico": 12,
        "Status Operacional": 18,
        "Previsão": 13,
        "Último Retorno": 16,
        "Observação": 22,
        "Risco": 11,
        "Ação Necessária": 24,
        # Compatibilidade com versões v5/v5.1, caso alguma aba antiga seja reformatada.
        "Serviço / Mobyan": 22,
        "Risco / Ação": 26,
    }

    colunas_com_quebra = {
        "Prestador",
        "Cliente",
        "Serviço",
        "Técnico",
        "Status Operacional",
        "Previsão",
        "Último Retorno",
        "Observação",
        "Ação Necessária",
        "Serviço / Mobyan",
        "Risco / Ação",
    }

    colunas_centralizadas = {
        "SITUAÇÃO",
        "Chamado",
        "Status Mobyan",
        "Data Limite",
        "Risco",
    }

    larguras_anteriores = layout_anterior.get("larguras", {}) if isinstance(layout_anterior, dict) else {}

    for col in range(1, max_col + 1):
        letra = get_column_letter(col)
        cabecalho = limpar_valor(ws.cell(row=1, column=col).value)

        largura_padrao = larguras_operacionais.get(cabecalho, 14)
        largura_anterior = larguras_anteriores.get(cabecalho)

        # Se o usuário ajustou a largura manualmente no arquivo anterior, preserva essa largura.
        # Limites evitam que uma coluna fique pequena/inutilizável ou gigante demais.
        if largura_anterior:
            largura_final = min(max(float(largura_anterior), 6), 45)
        else:
            largura_final = largura_padrao

        ws.column_dimensions[letra].width = largura_final

        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cabecalho in colunas_com_quebra:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            elif cabecalho in colunas_centralizadas:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=False,
                )
            else:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=False,
                )

    # Visual padrão em 120% para leitura confortável.
    zoom_anterior = layout_anterior.get("zoom") if isinstance(layout_anterior, dict) else None
    if zoom_anterior:
        try:
            zoom = int(zoom_anterior)
            zoom = min(max(zoom, 80), 140)
        except Exception:
            zoom = 120
    else:
        zoom = 120

    ws.sheet_view.zoomScale = zoom
    ws.sheet_view.zoomScaleNormal = zoom
    ws.sheet_view.showGridLines = False

    fonte_dados = Font(name="Arial", size=10)
    fonte_cabecalho_compacta = Font(name="Arial", size=9, color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.font = fonte_cabecalho_compacta
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).font = fonte_dados

    ws.row_dimensions[1].height = 24

    # Altura suficiente para duas linhas sem deixar a planilha pesada.
    altura_anterior = layout_anterior.get("altura_linhas") if isinstance(layout_anterior, dict) else None
    if altura_anterior:
        try:
            altura_linhas = min(max(float(altura_anterior), 24), 60)
        except Exception:
            altura_linhas = 38
    else:
        altura_linhas = 38

    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = altura_linhas

    wb.save(caminho_xlsx)


def gerar_planilha_pendencias(caminho_csv_completo):
    print("Gerando planilha de pendências do dia...")

    PASTA_PENDENCIAS.mkdir(parents=True, exist_ok=True)

    df_bruto = ler_relatorio_csv(caminho_csv_completo)
    df_bruto = padronizar_colunas_pendencias(df_bruto)

    coluna_data_limite = encontrar_coluna_data_limite(df_bruto)

    if coluna_data_limite != "Data Limite":
        df_bruto["Data Limite"] = df_bruto[coluna_data_limite]

    hoje = date.today()

    df_chamados = preparar_dataframe_final(
        df=df_bruto.copy(),
        hoje=hoje,
    )

    df_pendencias = df_chamados[
        df_chamados["SITUAÇÃO"].isin(["VENCIDA", "VENCE HOJE"])
    ].copy()

    acompanhamento_anterior = ler_acompanhamento_anterior()
    layout_acompanhamento_anterior = ler_layout_acompanhamento_anterior()
    df_acompanhamento = criar_dataframe_acompanhamento(
        df_pendencias=df_pendencias,
        acompanhamento_anterior=acompanhamento_anterior,
    )

    imagens_por_prestador = gerar_imagens_por_prestador(df_pendencias)

    df_envios = criar_dataframe_envios(df_pendencias, imagens_por_prestador)

    caminho_saida = ARQUIVO_FINAL_PENDENCIAS

    criar_backup_planilha_atual()

    if caminho_saida.exists():
        try:
            caminho_saida.unlink()
        except Exception:
            pass

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        criar_aba_resumo(writer, df_pendencias, df_acompanhamento)
        df_pendencias.to_excel(writer, index=False, sheet_name="Pendências")
        df_acompanhamento.to_excel(writer, index=False, sheet_name="Acompanhamento")
        df_chamados.to_excel(writer, index=False, sheet_name="Chamados")
        df_envios.to_excel(writer, index=False, sheet_name="Envios")

    formatar_planilha_dados(caminho_saida, "Pendências")
    formatar_aba_acompanhamento(caminho_saida, layout_acompanhamento_anterior)
    formatar_planilha_dados(caminho_saida, "Chamados")
    formatar_aba_envios(caminho_saida)
    formatar_aba_resumo(caminho_saida)

    print(f"Planilha final salva em: {caminho_saida}")
    print(f"Total de pendências encontradas: {len(df_pendencias)}")
    print(f"Total de OSs em acompanhamento: {len(df_acompanhamento)}")
    print(f"Total de chamados no relatório completo: {len(df_chamados)}")
    print(f"Imagens por prestador salvas em: {PASTA_IMAGENS_PRESTADOR}")

    return caminho_saida

def formatar_planilha_dados(caminho_xlsx, nome_aba):
    wb = load_workbook(caminho_xlsx)

    if nome_aba not in wb.sheetnames:
        wb.save(caminho_xlsx)
        return

    ws = wb[nome_aba]

    max_row = ws.max_row
    max_col = ws.max_column

    if max_row <= 1:
        wb.save(caminho_xlsx)
        return

    preenchimento_cabecalho = PatternFill("solid", fgColor="1F4E78")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)

    preenchimento_vencida = PatternFill("solid", fgColor="F4CCCC")
    preenchimento_vence_hoje = PatternFill("solid", fgColor="FFF2CC")
    preenchimento_futura = PatternFill("solid", fgColor="D9EAD3")
    preenchimento_sem_data = PatternFill("solid", fgColor="D9D9D9")

    preenchimento_situacao_vencida = PatternFill("solid", fgColor="CC0000")
    preenchimento_situacao_hoje = PatternFill("solid", fgColor="F1C232")
    preenchimento_situacao_futura = PatternFill("solid", fgColor="70AD47")
    preenchimento_situacao_sem_data = PatternFill("solid", fgColor="808080")

    preenchimento_falta_abonar = PatternFill("solid", fgColor="FF0000")
    fonte_falta_abonar = Font(color="FFFFFF", bold=True)

    preenchimento_base = PatternFill("solid", fgColor="D9EAD3")
    fonte_base = Font(color="000000", bold=True)

    fonte_branca_negrito = Font(color="FFFFFF", bold=True)
    fonte_preta_negrito = Font(color="000000", bold=True)

    borda_fina = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for cell in ws[1]:
        cell.fill = preenchimento_cabecalho
        cell.font = fonte_cabecalho
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda_fina

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    headers = [cell.value for cell in ws[1]]

    col_situacao = headers.index("SITUAÇÃO") + 1 if "SITUAÇÃO" in headers else None
    col_data_limite = headers.index("Data Limite") + 1 if "Data Limite" in headers else None
    col_justificativa = (
        headers.index("Justificativa do Abono") + 1
        if "Justificativa do Abono" in headers
        else None
    )
    col_motivo_base = headers.index("Motivo Base") + 1 if "Motivo Base" in headers else None
    col_observacao_base = headers.index("Observação Base") + 1 if "Observação Base" in headers else None

    for row in range(2, max_row + 1):
        situacao = ws.cell(row=row, column=col_situacao).value if col_situacao else ""

        if situacao == "VENCIDA":
            preenchimento_linha = preenchimento_vencida
            preenchimento_situacao = preenchimento_situacao_vencida
            fonte_situacao = fonte_branca_negrito
        elif situacao == "VENCE HOJE":
            preenchimento_linha = preenchimento_vence_hoje
            preenchimento_situacao = preenchimento_situacao_hoje
            fonte_situacao = fonte_preta_negrito
        elif situacao == "FUTURA":
            preenchimento_linha = preenchimento_futura
            preenchimento_situacao = preenchimento_situacao_futura
            fonte_situacao = fonte_branca_negrito
        else:
            preenchimento_linha = preenchimento_sem_data
            preenchimento_situacao = preenchimento_situacao_sem_data
            fonte_situacao = fonte_branca_negrito

        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = preenchimento_linha
            cell.border = borda_fina
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        if col_situacao:
            cell_situacao = ws.cell(row=row, column=col_situacao)
            cell_situacao.fill = preenchimento_situacao
            cell_situacao.font = fonte_situacao
            cell_situacao.alignment = Alignment(horizontal="center", vertical="center")

        if col_data_limite:
            ws.cell(row=row, column=col_data_limite).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        if col_justificativa:
            cell_justificativa = ws.cell(row=row, column=col_justificativa)

            if str(cell_justificativa.value).strip().upper() == "FALTA ABONAR":
                cell_justificativa.fill = preenchimento_falta_abonar
                cell_justificativa.font = fonte_falta_abonar
                cell_justificativa.alignment = Alignment(horizontal="center", vertical="center")

        if col_motivo_base:
            cell_motivo_base = ws.cell(row=row, column=col_motivo_base)

            if str(cell_motivo_base.value).strip() != "":
                cell_motivo_base.fill = preenchimento_base
                cell_motivo_base.font = fonte_base

        if col_observacao_base:
            cell_observacao_base = ws.cell(row=row, column=col_observacao_base)

            if str(cell_observacao_base.value).strip() != "":
                cell_observacao_base.fill = preenchimento_base

    larguras_personalizadas = {
        "SITUAÇÃO": 14,
        "Chamado": 12,
        "Numero Referencia": 18,
        "Contratante": 14,
        "Serviço": 18,
        "Status": 18,
        "Data Limite": 16,
        "Cliente": 36,
        "CNPJ / CPF": 18,
        "Cidade": 20,
        "Técnico": 28,
        "Prestador": 28,
        "Justificativa do Abono": 32,
        "Motivo Base": 26,
        "Observação Base": 42,
    }

    for col in range(1, max_col + 1):
        letra_coluna = get_column_letter(col)
        cabecalho = ws.cell(row=1, column=col).value

        if cabecalho in larguras_personalizadas:
            ws.column_dimensions[letra_coluna].width = larguras_personalizadas[cabecalho]
        else:
            maior_tamanho = 0

            for row in range(1, min(max_row, 200) + 1):
                valor = ws.cell(row=row, column=col).value

                if valor is not None:
                    maior_tamanho = max(maior_tamanho, len(str(valor)))

            largura = min(max(maior_tamanho + 2, 12), 38)
            ws.column_dimensions[letra_coluna].width = largura

    ws.row_dimensions[1].height = 24

    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = 18

    wb.save(caminho_xlsx)


def formatar_aba_envios(caminho_xlsx):
    wb = load_workbook(caminho_xlsx)

    if "Envios" not in wb.sheetnames:
        wb.save(caminho_xlsx)
        return

    ws = wb["Envios"]

    max_row = ws.max_row
    max_col = ws.max_column

    preenchimento_cabecalho = PatternFill("solid", fgColor="1F4E78")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)

    preenchimento_pronto = PatternFill("solid", fgColor="D9EAD3")
    preenchimento_atencao = PatternFill("solid", fgColor="FFF2CC")
    preenchimento_erro = PatternFill("solid", fgColor="F4CCCC")

    borda_fina = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for cell in ws[1]:
        cell.fill = preenchimento_cabecalho
        cell.font = fonte_cabecalho
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda_fina

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    headers = [cell.value for cell in ws[1]]

    col_status = headers.index("Status Envio") + 1 if "Status Envio" in headers else None
    col_mensagem = headers.index("Mensagem") + 1 if "Mensagem" in headers else None
    col_mensagem_manha = headers.index("Mensagem Manhã") + 1 if "Mensagem Manhã" in headers else None
    col_mensagem_acompanhamento = (
        headers.index("Mensagem Acompanhamento") + 1
        if "Mensagem Acompanhamento" in headers
        else None
    )

    for row in range(2, max_row + 1):
        status = ws.cell(row=row, column=col_status).value if col_status else ""

        if status == "Pronto para envio":
            preenchimento_linha = preenchimento_pronto
        elif status in ["Sem pendências", "Não marcado para envio"]:
            preenchimento_linha = preenchimento_atencao
        else:
            preenchimento_linha = preenchimento_erro

        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = preenchimento_linha
            cell.border = borda_fina
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        for col_mensagem_texto in [col_mensagem_manha, col_mensagem_acompanhamento]:
            if col_mensagem_texto:
                ws.cell(row=row, column=col_mensagem_texto).alignment = Alignment(
                    vertical="top",
                    horizontal="left",
                    wrap_text=True
                )

        if col_mensagem:
            ws.cell(row=row, column=col_mensagem).alignment = Alignment(
                vertical="top",
                horizontal="left",
                wrap_text=True
            )

    larguras = {
        "Prestador": 34,
        "Responsável": 24,
        "WhatsApp": 20,
        "Enviar": 12,
        "Imagem": 70,
        "Total Pendências": 18,
        "Vencidas": 12,
        "Vence Hoje": 14,
        "Status Envio": 24,
        "Mensagem": 18,
        "Mensagem Manhã": 72,
        "Mensagem Acompanhamento": 86,
        "Observação": 36,
    }

    for col in range(1, max_col + 1):
        letra = get_column_letter(col)
        cabecalho = ws.cell(row=1, column=col).value

        if cabecalho in larguras:
            ws.column_dimensions[letra].width = larguras[cabecalho]
        else:
            ws.column_dimensions[letra].width = 18

    if col_mensagem:
        letra_mensagem = get_column_letter(col_mensagem)
        ws.column_dimensions[letra_mensagem].hidden = True

    ws.row_dimensions[1].height = 24

    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = 70

    wb.save(caminho_xlsx)



def formatar_aba_resumo(caminho_xlsx):
    wb = load_workbook(caminho_xlsx)

    if "Resumo" not in wb.sheetnames:
        wb.save(caminho_xlsx)
        return

    ws = wb["Resumo"]

    azul = PatternFill("solid", fgColor="1F4E78")
    azul_claro = PatternFill("solid", fgColor="D9EAF7")
    vermelho = PatternFill("solid", fgColor="CC0000")
    vermelho_claro = PatternFill("solid", fgColor="F4CCCC")
    amarelo = PatternFill("solid", fgColor="F1C232")
    amarelo_claro = PatternFill("solid", fgColor="FFF2CC")
    verde = PatternFill("solid", fgColor="70AD47")
    verde_claro = PatternFill("solid", fgColor="D9EAD3")
    cinza = PatternFill("solid", fgColor="D9E2F3")
    branco = PatternFill("solid", fgColor="FFFFFF")

    fonte_branca = Font(color="FFFFFF", bold=True)
    fonte_titulo = Font(color="FFFFFF", bold=True, size=13)
    fonte_card_numero = Font(color="000000", bold=True, size=16)
    fonte_negrito = Font(bold=True)

    borda = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    ws.insert_rows(1, 1)

    for row in range(1, 20):
        ws.row_dimensions[row].height = 22

    ws["A1"] = "PENDÊNCIAS DE HOJE"
    ws.merge_cells("A1:H1")
    ws["A1"].fill = azul
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Indicadores do dia"
    ws.merge_cells("A2:C2")
    ws["A2"].fill = cinza
    ws["A2"].font = fonte_negrito
    ws["A2"].alignment = Alignment(horizontal="center")

    for row in range(5, 11):
        indicador = ws.cell(row=row, column=1).value

        if indicador == "Total de Pendências":
            fill_nome = azul
            fill_valor = azul_claro
        elif indicador in ["Vencidas", "Falta Abonar", "Em Risco"]:
            fill_nome = vermelho
            fill_valor = vermelho_claro
        elif indicador == "Vence Hoje":
            fill_nome = amarelo
            fill_valor = amarelo_claro
        elif indicador == "Controladas":
            fill_nome = verde
            fill_valor = verde_claro
        else:
            fill_nome = azul
            fill_valor = azul_claro

        ws.cell(row=row, column=1).fill = fill_nome
        ws.cell(row=row, column=1).font = fonte_branca
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.cell(row=row, column=2).fill = fill_valor
        ws.cell(row=row, column=2).font = fonte_card_numero
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=row, column=3).fill = fill_valor
        ws.cell(row=row, column=3).font = fonte_negrito
        ws.cell(row=row, column=3).number_format = "0.0%"
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 4):
            ws.cell(row=row, column=col).border = borda

        ws.row_dimensions[row].height = 26

    # Linha 4: cabeçalho técnico escrito pelo pandas (Indicador/Quantidade/%).
    # Os cards acima já são autoexplicativos, então essa linha só é ruído — fica oculta.
    ws.row_dimensions[4].hidden = True

    titulos_tabelas = [
        ("A13:C13", "Top 5 Prestadores"),
        ("F13:H13", "Top 5 Cidades"),
    ]

    for intervalo, titulo in titulos_tabelas:
        celula_inicio = intervalo.split(":")[0]
        ws[celula_inicio] = titulo
        ws.merge_cells(intervalo)
        ws[celula_inicio].fill = azul
        ws[celula_inicio].font = fonte_titulo
        ws[celula_inicio].alignment = Alignment(horizontal="center", vertical="center")

    header_row = 14

    for col in [1, 2, 3, 6, 7, 8]:
        cell = ws.cell(row=header_row, column=col)
        cell.fill = verde
        cell.font = fonte_branca
        cell.border = borda
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(15, 20):
        for col in [1, 2, 3, 6, 7, 8]:
            cell = ws.cell(row=row, column=col)
            cell.fill = branco
            cell.border = borda
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if col in [3, 8]:
                cell.number_format = "0.0%"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 3
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12

    ws.sheet_view.showGridLines = False

    ordem = ["Resumo", "Pendências", "Acompanhamento", "Chamados", "Envios"]

    novas_sheets = []

    for nome in ordem:
        if nome in wb.sheetnames:
            novas_sheets.append(wb[nome])

    for sheet in wb.worksheets:
        if sheet.title not in ordem:
            novas_sheets.append(sheet)

    wb._sheets = novas_sheets

    wb.save(caminho_xlsx)

def exportar_relatorio_completo():
    validar_configuracoes()

    PASTA_DOWNLOADS.mkdir(parents=True, exist_ok=True)
    PASTA_PENDENCIAS.mkdir(parents=True, exist_ok=True)
    PASTA_IMAGENS_PRESTADOR.mkdir(parents=True, exist_ok=True)
    PASTA_LOGS.mkdir(parents=True, exist_ok=True)

    nome_csv_temporario = f"relatorio_completo_mobyan_temp_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.csv"
    caminho_csv_temporario = PASTA_DOWNLOADS / nome_csv_temporario

    with sync_playwright() as p:
        navegador = p.chromium.launch(headless=False)

        contexto = navegador.new_context(
            accept_downloads=True,
            viewport={"width": 1366, "height": 768}
        )

        page = contexto.new_page()

        print("Abrindo sistema da Mobyan...")
        page.goto(MOBYAN_URL)

        fazer_login(page)

        finalizar_sessao_anterior_se_aparecer(page)

        print("Aguardando entrada no sistema...")
        page.wait_for_timeout(5000)

        print("Abrindo menu de relatórios...")
        abrir_relatorio_ordem_servico_sla(page)

        print("Aguardando tela do relatório carregar...")
        page.wait_for_timeout(5000)

        frame_relatorio = page.locator('iframe[name^="loadext-"]').content_frame

        selecionar_status(frame_relatorio)
        selecionar_prestadores(frame_relatorio)
        selecionar_estado_rs(frame_relatorio)

        print("Pesquisando relatório...")
        frame_relatorio.get_by_role("img", name="Pesquisar").click()

        print("Aguardando resultado da pesquisa...")
        page.wait_for_timeout(8000)

        print("Exportando CSV temporário...")
        frame_lista = page.locator('iframe[name^="listext-"]').content_frame

        with page.expect_download() as download_info:
            frame_lista.get_by_role("img", name="Exportar CSV").click()

        download = download_info.value
        download.save_as(caminho_csv_temporario)

        print(f"CSV temporário salvo em: {caminho_csv_temporario}")

        contexto.close()
        navegador.close()

    caminho_pendencias = gerar_planilha_pendencias(caminho_csv_temporario)

    limpar_csv_temporario(caminho_csv_temporario)

    print("Abrindo planilha final...")
    abrir_caminho(caminho_pendencias)

    print("Processo finalizado com sucesso!")
    print(f"Planilha final: {caminho_pendencias}")

    return caminho_pendencias


if __name__ == "__main__":
    exportar_relatorio_completo()