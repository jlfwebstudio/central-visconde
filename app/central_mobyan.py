import ctypes
import os
import sys
import queue
import subprocess
import tempfile
import threading
import tkinter as tk
from datetime import date, datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from PIL import Image, ImageTk
except Exception:
    Image = None
    ImageTk = None

try:
    from gestao_rotas import abrir_gestao_rotas, obter_resumo_rotas
except Exception:
    abrir_gestao_rotas = None
    obter_resumo_rotas = None

try:
    from configurar_whatsapp import abrir_configurar_whatsapp
except Exception:
    abrir_configurar_whatsapp = None

from autenticacao import (
    baixar_atualizacao,
    carregar_sessao_local,
    garantir_sessao_valida,
    limpar_sessao_local,
    verificar_atualizacao,
)
from caminho_base import BASE_DIR, FROZEN, RECURSOS_DIR

SCRIPT_GERAR_PENDENCIAS = RECURSOS_DIR / "app" / "exportador_mobyan.py"
SCRIPT_ENVIAR_WHATSAPP = RECURSOS_DIR / "app" / "enviar_whatsapp.py"
SCRIPT_GERAR_ROTEIRIZACAO = RECURSOS_DIR / "app" / "gerar_roteirizacao.py"
SCRIPT_GERAR_PDFS = RECURSOS_DIR / "app" / "gerar_pdfs.py"
SCRIPT_ANALISAR_ABONOS = RECURSOS_DIR / "app" / "analisar_abonos_ogea.py"
SCRIPT_GERAR_OS_VISCONDE_TESTE = RECURSOS_DIR / "app" / "gerar_os_visconde_teste.py"

ARQUIVO_PLANILHA = BASE_DIR / "outputs" / "pendencias_do_dia" / "pendencias_do_dia_atual.xlsx"
PASTA_FILTROS_PAINEL = BASE_DIR / "outputs" / "pendencias_do_dia" / "filtros_painel"
ARQUIVO_ROTEIRIZACAO = BASE_DIR / "outputs" / "roteirizacao" / "roteirizacao_atual.xlsx"
ARQUIVO_ABONOS = BASE_DIR / "outputs" / "abonos_ogea" / "analise_abonos_ogea_atual.xlsx"
ARQUIVO_REGRAS_ROTEIRIZACAO = BASE_DIR / "bases" / "regras_roteirizacao.xlsx"
PASTA_IMAGENS = BASE_DIR / "outputs" / "por_prestador_imagens"
PASTA_PDFS = BASE_DIR / "outputs" / "pdfs"
PASTA_ABONOS = BASE_DIR / "outputs" / "abonos_ogea"
PASTA_OS_VISCONDE_TESTE = BASE_DIR / "outputs" / "os_visconde_teste"
PASTA_LOGS = BASE_DIR / "logs"
ARQUIVO_LOGO = RECURSOS_DIR / "assets" / "logo_visconde.png"
ARQUIVO_ICONE_PNG = RECURSOS_DIR / "assets" / "logo_visconde_app.png"
ARQUIVO_ICONE_ICO = RECURSOS_DIR / "assets" / "logo_visconde.ico"



def configurar_identidade_windows():
    """Mantém a janela agrupada no atalho premium do ViscondeApp."""
    if os.name != "nt":
        return
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
            "Visconde.CentralVisconde.Windows"
        )
    except Exception:
        pass


def configurar_icone_janela(root):
    icone_carregado = False
    if ARQUIVO_ICONE_PNG.exists() and Image is not None and ImageTk is not None:
        try:
            imagem = Image.open(ARQUIVO_ICONE_PNG)
            imagem.thumbnail((256, 256), Image.Resampling.LANCZOS)
            icone = ImageTk.PhotoImage(imagem)
            root._icone_visconde = icone
            root.iconphoto(True, icone)
            icone_carregado = True
        except Exception:
            icone_carregado = False

    if os.name == "nt" and ARQUIVO_ICONE_ICO.exists():
        try:
            root.iconbitmap(default=str(ARQUIVO_ICONE_ICO))
            icone_carregado = True
        except Exception:
            pass

    return icone_carregado


def obter_python_console():
    executavel = Path(sys.executable)
    if os.name == "nt" and executavel.name.lower() == "pythonw.exe":
        python_console = executavel.with_name("python.exe")
        if python_console.exists():
            return python_console
    return executavel


PYTHON_AUTOMACOES = obter_python_console()


def abrir_caminho(caminho):
    caminho = Path(caminho).resolve()
    if os.name == "nt":
        os.startfile(str(caminho))
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(caminho)])
    else:
        subprocess.Popen(["xdg-open", str(caminho)])


from estilo_visconde import (
    ALTURA_CARD_CHEIO,
    ALTURA_CARD_PAR,
    COR_AZUL,
    COR_AZUL_HOVER,
    COR_BORDA,
    COR_BRANCO,
    COR_CARD,
    COR_CARD_HOVER,
    COR_CINZA,
    COR_DOURADO,
    COR_DOURADO_ESCURO,
    COR_DOURADO_HOVER,
    COR_FUNDO,
    COR_FUNDO_2,
    COR_LARANJA,
    COR_LARANJA_HOVER,
    COR_TEXTO_FRACO,
    COR_TEXTO_SECUNDARIO,
    COR_VERDE,
    COR_VERDE_HOVER,
    COR_VERMELHO,
    COR_VERMELHO_HOVER,
    LARGURA_CARD_CHEIO,
    LARGURA_CARD_PAR,
)


class NavButton(tk.Frame):
    def __init__(self, master, texto, comando, largura=210):
        super().__init__(master, bg=COR_FUNDO)
        self.comando = comando
        self.selecionado = False
        self.container = tk.Frame(
            self,
            bg=COR_FUNDO,
            width=largura,
            height=46,
            cursor="hand2",
            highlightthickness=0,
        )
        self.container.pack(fill="x")
        self.container.pack_propagate(False)
        self.indicador = tk.Frame(self.container, bg=COR_FUNDO, width=4)
        self.indicador.pack(side="left", fill="y")
        self.label = tk.Label(
            self.container,
            text=texto,
            bg=COR_FUNDO,
            fg=COR_TEXTO_SECUNDARIO,
            font=("Arial", 11, "bold"),
            anchor="w",
            padx=15,
            cursor="hand2",
        )
        self.label.pack(side="left", fill="both", expand=True)
        for widget in [self.container, self.indicador, self.label]:
            widget.bind("<Button-1>", self.clicar)
            widget.bind("<Enter>", self.entrar)
            widget.bind("<Leave>", self.sair)

    def clicar(self, event=None):
        self.comando()

    def entrar(self, event=None):
        if not self.selecionado:
            self.container.config(bg=COR_CARD)
            self.label.config(bg=COR_CARD, fg=COR_BRANCO)

    def sair(self, event=None):
        if not self.selecionado:
            self.container.config(bg=COR_FUNDO)
            self.label.config(bg=COR_FUNDO, fg=COR_TEXTO_SECUNDARIO)

    def selecionar(self, valor):
        self.selecionado = valor
        if valor:
            self.container.config(bg=COR_CARD)
            self.label.config(bg=COR_CARD, fg=COR_DOURADO)
            self.indicador.config(bg=COR_DOURADO)
        else:
            self.container.config(bg=COR_FUNDO)
            self.label.config(bg=COR_FUNDO, fg=COR_TEXTO_SECUNDARIO)
            self.indicador.config(bg=COR_FUNDO)


class ActionCard(tk.Frame):
    def __init__(
        self,
        master,
        titulo,
        descricao,
        comando,
        destaque=COR_DOURADO,
        destaque_hover=COR_DOURADO_HOVER,
        largura=330,
        altura=112,
    ):
        super().__init__(master, bg=COR_FUNDO_2)
        self.comando = comando
        self.ativo = True
        self.destaque = destaque
        self.destaque_hover = destaque_hover
        self.container = tk.Frame(
            self,
            bg=COR_CARD,
            width=largura,
            height=altura,
            cursor="hand2",
            highlightbackground=COR_BORDA,
            highlightthickness=1,
        )
        self.container.pack(fill="both", expand=True)
        self.container.pack_propagate(False)
        self.faixa = tk.Frame(self.container, bg=destaque, height=5)
        self.faixa.pack(fill="x")
        self.titulo = tk.Label(
            self.container,
            text=titulo,
            bg=COR_CARD,
            fg=COR_BRANCO,
            font=("Arial", 15, "bold"),
            anchor="w",
            cursor="hand2",
        )
        self.titulo.pack(fill="x", padx=18, pady=(16, 4))
        self.descricao = tk.Label(
            self.container,
            text=descricao,
            bg=COR_CARD,
            fg=COR_TEXTO_SECUNDARIO,
            font=("Arial", 10),
            anchor="w",
            justify="left",
            wraplength=max(230, largura - 42),
            cursor="hand2",
        )
        self.descricao.pack(fill="x", padx=18)
        for widget in [self.container, self.faixa, self.titulo, self.descricao]:
            widget.bind("<Button-1>", self.clicar)
            widget.bind("<Enter>", self.entrar)
            widget.bind("<Leave>", self.sair)

    def clicar(self, event=None):
        if self.ativo and self.comando:
            self.comando()

    def entrar(self, event=None):
        if self.ativo:
            self.container.config(bg=COR_CARD_HOVER)
            self.titulo.config(bg=COR_CARD_HOVER)
            self.descricao.config(bg=COR_CARD_HOVER)
            self.faixa.config(bg=self.destaque_hover)

    def sair(self, event=None):
        if self.ativo:
            self.container.config(bg=COR_CARD)
            self.titulo.config(bg=COR_CARD)
            self.descricao.config(bg=COR_CARD)
            self.faixa.config(bg=self.destaque)

    def set_enabled(self, enabled=True):
        self.ativo = enabled
        cor = self.destaque if enabled else COR_CINZA
        cursor = "hand2" if enabled else "arrow"
        self.container.config(cursor=cursor, bg=COR_CARD if enabled else "#252525")
        self.titulo.config(cursor=cursor, bg=COR_CARD if enabled else "#252525")
        self.descricao.config(cursor=cursor, bg=COR_CARD if enabled else "#252525")
        self.faixa.config(bg=cor)

    def definir_conteudo(self, titulo, descricao, comando, destaque=None, destaque_hover=None):
        self.comando = comando
        self.titulo.config(text=titulo)
        self.descricao.config(text=descricao)
        if destaque:
            self.destaque = destaque
            self.destaque_hover = destaque_hover or destaque
            self.faixa.config(bg=self.destaque)


class SmallButton(tk.Frame):
    def __init__(self, master, texto, comando, largura=170, altura=42):
        super().__init__(master, bg=COR_FUNDO_2)
        self.comando = comando
        self.container = tk.Frame(
            self,
            bg=COR_CARD,
            height=altura,
            width=largura,
            cursor="hand2",
            highlightbackground=COR_BORDA,
            highlightthickness=1,
        )
        self.container.pack(fill="both", expand=True)
        self.container.pack_propagate(False)
        self.label = tk.Label(
            self.container,
            text=texto,
            bg=COR_CARD,
            fg=COR_BRANCO,
            font=("Arial", 9, "bold"),
            cursor="hand2",
            wraplength=largura - 16,
            justify="center",
        )
        self.label.pack(expand=True)
        for widget in [self.container, self.label]:
            widget.bind("<Button-1>", lambda event: self.comando())
            widget.bind("<Enter>", self.entrar)
            widget.bind("<Leave>", self.sair)

    def entrar(self, event=None):
        self.container.config(bg=COR_CARD_HOVER)
        self.label.config(bg=COR_CARD_HOVER, fg=COR_DOURADO)

    def sair(self, event=None):
        self.container.config(bg=COR_CARD)
        self.label.config(bg=COR_CARD, fg=COR_BRANCO)


class MetricCard(tk.Frame):
    def __init__(self, master, titulo, valor="0", cor=COR_DOURADO, subtitulo="", tamanho_valor=23, comando=None):
        cursor = "hand2" if comando else "arrow"
        super().__init__(
            master,
            bg=COR_CARD,
            highlightbackground=COR_BORDA,
            highlightthickness=1,
            cursor=cursor,
        )
        self.comando = comando
        self.label_titulo = tk.Label(
            self,
            text=titulo,
            bg=COR_CARD,
            fg=COR_TEXTO_SECUNDARIO,
            font=("Arial", 10, "bold"),
            cursor=cursor,
        )
        self.label_titulo.pack(pady=(12, 4))
        self.label_valor = tk.Label(
            self,
            text=valor,
            bg=COR_CARD,
            fg=cor,
            font=("Arial", tamanho_valor, "bold"),
            wraplength=220,
            justify="center",
            cursor=cursor,
        )
        self.label_valor.pack(pady=(0, 2 if subtitulo else 12))
        self.label_subtitulo = tk.Label(
            self,
            text=subtitulo,
            bg=COR_CARD,
            fg=COR_TEXTO_FRACO,
            font=("Arial", 9),
            cursor=cursor,
        )
        if subtitulo:
            self.label_subtitulo.pack(pady=(0, 12))
        if comando:
            for widget in (self, self.label_titulo, self.label_valor, self.label_subtitulo):
                widget.bind("<Button-1>", lambda event: self.comando())
                widget.bind("<Enter>", self.entrar)
                widget.bind("<Leave>", self.sair)

    def entrar(self, event=None):
        self.config(highlightbackground=COR_DOURADO)
        for widget in (self, self.label_titulo, self.label_valor, self.label_subtitulo):
            widget.config(bg=COR_CARD_HOVER)

    def sair(self, event=None):
        self.config(highlightbackground=COR_BORDA)
        for widget in (self, self.label_titulo, self.label_valor, self.label_subtitulo):
            widget.config(bg=COR_CARD)

    def atualizar(self, valor, cor=None, subtitulo=None):
        self.label_valor.config(text=str(valor))
        if cor:
            self.label_valor.config(fg=cor)
        if subtitulo is not None:
            if subtitulo:
                self.label_subtitulo.config(text=subtitulo)
                if not self.label_subtitulo.winfo_manager():
                    self.label_valor.pack_configure(pady=(0, 2))
                    self.label_subtitulo.pack(pady=(0, 12))
            else:
                self.label_subtitulo.config(text="")
                if self.label_subtitulo.winfo_manager():
                    self.label_subtitulo.pack_forget()
                    self.label_valor.pack_configure(pady=(0, 12))


class CentralVisconde:
    def __init__(self, root):
        self.root = root
        self.root.title("ViscondeApp")
        configurar_icone_janela(self.root)
        largura_tela = self.root.winfo_screenwidth()
        altura_tela = self.root.winfo_screenheight()
        largura = min(1460, max(1120, largura_tela - 70))
        altura = min(900, max(720, altura_tela - 90))
        self.root.geometry(f"{largura}x{altura}")
        self.root.minsize(min(1100, largura), min(700, altura))
        self.root.configure(bg=COR_FUNDO)

        self.processo_rodando = False
        self.log_queue = queue.Queue()
        self.acao_atual = None
        self.janela_gestao_rotas = None
        self.janela_configurar_whatsapp = None
        self.botoes_processo = []
        self.paginas = {}
        self.nav_botoes = {}
        self.logo_image = None
        self.ultimo_prestador_top = None

        PASTA_LOGS.mkdir(parents=True, exist_ok=True)
        PASTA_ABONOS.mkdir(parents=True, exist_ok=True)

        self.montar_interface()
        self.mostrar_pagina("inicio")
        self.atualizar_resumos()
        self.root.after(200, self.processar_logs)
        # A Gestão Inteligente de Rotas não é aberta na inicialização.
        # A verificação automática ocorre somente após Gerar Roteirização.

    def montar_interface(self):
        self.sidebar = tk.Frame(self.root, bg=COR_FUNDO, width=245)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        self.main = tk.Frame(self.root, bg=COR_FUNDO_2)
        self.main.pack(side="left", fill="both", expand=True)

        self.montar_sidebar()
        self.montar_main()

    def montar_sidebar(self):
        logo_container = tk.Frame(self.sidebar, bg=COR_FUNDO, height=180)
        logo_container.pack(fill="x", padx=14, pady=(16, 10))
        logo_container.pack_propagate(False)

        carregou_logo = False
        if Image is not None and ImageTk is not None and ARQUIVO_LOGO.exists():
            try:
                imagem = Image.open(ARQUIVO_LOGO).convert("RGB")
                imagem.thumbnail((205, 135), Image.Resampling.LANCZOS)
                self.logo_image = ImageTk.PhotoImage(imagem)
                label_logo = tk.Label(logo_container, image=self.logo_image, bg=COR_FUNDO)
                label_logo.pack(expand=True)
                carregou_logo = True
            except Exception:
                carregou_logo = False

        if not carregou_logo:
            tk.Label(
                logo_container,
                text="VISCONDE",
                bg=COR_FUNDO,
                fg=COR_DOURADO,
                font=("Arial", 25, "bold"),
            ).pack(expand=True)
            tk.Label(
                logo_container,
                text="INSTALAÇÃO E MANUTENÇÃO",
                bg=COR_FUNDO,
                fg=COR_BRANCO,
                font=("Arial", 8, "bold"),
            ).pack()

        tk.Frame(self.sidebar, bg=COR_BORDA, height=1).pack(fill="x", padx=18, pady=(0, 15))

        itens = [
            ("inicio", "Painel"),
            ("operacao", "Operação"),
            ("whatsapp", "WhatsApp"),
            ("roteiros", "Roteiros"),
            ("ferramentas", "Abonos"),
        ]
        for chave, texto in itens:
            botao = NavButton(self.sidebar, texto, lambda c=chave: self.mostrar_pagina(c))
            botao.pack(fill="x", padx=8, pady=2)
            self.nav_botoes[chave] = botao

        rodape = tk.Frame(self.sidebar, bg=COR_FUNDO)
        rodape.pack(side="bottom", fill="x", padx=18, pady=18)
        tk.Label(
            rodape,
            text="ViscondeApp",
            bg=COR_FUNDO,
            fg=COR_DOURADO,
            font=("Arial", 10, "bold"),
        ).pack(anchor="w")
        tk.Label(
            rodape,
            text="Automação operacional",
            bg=COR_FUNDO,
            fg=COR_TEXTO_FRACO,
            font=("Arial", 9),
        ).pack(anchor="w", pady=(2, 0))

        sessao_atual = carregar_sessao_local()
        if sessao_atual:
            tk.Label(
                rodape,
                text=f"Conta: {sessao_atual.get('usuario', '')}",
                bg=COR_FUNDO,
                fg=COR_TEXTO_FRACO,
                font=("Arial", 8),
            ).pack(anchor="w", pady=(10, 0))

            link_sair = tk.Label(
                rodape,
                text="Sair da conta",
                bg=COR_FUNDO,
                fg=COR_DOURADO,
                font=("Arial", 9, "underline"),
                cursor="hand2",
            )
            link_sair.pack(anchor="w", pady=(2, 0))
            link_sair.bind("<Button-1>", lambda _evento: self.sair_da_conta())

    def montar_main(self):
        header = tk.Frame(self.main, bg=COR_FUNDO_2)
        header.pack(fill="x", padx=26, pady=(20, 8))
        self.page_title = tk.Label(
            header,
            text="",
            bg=COR_FUNDO_2,
            fg=COR_DOURADO,
            font=("Arial", 25, "bold"),
            anchor="w",
        )
        self.page_title.pack(fill="x")
        self.page_subtitle = tk.Label(
            header,
            text="",
            bg=COR_FUNDO_2,
            fg=COR_TEXTO_SECUNDARIO,
            font=("Arial", 11),
            anchor="w",
        )
        self.page_subtitle.pack(fill="x", pady=(4, 0))

        self.page_host = tk.Frame(self.main, bg=COR_FUNDO_2)
        self.page_host.pack(fill="both", expand=True, padx=26, pady=(8, 10))

        self.criar_pagina_inicio()
        self.criar_pagina_operacao()
        self.criar_pagina_whatsapp()
        self.criar_pagina_roteiros()
        self.criar_pagina_ferramentas()

        status_frame = tk.Frame(self.main, bg=COR_FUNDO_2)
        status_frame.pack(fill="x", padx=26, pady=(0, 8))
        self.status = tk.Label(
            status_frame,
            text="Status: aguardando ação...",
            bg=COR_FUNDO_2,
            fg=COR_DOURADO,
            font=("Arial", 11, "bold"),
            anchor="w",
        )
        self.status.pack(fill="x")

        log_container = tk.Frame(
            self.main,
            bg=COR_CARD,
            height=240,
            highlightbackground=COR_BORDA,
            highlightthickness=1,
        )
        log_container.pack(fill="both", padx=26, pady=(0, 18))
        log_container.pack_propagate(False)
        log_header = tk.Frame(log_container, bg=COR_CARD)
        log_header.pack(fill="x", padx=12, pady=(8, 0))
        tk.Label(
            log_header,
            text="Log da automação",
            bg=COR_CARD,
            fg=COR_DOURADO,
            font=("Arial", 10, "bold"),
        ).pack(side="left")
        tk.Label(
            log_header,
            text="Os processos continuam visíveis aqui em qualquer tela.",
            bg=COR_CARD,
            fg=COR_TEXTO_FRACO,
            font=("Arial", 9),
        ).pack(side="right")
        log_frame = tk.Frame(log_container, bg=COR_CARD)
        log_frame.pack(fill="both", expand=True, padx=12, pady=(6, 12))
        self.log_text = tk.Text(
            log_frame,
            bg="#050505",
            fg="#EDEDED",
            insertbackground=COR_DOURADO,
            font=(("Consolas" if os.name == "nt" else "Menlo"), 10),
            relief="flat",
            wrap="word",
            padx=10,
            pady=9,
        )
        self.log_text.pack(side="left", fill="both", expand=True)
        scroll = tk.Scrollbar(log_frame, command=self.log_text.yview)
        scroll.pack(side="right", fill="y")
        self.log_text.config(yscrollcommand=scroll.set)
        self.escrever_log("ViscondeApp iniciado.\n")
        self.escrever_log("Escolha uma área no menu lateral.\n")

    def nova_pagina(self, chave):
        """Cria uma página rolável: cada seção pode crescer sem sobrepor o rodapé."""
        wrapper = tk.Frame(self.page_host, bg=COR_FUNDO_2)
        wrapper.place(relx=0, rely=0, relwidth=1, relheight=1)

        canvas = tk.Canvas(wrapper, bg=COR_FUNDO_2, highlightthickness=0)
        scrollbar = tk.Scrollbar(wrapper, orient="vertical", command=canvas.yview)
        pagina = tk.Frame(canvas, bg=COR_FUNDO_2)

        pagina.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        janela_id = canvas.create_window((0, 0), window=pagina, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(janela_id, width=e.width))

        def rolar(event):
            passos = int(-1 * (event.delta / 120)) if os.name == "nt" else int(-1 * event.delta)
            canvas.yview_scroll(passos, "units")

        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", rolar))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
        canvas.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.paginas[chave] = wrapper
        return pagina

    def registrar_action(self, botao):
        self.botoes_processo.append(botao)
        return botao

    def criar_atalhos(self, master, itens, largura=134, altura=68):
        linha = tk.Frame(master, bg=COR_FUNDO_2)
        for idx, (rotulo, comando) in enumerate(itens):
            botao = SmallButton(linha, rotulo, comando, largura=largura, altura=altura)
            botao.grid(row=0, column=idx, padx=(0 if idx == 0 else 8, 0), sticky="w")
        return linha

    def criar_pagina_inicio(self):
        pagina = self.nova_pagina("inicio")
        resumo = tk.Frame(pagina, bg=COR_FUNDO_2)
        resumo.pack(fill="x", pady=(2, 14))
        self.card_vencendo_hoje = MetricCard(
            resumo, "OSs vencendo hoje", "0", COR_LARANJA,
            comando=lambda: self.abrir_pendencias_filtradas("vencendo_hoje"),
        )
        self.card_vencendo_hoje.grid(row=0, column=0, padx=(0, 8), sticky="ew")
        self.card_prestador_top = MetricCard(
            resumo, "Maior volume hoje", "—", COR_AZUL, tamanho_valor=16,
            comando=lambda: self.abrir_pendencias_filtradas("prestador_top", self.ultimo_prestador_top),
        )
        self.card_prestador_top.grid(row=0, column=1, padx=8, sticky="ew")
        self.card_criticas = MetricCard(
            resumo, "OSs críticas (5+ dias)", "0", COR_VERMELHO,
            comando=lambda: self.abrir_pendencias_filtradas("criticas"),
        )
        self.card_criticas.grid(row=0, column=2, padx=(8, 0), sticky="ew")
        for col in range(3):
            resumo.grid_columnconfigure(col, weight=1)

        self.criar_atalhos(
            pagina,
            [
                ("Abrir\nPendências", self.abrir_planilha),
                ("Abrir\nRoteirização", self.abrir_roteirizacao),
                ("Abrir\nRoteiros PDF", self.abrir_pasta_pdfs),
                ("Abrir Gestão\nde Rotas", self.abrir_gestao_rotas),
            ],
        ).pack(anchor="w", pady=(2, 12))

        self.adicionar_painel_info(
            pagina,
            "Como funciona esse painel",
            "Os números acima vêm da última planilha de pendências gerada (\"Gerar Pendências\"), sempre desconsiderando envios de bobina, que não contam no SLA mensal: quantas OSs vencem hoje, qual prestador concentra o maior volume de vencimentos hoje e quantas OSs já estão críticas, com 5 dias de atraso ou mais. Gere as pendências de novo ou use \"Atualizar Resumos\" para recalcular.",
            altura=110,
        )

    def criar_pagina_operacao(self):
        pagina = self.nova_pagina("operacao")
        acoes = tk.Frame(pagina, bg=COR_FUNDO_2)
        acoes.pack(fill="x", pady=(2, 14))
        b1 = self.registrar_action(ActionCard(
            acoes,
            "Gerar Pendências",
            "Busca as ordens de serviço pendentes na Mobyan e prepara tudo pra envio aos prestadores.",
            self.gerar_pendencias,
            COR_AZUL,
            COR_AZUL_HOVER,
            largura=LARGURA_CARD_CHEIO,
            altura=ALTURA_CARD_CHEIO,
        ))
        b1.pack(fill="x")

        self.criar_atalhos(
            pagina,
            [
                ("Abrir\nPendências", self.abrir_planilha),
                ("Atualizar\nResumo", self.atualizar_resumos),
            ],
        ).pack(anchor="w", pady=(8, 0))

        self.adicionar_painel_info(
            pagina,
            "Fluxo operacional",
            "1. Gere as pendências.  2. Vá em WhatsApp pra revisar e enviar a cobrança do dia.",
        )

    def criar_pagina_whatsapp(self):
        pagina = self.nova_pagina("whatsapp")
        acoes = tk.Frame(pagina, bg=COR_FUNDO_2)
        acoes.pack(fill="x", pady=(2, 14))
        b1 = self.registrar_action(ActionCard(
            acoes,
            "Enviar WhatsApp",
            "Envia a cobrança do dia pelos prestadores selecionados.",
            self.enviar_whatsapp,
            COR_VERDE,
            COR_VERDE_HOVER,
            largura=LARGURA_CARD_PAR,
            altura=ALTURA_CARD_PAR,
        ))
        b1.grid(row=0, column=0, padx=(0, 8), sticky="ew")
        b2 = ActionCard(
            acoes,
            "Configurar Contatos",
            "Cadastre o número de WhatsApp de cada base pra receber a cobrança.",
            self.abrir_configurar_whatsapp,
            COR_DOURADO,
            COR_DOURADO_HOVER,
            largura=LARGURA_CARD_PAR,
            altura=ALTURA_CARD_PAR,
        )
        b2.grid(row=0, column=1, padx=(8, 0), sticky="ew")
        acoes.grid_columnconfigure(0, weight=1)
        acoes.grid_columnconfigure(1, weight=1)

        self.criar_atalhos(
            pagina,
            [
                ("Abrir\nAcompanhamento", self.abrir_acompanhamento),
                ("Abrir\nImagens", self.abrir_pasta_imagens),
            ],
        ).pack(anchor="w", pady=(4, 0))

        self.adicionar_painel_info(
            pagina,
            "Fluxo de envio",
            "1. Cadastre o WhatsApp de cada base em Configurar Contatos (uma vez só).  2. Revise Acompanhamento e Envios.  3. Marque só as bases prontas.  4. Envie pelo WhatsApp.",
        )

    def criar_pagina_roteiros(self):
        pagina = self.nova_pagina("roteiros")
        acoes = tk.Frame(pagina, bg=COR_FUNDO_2)
        acoes.pack(fill="x", pady=(2, 14))
        b1 = self.registrar_action(ActionCard(
            acoes,
            "Gerar Roteirização",
            "Organiza as ordens de serviço da Mobyan e OGEA em rotas por técnico.",
            self.gerar_roteirizacao,
            COR_DOURADO,
            COR_DOURADO_HOVER,
            largura=LARGURA_CARD_PAR,
            altura=ALTURA_CARD_PAR,
        ))
        b1.grid(row=0, column=0, padx=(0, 8), sticky="ew")
        b2 = self.registrar_action(ActionCard(
            acoes,
            "Gerar Roteiros PDF",
            "Monta o roteiro em PDF de cada técnico, pronto pra imprimir.",
            self.gerar_pdfs,
            COR_LARANJA,
            COR_LARANJA_HOVER,
            largura=LARGURA_CARD_PAR,
            altura=ALTURA_CARD_PAR,
        ))
        b2.grid(row=0, column=1, padx=(8, 0), sticky="ew")
        acoes.grid_columnconfigure(0, weight=1)
        acoes.grid_columnconfigure(1, weight=1)

        self.criar_atalhos(
            pagina,
            [
                ("Abrir\nRoteirização", self.abrir_roteirizacao),
                ("Abrir\nRoteiros PDF", self.abrir_pasta_pdfs),
                ("Abrir base\nde regras", self.abrir_base_rotas),
            ],
        ).pack(anchor="w", pady=(0, 18))

        tk.Label(
            pagina,
            text="Gestão de Rotas",
            bg=COR_FUNDO_2,
            fg=COR_BRANCO,
            font=("Arial", 13, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 8))
        gestao_rotas_linha = tk.Frame(pagina, bg=COR_FUNDO_2)
        gestao_rotas_linha.pack(fill="x")
        b3 = ActionCard(
            gestao_rotas_linha,
            "Abrir Gestão Inteligente de Rotas",
            "Ajuste rotas, técnicos e regras de roteirização num só lugar.",
            self.abrir_gestao_rotas,
            COR_DOURADO,
            COR_DOURADO_HOVER,
            largura=LARGURA_CARD_PAR,
            altura=ALTURA_CARD_PAR,
        )
        b3.grid(row=0, column=0, padx=(0, 8), sticky="ew")
        b4 = ActionCard(
            gestao_rotas_linha,
            "Gerenciar Técnicos",
            "Cadastre, edite ou desative técnicos pra eles já aparecerem prontos nas regras e pendências.",
            self.abrir_gestao_tecnicos,
            COR_LARANJA,
            COR_LARANJA_HOVER,
            largura=LARGURA_CARD_PAR,
            altura=ALTURA_CARD_PAR,
        )
        b4.grid(row=0, column=1, padx=(8, 0), sticky="ew")
        gestao_rotas_linha.grid_columnconfigure(0, weight=1)
        gestao_rotas_linha.grid_columnconfigure(1, weight=1)

    def criar_pagina_ferramentas(self):
        pagina = self.nova_pagina("ferramentas")

        b = self.registrar_action(ActionCard(
            pagina,
            "Selecionar relatório e analisar",
            "Analisa o relatório da OGEA e aponta quais OSs têm direito a abono.",
            self.analisar_abonos,
            COR_VERDE,
            COR_VERDE_HOVER,
            largura=LARGURA_CARD_CHEIO,
            altura=ALTURA_CARD_CHEIO,
        ))
        b.pack(fill="x")

        self.criar_atalhos(
            pagina,
            [
                ("Abrir última\nanálise", self.abrir_analise_abonos),
                ("Abrir pasta\nde abonos", self.abrir_pasta_abonos),
            ],
            altura=58,
        ).pack(anchor="w", pady=(8, 0))

        self.adicionar_painel_info(
            pagina,
            "Como o analista decide",
            "• Abre sábado de forma explícita: NÃO ABONAR.  • Não abre sábado ou funciona apenas de segunda a sexta: ABONAR.\n• Nenhuma informação sobre sábado: ABONAR, seguindo a regra operacional informada.  • Informações conflitantes: REVISAR MANUALMENTE.\nPesquisa todos os campos: bairro, endereço, observação, cliente e serviço.",
            altura=112,
        )

        tk.Frame(pagina, bg=COR_BORDA, height=1).pack(fill="x", pady=(12, 8))

        tk.Label(
            pagina,
            text="Sistema",
            bg=COR_FUNDO_2,
            fg=COR_BRANCO,
            font=("Arial", 13, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 4))
        tk.Label(
            pagina,
            text=f"Pasta principal: {BASE_DIR}",
            bg=COR_FUNDO_2,
            fg=COR_TEXTO_FRACO,
            font=("Arial", 9),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))
        self.criar_atalhos(
            pagina,
            [
                ("Abrir\nProjeto", self.abrir_pasta_projeto),
                ("Abrir\nLogs", self.abrir_pasta_logs),
                ("Abrir\nPDFs", self.abrir_pasta_pdfs),
                ("Atualizar\nResumos", self.atualizar_resumos),
            ],
            altura=58,
        ).pack(anchor="w")

    def adicionar_painel_info(self, master, titulo, texto, altura=110):
        painel = tk.Frame(
            master,
            bg=COR_CARD,
            height=altura,
            highlightbackground=COR_BORDA,
            highlightthickness=1,
        )
        painel.pack(fill="x", pady=(16, 0))
        painel.pack_propagate(False)
        tk.Label(
            painel,
            text=titulo,
            bg=COR_CARD,
            fg=COR_DOURADO,
            font=("Arial", 12, "bold"),
            anchor="w",
        ).pack(fill="x", padx=16, pady=(12, 4))
        tk.Label(
            painel,
            text=texto,
            bg=COR_CARD,
            fg=COR_TEXTO_SECUNDARIO,
            font=("Arial", 10),
            anchor="nw",
            justify="left",
            wraplength=980,
        ).pack(fill="both", expand=True, padx=16, pady=(0, 12))
        return painel

    def mostrar_pagina(self, chave):
        titulos = {
            "inicio": ("Painel", "Panorama das OSs vencendo hoje, desconsiderando envios de bobina."),
            "operacao": ("Operação", "Geração e revisão das pendências operacionais."),
            "whatsapp": ("WhatsApp", "Envio das pendências e cadastro dos contatos de cada base."),
            "roteiros": ("Roteiros", "Roteirização, ajuste de rotas e geração dos PDFs por técnico."),
            "ferramentas": ("Abonos", "Analista de abonos OGEA e acesso a arquivos e logs do sistema."),
        }
        for nome, pagina in self.paginas.items():
            if nome == chave:
                pagina.lift()
        for nome, botao in self.nav_botoes.items():
            botao.selecionar(nome == chave)
        titulo, subtitulo = titulos[chave]
        self.page_title.config(text=titulo)
        self.page_subtitle.config(text=subtitulo)
        if chave == "inicio":
            self.atualizar_metricas_dia()
        elif chave == "roteiros":
            self.atualizar_resumo_rotas()

    def escrever_log(self, texto):
        self.log_text.insert("end", texto)
        self.log_text.see("end")
        self.root.update_idletasks()

    def definir_status(self, texto, cor=COR_DOURADO):
        self.status.config(text=f"Status: {texto}", fg=cor)

    def notificar(self, titulo, mensagem):
        try:
            if os.name == "nt":
                import winsound
                winsound.MessageBeep(winsound.MB_ICONASTERISK)
            elif sys.platform == "darwin":
                mensagem_segura = str(mensagem).replace('"', "'")
                titulo_seguro = str(titulo).replace('"', "'")
                subprocess.run([
                    "osascript",
                    "-e",
                    f'display notification "{mensagem_segura}" with title "{titulo_seguro}"',
                ], check=False)
            else:
                subprocess.run(["notify-send", str(titulo), str(mensagem)], check=False)
        except Exception:
            try:
                self.root.bell()
            except Exception:
                pass

    def travar_botoes(self):
        self.processo_rodando = True
        for botao in self.botoes_processo:
            botao.set_enabled(False)

    def liberar_botoes(self):
        self.processo_rodando = False
        for botao in self.botoes_processo:
            botao.set_enabled(True)

    def ler_envios_para_resumo(self):
        vazio = {
            "existe": False,
            "prontos": 0,
            "sem_retorno": 0,
            "em_risco": 0,
            "total": 0,
        }
        if not ARQUIVO_PLANILHA.exists():
            return vazio
        try:
            envios = pd.read_excel(ARQUIVO_PLANILHA, sheet_name="Envios", dtype=str, keep_default_na=False, na_filter=False)
            pendencias = pd.read_excel(ARQUIVO_PLANILHA, sheet_name="Pendências", dtype=str, keep_default_na=False, na_filter=False)
            try:
                acompanhamento = pd.read_excel(ARQUIVO_PLANILHA, sheet_name="Acompanhamento", dtype=str, keep_default_na=False, na_filter=False)
            except Exception:
                acompanhamento = pd.DataFrame()

            for coluna in envios.columns:
                envios[coluna] = envios[coluna].astype(str).str.strip()
            if "Enviar" not in envios.columns:
                envios["Enviar"] = ""
            if "Status Envio" not in envios.columns:
                envios["Status Envio"] = ""
            enviar_norm = envios["Enviar"].astype(str).str.strip().str.lower()
            prontos = len(envios[(envios["Status Envio"] == "Pronto para envio") & enviar_norm.isin(["sim", "s", "yes", "y"])])
            sem_retorno = 0
            em_risco = 0
            if not acompanhamento.empty:
                for coluna in acompanhamento.columns:
                    acompanhamento[coluna] = acompanhamento[coluna].astype(str).str.strip()
                if "Status Operacional" in acompanhamento.columns:
                    sem_retorno = len(acompanhamento[acompanhamento["Status Operacional"] == "Sem retorno"])
                if "Risco" in acompanhamento.columns:
                    riscos = acompanhamento["Risco"].astype(str).str.strip().str.upper()
                    em_risco = len(acompanhamento[riscos.isin(["CRÍTICO", "CRITICO", "ALTO RISCO", "ABONAR"])])
            return {
                "existe": True,
                "prontos": prontos,
                "sem_retorno": sem_retorno,
                "em_risco": em_risco,
                "total": len(pendencias),
            }
        except Exception as erro:
            self.escrever_log(f"\nAviso: não consegui atualizar o resumo: {erro}\n")
            return vazio

    def atualizar_resumo_rotas(self):
        resumo = {"sem_rota": 0, "conflitos": 0, "regras": 0, "aliases": 0}
        if obter_resumo_rotas is not None:
            try:
                resumo.update(obter_resumo_rotas())
            except Exception as erro:
                self.escrever_log(f"\nAviso: não consegui atualizar o resumo de rotas: {erro}\n")
        if hasattr(self, "card_rotas_sem_rota"):
            self.card_rotas_sem_rota.atualizar(resumo["sem_rota"], COR_VERMELHO)
            self.card_rotas_conflitos.atualizar(resumo["conflitos"], COR_LARANJA)
            self.card_rotas_regras.atualizar(resumo["regras"], COR_DOURADO)
            self.card_rotas_aliases.atualizar(resumo["aliases"], COR_AZUL)
        return resumo

    def atualizar_resumos(self):
        self.atualizar_resumo_rotas()
        self.atualizar_metricas_dia()
        self.definir_status("resumos atualizados", COR_TEXTO_SECUNDARIO)

    def calcular_metricas_painel(self):
        vazio = {
            "existe": False,
            "vencendo_hoje": 0,
            "criticas": 0,
            "prestador_top": None,
            "prestador_top_qtd": 0,
        }
        if not ARQUIVO_PLANILHA.exists():
            return vazio
        try:
            df = pd.read_excel(
                ARQUIVO_PLANILHA, sheet_name="Pendências", dtype=str,
                keep_default_na=False, na_filter=False,
            )
        except Exception as erro:
            self.escrever_log(f"\nAviso: não consegui atualizar o painel do dia: {erro}\n")
            return vazio
        if "Serviço" in df.columns:
            df = df[~df["Serviço"].astype(str).str.upper().str.contains("BOBINA", na=False)]
        if "Data Limite" not in df.columns:
            return {**vazio, "existe": True}

        datas = pd.to_datetime(df["Data Limite"], format="%d/%m/%Y", errors="coerce").dt.date
        hoje = date.today()
        vencendo_hoje_mask = datas == hoje
        criticas_mask = datas.apply(lambda d: d is not None and not pd.isna(d) and (hoje - d).days >= 5)

        prestador_top = None
        prestador_top_qtd = 0
        if "Prestador" in df.columns:
            prestadores = df.loc[vencendo_hoje_mask, "Prestador"].astype(str).str.strip()
            prestadores = prestadores[prestadores != ""]
            if not prestadores.empty:
                contagem = prestadores.value_counts()
                prestador_top = contagem.idxmax()
                prestador_top_qtd = int(contagem.max())

        return {
            "existe": True,
            "vencendo_hoje": int(vencendo_hoje_mask.sum()),
            "criticas": int(criticas_mask.sum()),
            "prestador_top": prestador_top,
            "prestador_top_qtd": prestador_top_qtd,
        }

    def atualizar_metricas_dia(self):
        if not hasattr(self, "card_vencendo_hoje"):
            return
        dados = self.calcular_metricas_painel()
        self.ultimo_prestador_top = dados["prestador_top"]
        self.card_vencendo_hoje.atualizar(dados["vencendo_hoje"], COR_LARANJA)
        self.card_criticas.atualizar(dados["criticas"], COR_VERMELHO)
        if dados["prestador_top"]:
            self.card_prestador_top.atualizar(
                dados["prestador_top"], COR_AZUL,
                subtitulo=f"{dados['prestador_top_qtd']} OS(s) vencendo hoje",
            )
        else:
            self.card_prestador_top.atualizar("—", COR_AZUL, subtitulo="")

    def rodar_script(self, script_path, nome_acao, mensagem_sucesso, env_extra=None, argumentos=None):
        if self.processo_rodando:
            messagebox.showwarning("Processo em andamento", "Já existe uma automação rodando. Aguarde finalizar.")
            return
        if not FROZEN and not script_path.exists():
            messagebox.showerror("Arquivo não encontrado", f"Não encontrei o script:\n{script_path}")
            return
        self.acao_atual = nome_acao
        self.travar_botoes()
        self.definir_status(f"{nome_acao} em andamento...", COR_DOURADO)
        self.log_text.delete("1.0", "end")
        self.escrever_log(f"Iniciando: {nome_acao}\n")
        self.escrever_log("=" * 70 + "\n\n")
        thread = threading.Thread(
            target=self.executar_script_thread,
            args=(script_path, nome_acao, mensagem_sucesso, env_extra, argumentos or []),
            daemon=True,
        )
        thread.start()

    def executar_script_thread(self, script_path, nome_acao, mensagem_sucesso, env_extra=None, argumentos=None):
        try:
            env = os.environ.copy()
            if env_extra:
                env.update(env_extra)
            env.setdefault("PYTHONUTF8", "1")
            env.setdefault("PYTHONIOENCODING", "utf-8")
            kwargs_processo = {}
            if os.name == "nt":
                kwargs_processo["creationflags"] = subprocess.CREATE_NO_WINDOW
            if FROZEN:
                comando = [
                    str(PYTHON_AUTOMACOES), "--rodar-automacao", script_path.stem,
                ] + list(argumentos or [])
            else:
                comando = [str(PYTHON_AUTOMACOES), str(script_path)] + list(argumentos or [])
            processo = subprocess.Popen(
                comando,
                cwd=str(BASE_DIR),
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                bufsize=1,
                **kwargs_processo,
            )
            for linha in processo.stdout:
                self.log_queue.put(("log", linha))
            processo.wait()
            if processo.returncode == 0:
                self.log_queue.put(("sucesso", mensagem_sucesso))
            else:
                self.log_queue.put(("erro", f"{nome_acao} terminou com erro. Verifique o log acima."))
        except Exception as erro:
            self.log_queue.put(("erro", f"Erro ao executar {nome_acao}: {erro}"))

    def processar_logs(self):
        try:
            while True:
                tipo, conteudo = self.log_queue.get_nowait()
                if tipo == "log":
                    self.escrever_log(conteudo)
                elif tipo == "sucesso":
                    acao_concluida = self.acao_atual
                    self.acao_atual = None
                    self.escrever_log("\n" + "=" * 70 + "\n")
                    self.escrever_log(conteudo + "\n")
                    self.definir_status(conteudo, COR_VERDE)
                    self.notificar("ViscondeApp", conteudo)
                    self.liberar_botoes()
                    self.atualizar_resumos()
                    if acao_concluida == "Gerar Roteirização":
                        self.root.after(450, self.finalizar_roteirizacao_pos_execucao)
                elif tipo == "erro":
                    self.acao_atual = None
                    self.escrever_log("\n" + "=" * 70 + "\n")
                    self.escrever_log(conteudo + "\n")
                    self.definir_status("erro na automação", COR_VERMELHO)
                    self.notificar("ViscondeApp", "A automação terminou com erro.")
                    self.liberar_botoes()
                    self.atualizar_resumos()
        except queue.Empty:
            pass
        self.root.after(200, self.processar_logs)

    def finalizar_roteirizacao_pos_execucao(self):
        """Abre a Gestão somente se o roteiro recém-gerado tiver pendências.

        Quando não há pendências, mantém o comportamento anterior e abre a
        planilha final. Assim o Excel nunca bloqueia o arquivo durante a etapa
        de correção e reprocessamento das rotas.
        """
        abriu_gestao = self.verificar_pendencias_rotas_automaticamente()
        if not abriu_gestao and ARQUIVO_ROTEIRIZACAO.exists():
            try:
                abrir_caminho(ARQUIVO_ROTEIRIZACAO)
            except Exception as erro:
                self.escrever_log(f"\nAviso: não consegui abrir a roteirização: {erro}\n")

    def obter_pendencias_rotas_atuais(self):
        resumo = {"sem_rota": 0, "conflitos": 0, "regras": 0, "aliases": 0}
        if obter_resumo_rotas is not None:
            try:
                resumo.update(obter_resumo_rotas())
            except Exception as erro:
                self.escrever_log(f"\nAviso: não consegui verificar pendências de rotas: {erro}\n")
        return resumo

    def verificar_pendencias_rotas_automaticamente(self):
        """Verifica o roteiro recém-gerado e abre a Gestão se houver anomalias."""
        if not ARQUIVO_ROTEIRIZACAO.exists() or abrir_gestao_rotas is None:
            return False
        resumo = self.obter_pendencias_rotas_atuais()
        sem_rota = int(resumo.get("sem_rota", 0) or 0)
        conflitos = int(resumo.get("conflitos", 0) or 0)
        total = sem_rota + conflitos
        self.atualizar_resumo_rotas()
        if total <= 0:
            return False

        self.mostrar_pagina("roteiros")
        self.definir_status(f"{total} OS(s) aguardando ajuste de rota", COR_LARANJA)
        self.escrever_log(
            f"\nATENÇÃO: a roteirização possui {sem_rota} OS(s) sem rota e "
            f"{conflitos} conflito(s). Abrindo a Gestão Inteligente de Rotas.\n"
        )
        self.notificar(
            "Rotas precisam de atenção",
            f"Existem {total} OS(s) que precisam ser resolvidas antes dos PDFs.",
        )
        self.abrir_gestao_rotas()
        return True

    def exigir_rotas_resolvidas(self, acao):
        """Impede etapas dependentes do roteiro enquanto existirem anomalias."""
        if not ARQUIVO_ROTEIRIZACAO.exists():
            return True
        resumo = self.obter_pendencias_rotas_atuais()
        sem_rota = int(resumo.get("sem_rota", 0) or 0)
        conflitos = int(resumo.get("conflitos", 0) or 0)
        total = sem_rota + conflitos
        if total <= 0:
            return True

        messagebox.showwarning(
            "Rotas pendentes",
            f"Não é seguro executar '{acao}' ainda.\n\n"
            f"Sem rota: {sem_rota}\n"
            f"Conflitos: {conflitos}\n\n"
            "A Gestão Inteligente de Rotas será aberta para você resolver os casos. "
            "Depois de gerar novamente o roteiro sem pendências, tente esta etapa outra vez.",
            parent=self.root,
        )
        self.mostrar_pagina("roteiros")
        self.abrir_gestao_rotas()
        return False

    def gerar_pendencias(self):
        self.rodar_script(SCRIPT_GERAR_PENDENCIAS, "Gerar Pendências", "Planilha de pendências criada com sucesso.")

    def gerar_roteirizacao(self):
        # A planilha permanece fechada durante a validação das rotas. Se não
        # houver pendências, ela será aberta ao final por esta própria Central.
        self.rodar_script(
            SCRIPT_GERAR_ROTEIRIZACAO,
            "Gerar Roteirização",
            "Roteirização criada com sucesso.",
            argumentos=["--sem-abrir"],
        )

    def gerar_pdfs(self):
        if not ARQUIVO_ROTEIRIZACAO.exists():
            messagebox.showwarning("Roteirização não encontrada", "Gere a roteirização antes de criar os PDFs.")
            return
        if not self.exigir_rotas_resolvidas("Gerar Roteiros PDF"):
            return
        self.rodar_script(SCRIPT_GERAR_PDFS, "Gerar Roteiros PDF", "Roteiros unificados criados com sucesso.")

    def gerar_os_visconde_teste(self):
        if not ARQUIVO_ROTEIRIZACAO.exists():
            messagebox.showwarning(
                "Roteirização não encontrada",
                "Gere a roteirização oficial antes de criar as OSs piloto da Visconde.",
            )
            return
        if not self.exigir_rotas_resolvidas("Gerar OS Visconde - TESTE"):
            return
        confirmar = messagebox.askyesno(
            "OS Visconde - Modo TESTE",
            "Esta função cria arquivos somente na pasta piloto e NÃO altera a roteirização nem os PDFs oficiais.\n\nDeseja continuar?",
        )
        if not confirmar:
            return
        self.rodar_script(
            SCRIPT_GERAR_OS_VISCONDE_TESTE,
            "Gerar OS Visconde - TESTE",
            "OSs Visconde de teste criadas com sucesso. O fluxo oficial não foi alterado.",
        )

    def analisar_abonos(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o relatório da OGEA",
            filetypes=[
                ("Relatórios OGEA", "*.csv *.txt *.xlsx *.xlsm *.xls"),
                ("Todos os arquivos", "*.*"),
            ],
        )
        if not arquivo:
            return
        self.rodar_script(
            SCRIPT_ANALISAR_ABONOS,
            "Analisar Abonos OGEA",
            "Análise de abonos criada com sucesso.",
            argumentos=["--arquivo", arquivo],
        )

    def escolher_modelo_mensagem(self):
        janela = tk.Toplevel(self.root)
        janela.title("Modelo de mensagem")
        janela.geometry("580x390")
        janela.resizable(False, False)
        janela.configure(bg=COR_FUNDO)
        janela.transient(self.root)
        janela.grab_set()
        escolha = {"valor": None}
        janela.update_idletasks()
        largura = 580
        altura = 390
        x = (janela.winfo_screenwidth() // 2) - (largura // 2)
        y = (janela.winfo_screenheight() // 2) - (altura // 2)
        janela.geometry(f"{largura}x{altura}+{x}+{y}")
        container = tk.Frame(janela, bg=COR_FUNDO)
        container.pack(fill="both", expand=True, padx=28, pady=24)
        tk.Label(container, text="Qual mensagem deseja enviar?", font=("Arial", 19, "bold"), bg=COR_FUNDO, fg=COR_DOURADO).pack(pady=(4, 6))
        tk.Label(container, text="Escolha o modelo que será usado no envio pelo WhatsApp.", font=("Arial", 11), bg=COR_FUNDO, fg=COR_TEXTO_SECUNDARIO).pack(pady=(0, 20))

        def selecionar(valor):
            escolha["valor"] = valor
            janela.destroy()

        ActionCard(container, "Mensagem da Manhã", "Primeiro envio do dia com cobrança das vencidas.", lambda: selecionar("manha"), COR_AZUL, COR_AZUL_HOVER, largura=510, altura=82).pack(fill="x", pady=6)
        ActionCard(container, "Acompanhamento", "Atualização das pendências ao longo do dia.", lambda: selecionar("acompanhamento"), COR_VERDE, COR_VERDE_HOVER, largura=510, altura=82).pack(fill="x", pady=6)
        SmallButton(container, "Cancelar", lambda: selecionar(None), largura=510).pack(fill="x", pady=(8, 0))
        janela.wait_window()
        return escolha["valor"]

    def enviar_whatsapp(self):
        dados = self.ler_envios_para_resumo()
        prontos = dados["prontos"]
        if prontos <= 0:
            messagebox.showwarning("Nenhum envio pronto", "Não há nenhuma base com Enviar = Sim e Status Envio = Pronto para envio.")
            self.atualizar_resumos()
            return
        modelo = self.escolher_modelo_mensagem()
        if modelo is None:
            return
        nome_modelo = "Mensagem da Manhã" if modelo == "manha" else "Mensagem de Acompanhamento"
        resposta = messagebox.askyesno(
            "Confirmar envio",
            f"Modelo selecionado: {nome_modelo}\n\nSerão enviados {prontos} prestador(es) por WhatsApp.\n\nDeseja continuar?",
        )
        if not resposta:
            return
        self.rodar_script(
            SCRIPT_ENVIAR_WHATSAPP,
            "Enviar Pendências WhatsApp",
            "Envio das pendências para as bases concluído.",
            env_extra={"MODELO_MENSAGEM": modelo},
        )

    def fechar_gestao_rotas(self):
        instancia = self.janela_gestao_rotas
        self.janela_gestao_rotas = None
        try:
            if instancia is not None and instancia.win.winfo_exists():
                instancia.win.destroy()
        except Exception:
            pass

    def abrir_gestao_rotas(self, aba_inicial=None):
        if abrir_gestao_rotas is None:
            messagebox.showerror(
                "Gestão de Rotas",
                "O módulo de Gestão de Rotas não foi encontrado. Reaplique a atualização da Central.",
            )
            return
        try:
            if (
                self.janela_gestao_rotas is not None
                and self.janela_gestao_rotas.win.winfo_exists()
            ):
                self.janela_gestao_rotas.recarregar_tudo()
                if aba_inicial:
                    self.janela_gestao_rotas.barra_abas.selecionar(aba_inicial)
                self.janela_gestao_rotas.win.deiconify()
                self.janela_gestao_rotas.win.lift()
                self.janela_gestao_rotas.win.focus_force()
                return

            self.janela_gestao_rotas = abrir_gestao_rotas(
                self.root, on_change=self.atualizar_resumo_rotas, aba_inicial=aba_inicial
            )
            self.janela_gestao_rotas.win.protocol("WM_DELETE_WINDOW", self.fechar_gestao_rotas)
            self.janela_gestao_rotas.win.lift()
            self.janela_gestao_rotas.win.focus_force()
        except Exception as erro:
            self.janela_gestao_rotas = None
            messagebox.showerror("Gestão de Rotas", f"Não consegui abrir o módulo:\n\n{erro}")

    def abrir_gestao_tecnicos(self):
        self.abrir_gestao_rotas(aba_inicial="Técnicos")

    def abrir_configurar_whatsapp(self):
        if abrir_configurar_whatsapp is None:
            messagebox.showerror(
                "Configurar WhatsApp",
                "O módulo de configuração de WhatsApp não foi encontrado. Reaplique a atualização da Central.",
            )
            return
        try:
            if (
                self.janela_configurar_whatsapp is not None
                and self.janela_configurar_whatsapp.win.winfo_exists()
            ):
                self.janela_configurar_whatsapp.recarregar()
                self.janela_configurar_whatsapp.win.deiconify()
                self.janela_configurar_whatsapp.win.lift()
                self.janela_configurar_whatsapp.win.focus_force()
                return

            self.janela_configurar_whatsapp = abrir_configurar_whatsapp(self.root)
            self.janela_configurar_whatsapp.win.lift()
            self.janela_configurar_whatsapp.win.focus_force()
        except Exception as erro:
            self.janela_configurar_whatsapp = None
            messagebox.showerror("Configurar WhatsApp", f"Não consegui abrir o módulo:\n\n{erro}")

    def abrir_base_rotas(self):
        if ARQUIVO_REGRAS_ROTEIRIZACAO.exists():
            abrir_caminho(ARQUIVO_REGRAS_ROTEIRIZACAO)
        else:
            messagebox.showwarning("Base não encontrada", "A base regras_roteirizacao.xlsx não foi encontrada.")

    def abrir_planilha(self):
        if ARQUIVO_PLANILHA.exists():
            abrir_caminho(ARQUIVO_PLANILHA)
        else:
            messagebox.showwarning("Planilha não encontrada", "A planilha ainda não foi criada.")

    def montar_filtro_pendencias(self, modo, prestador=None):
        if modo == "prestador_top" and not prestador:
            return None, "Nenhum prestador com OS vencendo hoje ainda."
        if not ARQUIVO_PLANILHA.exists():
            return None, "A planilha de pendências ainda não foi gerada."
        try:
            df = pd.read_excel(
                ARQUIVO_PLANILHA, sheet_name="Pendências", dtype=str,
                keep_default_na=False, na_filter=False,
            )
        except Exception as erro:
            return None, f"Não consegui ler a planilha de pendências:\n{erro}"
        if "Serviço" in df.columns:
            df = df[~df["Serviço"].astype(str).str.upper().str.contains("BOBINA", na=False)]
        if "Data Limite" not in df.columns:
            return None, "A coluna 'Data Limite' não foi encontrada na planilha."

        datas = pd.to_datetime(df["Data Limite"], format="%d/%m/%Y", errors="coerce").dt.date
        hoje = date.today()

        if modo == "vencendo_hoje":
            filtro = datas == hoje
            titulo = "OSs vencendo hoje"
            nome_arquivo = "vencendo_hoje.xlsx"
        elif modo == "criticas":
            filtro = datas.apply(lambda d: d is not None and not pd.isna(d) and (hoje - d).days >= 5)
            titulo = "OSs críticas (5+ dias de atraso)"
            nome_arquivo = "criticas.xlsx"
        elif modo == "prestador_top":
            filtro = (datas == hoje) & (df["Prestador"].astype(str).str.strip() == str(prestador).strip())
            titulo = f"Vencendo hoje — {prestador}"
            nome_arquivo = "prestador_top.xlsx"
        else:
            return None, "Filtro desconhecido."

        resultado = df[filtro].reset_index(drop=True)
        if resultado.empty:
            return None, "Nenhuma OS encontrada para esse filtro."

        PASTA_FILTROS_PAINEL.mkdir(parents=True, exist_ok=True)
        caminho = PASTA_FILTROS_PAINEL / nome_arquivo
        with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
            resultado.to_excel(writer, index=False, sheet_name="Filtro")

        wb = openpyxl.load_workbook(caminho)
        ws = wb["Filtro"]
        total_colunas = len(resultado.columns)
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value=(
            f"{titulo} — {len(resultado)} OS(s) — gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ))
        ws.cell(row=1, column=1).font = Font(bold=True, color="F4C430", size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_colunas)
        linha_cabecalho = 2
        for coluna in range(1, total_colunas + 1):
            celula = ws.cell(row=linha_cabecalho, column=coluna)
            celula.font = Font(bold=True, color="FFFFFF")
            celula.fill = PatternFill("solid", fgColor="171717")
        ws.freeze_panes = ws.cell(row=linha_cabecalho + 1, column=1)
        ws.auto_filter.ref = (
            f"A{linha_cabecalho}:{get_column_letter(total_colunas)}{linha_cabecalho + len(resultado)}"
        )
        for coluna in range(1, total_colunas + 1):
            letra = get_column_letter(coluna)
            maior = max(
                [len(str(resultado.columns[coluna - 1]))]
                + [len(str(v)) for v in resultado.iloc[:, coluna - 1]]
            )
            ws.column_dimensions[letra].width = min(max(maior + 2, 10), 45)
        wb.save(caminho)
        return caminho, None

    def abrir_pendencias_filtradas(self, modo, prestador=None):
        caminho, erro = self.montar_filtro_pendencias(modo, prestador=prestador)
        if erro:
            messagebox.showinfo("Filtro de pendências", erro)
            return
        abrir_caminho(caminho)

    def abrir_roteirizacao(self):
        if ARQUIVO_ROTEIRIZACAO.exists():
            abrir_caminho(ARQUIVO_ROTEIRIZACAO)
        else:
            messagebox.showwarning("Roteirização não encontrada", "Gere a roteirização primeiro.")

    def abrir_analise_abonos(self):
        if ARQUIVO_ABONOS.exists():
            abrir_caminho(ARQUIVO_ABONOS)
        else:
            messagebox.showwarning("Análise não encontrada", "Nenhuma análise de abonos foi criada ainda.")

    def abrir_acompanhamento(self):
        self.abrir_planilha()

    def abrir_pasta_pdfs(self):
        PASTA_PDFS.mkdir(parents=True, exist_ok=True)
        abrir_caminho(PASTA_PDFS)

    def abrir_pasta_imagens(self):
        PASTA_IMAGENS.mkdir(parents=True, exist_ok=True)
        abrir_caminho(PASTA_IMAGENS)

    def abrir_pasta_os_visconde_teste(self):
        PASTA_OS_VISCONDE_TESTE.mkdir(parents=True, exist_ok=True)
        abrir_caminho(PASTA_OS_VISCONDE_TESTE)

    def abrir_pasta_abonos(self):
        PASTA_ABONOS.mkdir(parents=True, exist_ok=True)
        abrir_caminho(PASTA_ABONOS)

    def abrir_pasta_logs(self):
        PASTA_LOGS.mkdir(parents=True, exist_ok=True)
        abrir_caminho(PASTA_LOGS)

    def abrir_pasta_projeto(self):
        abrir_caminho(BASE_DIR)

    def sair_da_conta(self):
        if self.processo_rodando:
            messagebox.showwarning(
                "Sair da conta",
                "Espere a automação em andamento terminar antes de sair da conta.",
            )
            return

        if not messagebox.askyesno(
            "Sair da conta",
            "Isso encerra a sessão atual. Vai ser preciso fazer login de novo "
            "(pode ser em outra conta) pra continuar usando a Central. Deseja sair?",
        ):
            return

        limpar_sessao_local()
        self.root.destroy()
        main()


def despachar_automacao_frozen():
    """Num executável empacotado, sys.executable é o próprio app — não dá pra
    apontar um caminho de .py como fariam os scripts hoje via subprocess. Em vez
    disso, rodar_script/gestao_rotas relançam o próprio executável com
    "--rodar-automacao <modulo> ...args", e este ponto de entrada intercepta isso
    antes de abrir qualquer janela, despachando pro módulo certo via runpy —
    cada script mantém seu próprio bloco `if __name__ == "__main__":` inalterado.

    Retorna True se tratou uma chamada de despacho (o processo deve encerrar
    logo em seguida), False se é uma abertura normal da GUI."""
    if not FROZEN or len(sys.argv) < 3 or sys.argv[1] != "--rodar-automacao":
        return False

    import runpy

    nome_modulo = sys.argv[2]
    sys.argv = [nome_modulo] + sys.argv[3:]
    runpy.run_module(nome_modulo, run_name="__main__")
    return True


def chromium_instalado():
    """Confere se o Chromium do Playwright já está baixado. Num executável
    empacotado não existe mais um `playwright install chromium` rodado pelo
    instalador (esse passo some junto com a venv/pip) — então isso precisa ser
    verificado e resolvido na primeira abertura do app."""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as playwright:
            return Path(playwright.chromium.executable_path).exists()
    except Exception:
        return False


ARQUIVO_LOG_NAVEGADOR = BASE_DIR / "logs" / "instalacao_navegador.log"


def _registrar_log_navegador(linhas):
    try:
        ARQUIVO_LOG_NAVEGADOR.parent.mkdir(parents=True, exist_ok=True)
        with ARQUIVO_LOG_NAVEGADOR.open("a", encoding="utf-8") as arquivo:
            arquivo.write(f"\n{'=' * 70}\n{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            for linha in linhas:
                arquivo.write(str(linha) + "\n")
    except Exception:
        pass


def instalar_chromium_com_progresso(root, ao_concluir):
    """Baixa o Chromium do Playwright em segundo plano, mostrando um diálogo
    simples de progresso. Chama ao_concluir(erro) na thread principal do Tk
    quando terminar (erro=None em caso de sucesso).

    Tudo que acontece aqui é registrado em ARQUIVO_LOG_NAVEGADOR — se isso
    falhar numa máquina que não temos acesso direto, esse arquivo é o que
    permite diagnosticar sem depender de descrever o erro por texto."""
    configurar_icone_janela(root)
    janela = tk.Toplevel(root)
    janela.title("ViscondeApp")
    janela.configure(bg=COR_FUNDO)
    janela.geometry("440x150")
    janela.resizable(False, False)
    janela.protocol("WM_DELETE_WINDOW", lambda: None)
    janela.transient(root)
    janela.grab_set()

    tk.Label(
        janela,
        text="Preparando o navegador da automação (só na primeira vez)...",
        bg=COR_FUNDO,
        fg=COR_BRANCO,
        font=("Arial", 11),
        wraplength=400,
        justify="left",
    ).pack(pady=(26, 10), padx=20)

    barra = ttk.Progressbar(janela, mode="indeterminate")
    barra.pack(fill="x", padx=20, pady=8)
    barra.start(12)

    def worker():
        erro = None
        log = [
            "Iniciando instalação do Chromium...",
            f"PLAYWRIGHT_BROWSERS_PATH={os.environ.get('PLAYWRIGHT_BROWSERS_PATH', '(não definido)')}",
        ]
        try:
            from playwright._impl._driver import compute_driver_executable, get_driver_env
            driver_executable, driver_cli = compute_driver_executable()
            log.append(f"driver_executable={driver_executable} (existe={Path(driver_executable).exists()})")
            log.append(f"driver_cli={driver_cli} (existe={Path(driver_cli).exists()})")

            kwargs_processo = {}
            if os.name == "nt":
                kwargs_processo["creationflags"] = subprocess.CREATE_NO_WINDOW

            # --force evita que uma tentativa anterior parcial/corrompida (ex:
            # interrompida por antivírus) faça o instalador achar que já está
            # tudo pronto e não baixar nada. --no-shell pula o Chrome Headless
            # Shell, que não usamos (a automação sempre roda com navegador
            # visível, headless=False).
            resultado = subprocess.run(
                [str(driver_executable), driver_cli, "install", "--force", "--no-shell", "chromium"],
                env=get_driver_env(),
                capture_output=True,
                text=True,
                errors="replace",
                **kwargs_processo,
            )
            log.append(f"returncode={resultado.returncode}")
            if resultado.stdout:
                log.append("--- stdout ---\n" + resultado.stdout)
            if resultado.stderr:
                log.append("--- stderr ---\n" + resultado.stderr)

            if resultado.returncode != 0:
                erro = "A instalação do navegador terminou com código de erro."
            elif not chromium_instalado():
                erro = "A instalação terminou sem erro, mas o navegador continua não encontrado."
                try:
                    from playwright.sync_api import sync_playwright
                    with sync_playwright() as playwright:
                        caminho_esperado = Path(playwright.chromium.executable_path)
                        log.append(f"executable_path esperado: {caminho_esperado}")
                        pasta_navegadores = caminho_esperado.parent.parent
                        if pasta_navegadores.exists():
                            log.append(f"Conteúdo de {pasta_navegadores}:")
                            log.extend(f"  - {item.name}" for item in pasta_navegadores.iterdir())
                        else:
                            log.append(f"Pasta {pasta_navegadores} nem existe.")
                except Exception as exc_diag:
                    log.append(f"Não consegui inspecionar a pasta de navegadores: {exc_diag!r}")
        except Exception as exc:
            erro = str(exc)
            log.append(f"EXCEÇÃO: {exc!r}")

        log.append(f"Resultado final: {'OK' if not erro else erro}")
        _registrar_log_navegador(log)

        def finalizar():
            janela.destroy()
            if erro:
                erro_completo = f"{erro}\n\nDetalhes em: {ARQUIVO_LOG_NAVEGADOR}"
            else:
                erro_completo = None
            ao_concluir(erro_completo)

        root.after(0, finalizar)

    threading.Thread(target=worker, daemon=True).start()


def avisar_atualizacao_disponivel(root, info):
    """Pergunta se o usuário quer baixar a atualização já publicada e, se sim,
    baixa em segundo plano e pede pra fechar o app e rodar o instalador —
    não se autossubstitui em silêncio (ver Fase 4 do plano)."""
    versao_nova = info.get("versao", "?")
    mensagem = f"Uma nova versão do ViscondeApp está disponível ({versao_nova})."
    if info.get("notas"):
        mensagem += f"\n\n{info['notas']}"
    mensagem += "\n\nDeseja baixar agora?"

    if not messagebox.askyesno("Atualização disponível", mensagem):
        return

    nome_arquivo = Path(info["url_download"]).name or f"central-visconde-{versao_nova}"
    destino = Path(tempfile.gettempdir()) / nome_arquivo

    def finalizar(erro):
        if erro:
            messagebox.showerror("ViscondeApp", f"Não consegui baixar a atualização:\n{erro}")
            return
        messagebox.showinfo(
            "Atualização baixada",
            f"A atualização foi baixada em:\n{destino}\n\n"
            "Feche o ViscondeApp e rode o instalador baixado pra concluir.",
        )
        try:
            abrir_caminho(destino.parent)
        except Exception:
            pass

    def worker():
        erro = None
        try:
            baixar_atualizacao(info["url_download"], destino)
        except Exception as exc:
            erro = str(exc)
        root.after(0, lambda: finalizar(erro))

    threading.Thread(target=worker, daemon=True).start()


def verificar_atualizacao_em_segundo_plano(root):
    """Consulta o backend por versão nova sem travar a abertura do app (só faz
    sentido em builds empacotados — rodando do código-fonte já é sempre a
    versão mais recente)."""
    def worker():
        try:
            info = verificar_atualizacao()
        except Exception:
            info = None
        if info:
            root.after(0, lambda: avisar_atualizacao_disponivel(root, info))

    threading.Thread(target=worker, daemon=True).start()


def main():
    if despachar_automacao_frozen():
        return

    configurar_identidade_windows()

    if not garantir_sessao_valida():
        return

    root = tk.Tk()

    def abrir_app_principal():
        CentralVisconde(root)
        if FROZEN:
            root.after(2000, verificar_atualizacao_em_segundo_plano, root)

    if FROZEN and not chromium_instalado():
        root.withdraw()

        def ao_concluir(erro):
            root.deiconify()
            if erro:
                messagebox.showerror(
                    "ViscondeApp",
                    f"Não consegui preparar o navegador da automação:\n{erro}\n\n"
                    "Feche e abra o app de novo para tentar mais uma vez.",
                )
            abrir_app_principal()

        instalar_chromium_com_progresso(root, ao_concluir)
    else:
        abrir_app_principal()

    root.mainloop()


if __name__ == "__main__":
    main()
